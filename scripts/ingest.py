"""Script CLI de ingesta: Excel/CSV → PostgreSQL.

Detecta el tipo de archivo por las hojas/headers que contiene y carga los datos
usando UPSERT (ON CONFLICT DO UPDATE).

Uso:
    uv run python scripts/ingest.py data/cardio_ventas.xlsx
    uv run python scripts/ingest.py --all data/
    uv run python scripts/ingest.py --dry-run data/ddd_data.xlsx
"""

from __future__ import annotations

import csv
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import typer
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

# Agregar el directorio backend al path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from app.database import Base
from app.models.cardio import (
    Agreement,
    AgreementDetail,
    Brand,
    BudgetEntry,
    ChannelDistribution,
    KpiBrand,
    KpiGlobal,
    MarketAteneo,
    MarketAteneoNational,
    MarketPerformance,
    MarketPerformanceRegional,
    Molecule,
    Presentation,
    Prescription,
    PrescriptionCompetitor,
    PrescriptionMarketShare,
    Price,
    PriceCatalog,
    Region,
    StockBrand,
    StockPresentation,
)
from app.models.ddd import (
    DddBrand,
    DddBrandMonthly,
    DddMarket,
    DddRegionSummary,
    DddTotalMonthly,
)

app = typer.Typer(help="Siegfried BI — Ingesta de datos Excel")

MONTH_NAMES = [
    "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
]


def get_session() -> Session:
    db_url = os.environ.get(
        "DATABASE_URL", "postgresql://siegfried:siegfried_dev@localhost:5432/siegfried_bi"
    )
    engine = create_engine(db_url)
    return sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)()


def get_or_create_brand(db: Session, name: str, **kwargs) -> Brand:
    brand = db.query(Brand).filter(Brand.name == name).first()
    if not brand:
        brand = Brand(name=name, **kwargs)
        db.add(brand)
        db.flush()
    return brand


def get_or_create_molecule(db: Session, name: str, clase: str | None = None) -> Molecule:
    mol = db.query(Molecule).filter(Molecule.name == name).first()
    if not mol:
        mol = Molecule(name=name, clase=clase)
        db.add(mol)
        db.flush()
    return mol


def get_or_create_region(db: Session, name: str) -> Region:
    reg = db.query(Region).filter(Region.name == name).first()
    if not reg:
        reg = Region(name=name)
        db.add(reg)
        db.flush()
    return reg


def get_or_create_presentation(
    db: Session, brand_id: int, name: str, familia: str | None = None
) -> Presentation:
    pres = (
        db.query(Presentation)
        .filter(Presentation.brand_id == brand_id, Presentation.name == name)
        .first()
    )
    if not pres:
        pres = Presentation(brand_id=brand_id, name=name, familia=familia)
        db.add(pres)
        db.flush()
    return pres


def detect_file_type_xlsx(wb) -> str:
    sheets = set(wb.sheetnames)
    if "Budget" in sheets or "Canales" in sheets or "KPIs" in sheets:
        return "cardio_ventas"
    if "Productos" in sheets or "Performance" in sheets:
        return "cardio_mercado"
    if "Recetas" in sheets or "Stock_Marca" in sheets:
        return "cardio_recetas_stock"
    if "Mercados" in sheets or "Datos_Mensuales" in sheets:
        return "ddd_data"

    # Detectar por nombre de hoja para archivos nuevos
    if "DS_AR_PM_FV_Standard" in sheets:
        return "ateneo_national"

    # Detectar por headers de la primera hoja
    ws = wb[wb.sheetnames[0]]
    headers = []
    for row in ws.iter_rows(max_row=2, values_only=True):
        headers.extend([str(h).strip() if h else "" for h in row])
        break
    header_set = set(headers)

    if "Registro" in header_set and "Troquel" in header_set and "Droga" in header_set:
        return "price_catalog"
    if "ObraSocial1" in header_set and "Aporte Neto" in header_set:
        return "convenios_detalle"
    if "RegionCUP" in header_set and "Droga" in header_set:
        return "mercado_regional"

    # Stock pivot: header de 2 filas, fila 0 = "Mes-Año" + fechas, fila 1 = Familia/Producto + métricas
    if "Mes-Año" in header_set or "Mes-Ano" in header_set:
        return "stock_pivot"

    # Verificar fila 2 para Familia/Producto
    row2_headers = []
    for i, row in enumerate(ws.iter_rows(max_row=2, values_only=True)):
        if i == 1:
            row2_headers = [str(h).strip() if h else "" for h in row]
    row2_set = set(row2_headers)
    if ("Familia" in row2_set or "Producto" in row2_set) and len(row2_headers) > 20:
        return "stock_pivot"

    raise ValueError(f"No se pudo detectar el tipo de archivo. Hojas: {sheets}, Headers: {headers[:10]}")


def read_sheet_rows(ws):
    """Lee una hoja como lista de dicts usando la primera fila como headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


def ingest_cardio_ventas(db: Session, wb, dry_run: bool = False) -> dict:
    stats = {"budget": 0, "canales": 0, "kpis_global": 0, "kpis_brand": 0}
    errors = []

    # Budget
    if "Budget" in wb.sheetnames:
        for row in read_sheet_rows(wb["Budget"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                year = int(row["Año"])
                for m in range(1, 13):
                    b_key = f"{MONTH_NAMES[m-1]}_Budget"
                    a_key = f"{MONTH_NAMES[m-1]}_Real"
                    budget_val = row.get(b_key)
                    actual_val = row.get(a_key)
                    if budget_val is None and actual_val is None:
                        continue
                    existing = (
                        db.query(BudgetEntry)
                        .filter_by(brand_id=brand.id, year=year, month=m)
                        .first()
                    )
                    if existing:
                        existing.budget = int(budget_val) if budget_val else None
                        existing.actual = int(actual_val) if actual_val else None
                    else:
                        db.add(BudgetEntry(
                            brand_id=brand.id, year=year, month=m,
                            budget=int(budget_val) if budget_val else None,
                            actual=int(actual_val) if actual_val else None,
                        ))
                    stats["budget"] += 1
            except Exception as e:
                errors.append(f"Budget: {row.get('Marca', '?')} - {e}")

    # Canales
    if "Canales" in wb.sheetnames:
        for row in read_sheet_rows(wb["Canales"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                existing = db.query(ChannelDistribution).filter_by(brand_id=brand.id).first()
                if existing:
                    existing.units = int(row["Unidades"])
                    existing.conversion_pct = float(row["Convenios_%"]) if row.get("Convenios_%") else None
                    existing.counter_pct = float(row["Mostrador_%"]) if row.get("Mostrador_%") else None
                else:
                    db.add(ChannelDistribution(
                        brand_id=brand.id, units=int(row["Unidades"]),
                        conversion_pct=float(row["Convenios_%"]) if row.get("Convenios_%") else None,
                        counter_pct=float(row["Mostrador_%"]) if row.get("Mostrador_%") else None,
                    ))
                stats["canales"] += 1
            except Exception as e:
                errors.append(f"Canales: {row.get('Marca', '?')} - {e}")

    # KPIs
    if "KPIs" in wb.sheetnames:
        for row in read_sheet_rows(wb["KPIs"]):
            try:
                if "KPI_Global" in str(row.get("Tipo", "")):
                    data = {k: v for k, v in row.items() if k != "Tipo" and v is not None}
                    db.add(KpiGlobal(data=data))
                    stats["kpis_global"] += 1
                else:
                    brand_name = row.get("Marca")
                    if brand_name:
                        brand = get_or_create_brand(db, brand_name, is_siegfried=True)
                        data = {k: v for k, v in row.items() if k not in ("Tipo", "Marca") and v is not None}
                        db.add(KpiBrand(brand_id=brand.id, data=data))
                        stats["kpis_brand"] += 1
            except Exception as e:
                errors.append(f"KPIs: {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_cardio_mercado(db: Session, wb, dry_run: bool = False) -> dict:
    stats = {"products": 0, "performance": 0}
    errors = []

    # Productos (catálogo)
    if "Productos" in wb.sheetnames:
        for row in read_sheet_rows(wb["Productos"]):
            try:
                mol = get_or_create_molecule(db, row["Molécula"])
                brand = get_or_create_brand(
                    db, row["Producto"],
                    molecule_id=mol.id,
                    manufacturer=row.get("Laboratorio"),
                    is_siegfried=bool(row.get("Es_Siegfried")),
                )
                stats["products"] += 1
            except Exception as e:
                errors.append(f"Productos: {row.get('Producto', '?')} - {e}")

    # Performance (formato largo)
    if "Performance" in wb.sheetnames:
        for row in read_sheet_rows(wb["Performance"]):
            try:
                mol = get_or_create_molecule(db, row["Molécula"])
                brand = get_or_create_brand(db, row["Producto"], molecule_id=mol.id)
                month_val = row["Mes"]
                if isinstance(month_val, str):
                    parts = month_val.split("-")
                    month_date = date(int(parts[0]), int(parts[1]), 1)
                else:
                    month_date = date(month_val.year, month_val.month, 1)

                existing = (
                    db.query(MarketPerformance)
                    .filter_by(brand_id=brand.id, molecule_id=mol.id, month=month_date)
                    .first()
                )
                if existing:
                    existing.units = int(row["Unidades"]) if row.get("Unidades") else None
                    existing.market_share = float(row["MS_%"]) if row.get("MS_%") else None
                else:
                    db.add(MarketPerformance(
                        brand_id=brand.id, molecule_id=mol.id, month=month_date,
                        units=int(row["Unidades"]) if row.get("Unidades") else None,
                        market_share=float(row["MS_%"]) if row.get("MS_%") else None,
                    ))
                stats["performance"] += 1
            except Exception as e:
                errors.append(f"Performance: {row.get('Producto', '?')} - {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_cardio_recetas_stock(db: Session, wb, dry_run: bool = False) -> dict:
    stats = {"recetas": 0, "recetas_ms": 0, "convenios": 0, "stock_marca": 0, "stock_pres": 0, "precios": 0}
    errors = []

    def parse_month(val) -> date:
        if isinstance(val, str):
            parts = val.split("-")
            return date(int(parts[0]), int(parts[1]), 1)
        return date(val.year, val.month, 1)

    # Recetas
    if "Recetas" in wb.sheetnames:
        for row in read_sheet_rows(wb["Recetas"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                month_date = parse_month(row["Mes"])
                rx_val = row.get("Recetas")
                md_val = row.get("Médicos")
                if rx_val is None or md_val is None:
                    continue
                existing = db.query(Prescription).filter_by(brand_id=brand.id, month=month_date).first()
                if existing:
                    existing.prescriptions = int(rx_val)
                    existing.physicians = int(md_val)
                else:
                    db.add(Prescription(
                        brand_id=brand.id, month=month_date,
                        prescriptions=int(rx_val), physicians=int(md_val),
                    ))
                stats["recetas"] += 1
            except Exception as e:
                errors.append(f"Recetas: {row.get('Marca', '?')} - {e}")

    # Recetas_MS
    if "Recetas_MS" in wb.sheetnames:
        for row in read_sheet_rows(wb["Recetas_MS"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                month_date = parse_month(row["Mes"])
                existing = db.query(PrescriptionMarketShare).filter_by(brand_id=brand.id, month=month_date).first()
                if existing:
                    existing.sie_prescriptions = int(row["Recetas_SIE"]) if row.get("Recetas_SIE") else None
                    existing.market_prescriptions = int(row["Recetas_Mercado"]) if row.get("Recetas_Mercado") else None
                    existing.market_share = float(row["MS_%"]) if row.get("MS_%") else None
                else:
                    db.add(PrescriptionMarketShare(
                        brand_id=brand.id, month=month_date,
                        sie_prescriptions=int(row["Recetas_SIE"]) if row.get("Recetas_SIE") else None,
                        market_prescriptions=int(row["Recetas_Mercado"]) if row.get("Recetas_Mercado") else None,
                        market_share=float(row["MS_%"]) if row.get("MS_%") else None,
                    ))
                stats["recetas_ms"] += 1
            except Exception as e:
                errors.append(f"Recetas_MS: {row.get('Marca', '?')} - {e}")

    # Convenios
    if "Convenios" in wb.sheetnames:
        for row in read_sheet_rows(wb["Convenios"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                existing = db.query(Agreement).filter_by(brand_id=brand.id, health_plan=row["Obra_Social"]).first()
                if existing:
                    existing.units_current = int(row["Unid_2025"]) if row.get("Unid_2025") else None
                    existing.units_previous = int(row["Unid_2024"]) if row.get("Unid_2024") else None
                    existing.delta_pct = int(row["Delta_%"]) if row.get("Delta_%") else None
                    existing.net_amount = float(row["Neto_$"]) if row.get("Neto_$") else None
                else:
                    db.add(Agreement(
                        brand_id=brand.id, health_plan=row["Obra_Social"],
                        units_current=int(row["Unid_2025"]) if row.get("Unid_2025") else None,
                        units_previous=int(row["Unid_2024"]) if row.get("Unid_2024") else None,
                        delta_pct=int(row["Delta_%"]) if row.get("Delta_%") else None,
                        net_amount=float(row["Neto_$"]) if row.get("Neto_$") else None,
                    ))
                stats["convenios"] += 1
            except Exception as e:
                errors.append(f"Convenios: {row.get('Marca', '?')} - {e}")

    # Stock_Marca
    if "Stock_Marca" in wb.sheetnames:
        for row in read_sheet_rows(wb["Stock_Marca"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                month_date = parse_month(row["Mes"])
                existing = db.query(StockBrand).filter_by(brand_id=brand.id, month=month_date).first()
                if existing:
                    existing.days_cover = int(row["Dias"]) if row.get("Dias") else None
                    existing.sales = int(row["Ventas"]) if row.get("Ventas") else None
                    existing.stock_units = int(row.get("Stock", 0)) if row.get("Stock") else None
                else:
                    db.add(StockBrand(
                        brand_id=brand.id, month=month_date,
                        days_cover=int(row["Dias"]) if row.get("Dias") else None,
                        sales=int(row["Ventas"]) if row.get("Ventas") else None,
                        stock_units=int(row.get("Stock", 0)) if row.get("Stock") else None,
                    ))
                stats["stock_marca"] += 1
            except Exception as e:
                errors.append(f"Stock_Marca: {row.get('Marca', '?')} - {e}")

    # Stock_Presentación
    if "Stock_Presentación" in wb.sheetnames:
        for row in read_sheet_rows(wb["Stock_Presentación"]):
            try:
                brand = get_or_create_brand(db, row["Marca"], is_siegfried=True)
                pres = get_or_create_presentation(
                    db, brand.id, row["Presentación"], familia=row.get("Familia")
                )
                month_date = parse_month(row["Mes"])
                existing = db.query(StockPresentation).filter_by(presentation_id=pres.id, month=month_date).first()
                if existing:
                    existing.sales = int(row["Ventas"]) if row.get("Ventas") else None
                    existing.days_cover = int(row["Dias"]) if row.get("Dias") else None
                    existing.status = row.get("Estado")
                else:
                    db.add(StockPresentation(
                        presentation_id=pres.id, month=month_date,
                        sales=int(row["Ventas"]) if row.get("Ventas") else None,
                        days_cover=int(row["Dias"]) if row.get("Dias") else None,
                        status=row.get("Estado"),
                    ))
                stats["stock_pres"] += 1
            except Exception as e:
                errors.append(f"Stock_Pres: {row.get('Presentación', '?')} - {e}")

    # Precios
    if "Precios" in wb.sheetnames:
        for row in read_sheet_rows(wb["Precios"]):
            try:
                brand = get_or_create_brand(db, row["Marca_Ref"], is_siegfried=True)
                existing = (
                    db.query(Price)
                    .filter_by(brand_id=brand.id, presentation=row["Presentación"], product_name=row["Producto"])
                    .first()
                )
                if existing:
                    existing.laboratory = row["Laboratorio"]
                    existing.pvp_previous = float(row["PVP_Ant"]) if row.get("PVP_Ant") else None
                    existing.pvp_current = float(row["PVP_Act"]) if row.get("PVP_Act") else None
                    existing.variation = float(row["Var_%"]) if row.get("Var_%") else None
                    existing.is_siegfried = bool(row.get("Es_SIE"))
                else:
                    db.add(Price(
                        brand_id=brand.id, presentation=row["Presentación"],
                        laboratory=row["Laboratorio"], product_name=row["Producto"],
                        pvp_previous=float(row["PVP_Ant"]) if row.get("PVP_Ant") else None,
                        pvp_current=float(row["PVP_Act"]) if row.get("PVP_Act") else None,
                        variation=float(row["Var_%"]) if row.get("Var_%") else None,
                        is_siegfried=bool(row.get("Es_SIE")),
                    ))
                stats["precios"] += 1
            except Exception as e:
                errors.append(f"Precios: {row.get('Producto', '?')} - {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_ddd_data(db: Session, wb, dry_run: bool = False) -> dict:
    stats = {"mercados": 0, "marcas": 0, "datos_mensuales": 0}
    errors = []

    market_cache: dict[str, DddMarket] = {}

    # Mercados
    if "Mercados" in wb.sheetnames:
        for row in read_sheet_rows(wb["Mercados"]):
            try:
                name = row["Mercado"]
                existing = db.query(DddMarket).filter_by(name=name).first()
                if existing:
                    existing.clase = row.get("Clase")
                    existing.total_units = int(row["Total_Unid"]) if row.get("Total_Unid") else None
                    existing.sie_units = int(row["Unid_SIE"]) if row.get("Unid_SIE") else None
                    existing.global_ms = float(row["MS_%"]) if row.get("MS_%") else None
                    market_cache[name] = existing
                else:
                    m = DddMarket(
                        name=name, clase=row.get("Clase"),
                        total_units=int(row["Total_Unid"]) if row.get("Total_Unid") else None,
                        sie_units=int(row["Unid_SIE"]) if row.get("Unid_SIE") else None,
                        global_ms=float(row["MS_%"]) if row.get("MS_%") else None,
                    )
                    db.add(m)
                    db.flush()
                    market_cache[name] = m
                stats["mercados"] += 1
            except Exception as e:
                errors.append(f"Mercados: {row.get('Mercado', '?')} - {e}")

    # Marcas
    brand_cache: dict[tuple[str, str], DddBrand] = {}
    if "Marcas" in wb.sheetnames:
        for row in read_sheet_rows(wb["Marcas"]):
            try:
                mkt_name = row["Mercado"]
                if mkt_name not in market_cache:
                    market_cache[mkt_name] = db.query(DddMarket).filter_by(name=mkt_name).one()
                market = market_cache[mkt_name]
                brand_name = row["Marca"]
                existing = db.query(DddBrand).filter_by(market_id=market.id, name=brand_name).first()
                if existing:
                    existing.is_siegfried = bool(row.get("Es_SIE"))
                    brand_cache[(mkt_name, brand_name)] = existing
                else:
                    b = DddBrand(
                        market_id=market.id, name=brand_name,
                        is_siegfried=bool(row.get("Es_SIE")),
                    )
                    db.add(b)
                    db.flush()
                    brand_cache[(mkt_name, brand_name)] = b
                stats["marcas"] += 1
            except Exception as e:
                errors.append(f"Marcas DDD: {row.get('Marca', '?')} - {e}")

    # Datos_Mensuales (formato ancho)
    if "Datos_Mensuales" in wb.sheetnames:
        for row in read_sheet_rows(wb["Datos_Mensuales"]):
            try:
                mkt_name = row["Mercado"]
                brand_name = row.get("Marca")
                region_name = row["Región"]
                region = get_or_create_region(db, region_name)

                if mkt_name not in market_cache:
                    market_cache[mkt_name] = db.query(DddMarket).filter_by(name=mkt_name).one()
                market = market_cache[mkt_name]

                for m in range(1, 13):
                    val = row.get(MONTH_NAMES[m - 1])
                    if val is None:
                        continue

                    if brand_name and brand_name != "__TOTAL__":
                        key = (mkt_name, brand_name)
                        if key not in brand_cache:
                            existing_b = db.query(DddBrand).filter_by(
                                market_id=market.id, name=brand_name
                            ).first()
                            if not existing_b:
                                existing_b = DddBrand(market_id=market.id, name=brand_name, is_siegfried=False)
                                db.add(existing_b)
                                db.flush()
                            brand_cache[key] = existing_b
                        ddd_brand = brand_cache[key]
                        existing = db.query(DddBrandMonthly).filter_by(
                            ddd_brand_id=ddd_brand.id, region_id=region.id, month=m
                        ).first()
                        if existing:
                            existing.units = int(val)
                        else:
                            db.add(DddBrandMonthly(
                                ddd_brand_id=ddd_brand.id, region_id=region.id,
                                month=m, units=int(val),
                            ))
                    else:
                        existing = db.query(DddTotalMonthly).filter_by(
                            market_id=market.id, region_id=region.id, month=m
                        ).first()
                        if existing:
                            existing.units = int(val)
                        else:
                            db.add(DddTotalMonthly(
                                market_id=market.id, region_id=region.id,
                                month=m, units=int(val),
                            ))
                    stats["datos_mensuales"] += 1
            except Exception as e:
                errors.append(f"Datos_Mensuales: {row.get('Marca', '?')} - {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


MONTH_NAMES_FULL = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

MONTH_NAMES_SHORT_ES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}

MONTH_NAMES_EN = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def parse_date_from_filename(filename: str) -> date | None:
    """Extrae fecha de nombres como '21 de febrero de 2026'."""
    m = re.search(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", filename)
    if m:
        day = int(m.group(1))
        month_name = m.group(2).lower()
        year = int(m.group(3))
        month_num = MONTH_NAMES_FULL.get(month_name)
        if month_num:
            return date(year, month_num, day)
    return None


def safe_int(val) -> int | None:
    if val is None or val == "" or val == "-":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None or val == "" or val == "-":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def ingest_convenios_detalle(db: Session, wb, dry_run: bool = False, filepath: str = "") -> dict:
    """Archivo 1: Detalle consumos y aportes por convenio."""
    stats = {"details": 0, "skipped_subtotals": 0}
    errors = []

    report_date = parse_date_from_filename(filepath)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"stats": stats, "errors": ["Hoja vacía"]}

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    # Contexto jerárquico carry-forward
    ctx_lab = ctx_line = ctx_familia = None

    for row_vals in rows[1:]:
        row = dict(zip(headers, row_vals))
        if all(v is None for v in row_vals):
            continue

        # Actualizar contexto jerárquico
        if row.get("Laboratorio"):
            ctx_lab = str(row["Laboratorio"]).strip()
        if row.get("Linea"):
            ctx_line = str(row["Linea"]).strip()
        if row.get("Familia"):
            ctx_familia = str(row["Familia"]).strip()

        # Filtrar subtotales (filas sin Producto o con "Totales")
        product = row.get("Producto")
        obra_social = row.get("ObraSocial1")
        if not product or not obra_social:
            stats["skipped_subtotals"] += 1
            continue
        product_str = str(product).strip()
        if product_str.lower() in ("totales", "", "total"):
            stats["skipped_subtotals"] += 1
            continue

        try:
            # Buscar brand por familia
            brand = get_or_create_brand(db, ctx_familia, is_siegfried=True) if ctx_familia else None

            obra_social_str = str(obra_social).strip()
            obra_social2 = str(row.get("ObraSocial2", "")).strip() if row.get("ObraSocial2") else None

            existing = (
                db.query(AgreementDetail)
                .filter_by(
                    brand_id=brand.id if brand else None,
                    health_plan=obra_social_str,
                    health_plan_detail=obra_social2,
                    product_name=product_str,
                    report_date=report_date,
                )
                .first()
            )
            data = dict(
                laboratory=ctx_lab,
                line=ctx_line,
                familia=ctx_familia,
                health_plan_detail=obra_social2,
                units=safe_int(row.get("Consumo uni")),
                pvp_amount=safe_float(row.get("Consumo PVP $")),
                system_coverage=safe_float(row.get("Cob sistema $")),
                lab_pvp_contribution=safe_float(row.get("Aporte PVP lab $")),
                lab_adjustment=safe_float(row.get("Ajuste lab $")),
                lab_total_contribution=safe_float(row.get("Aporte Lab Total")),
                net_contribution=safe_float(row.get("Aporte Neto")),
            )
            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
            else:
                db.add(AgreementDetail(
                    brand_id=brand.id if brand else None,
                    health_plan=obra_social_str,
                    product_name=product_str,
                    report_date=report_date,
                    **data,
                ))
            stats["details"] += 1
        except Exception as e:
            errors.append(f"Convenio detalle: {product_str} - {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_stock_pivot(db: Session, wb, dry_run: bool = False, filepath: str = "") -> dict:
    """Archivo 2: Stock pivot (Familia/Producto × meses)."""
    stats = {"presentations": 0, "brands": 0}
    errors = []

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {"stats": stats, "errors": ["Menos de 3 filas"]}

    # Parsear header de 2 filas
    row0 = list(rows[0])  # Mes-Año (fechas)
    row1 = list(rows[1])  # Familia, Producto, métricas

    # Construir mapeo: col_index -> (date, metric)
    col_map: dict[int, tuple[date, str]] = {}
    current_date = None
    for i in range(2, len(row1)):
        # row0 tiene fechas (datetime objects) o None
        if i < len(row0) and row0[i] is not None:
            dt = row0[i]
            if isinstance(dt, datetime):
                current_date = date(dt.year, dt.month, 1)
            elif isinstance(dt, date):
                current_date = date(dt.year, dt.month, 1)
        metric = str(row1[i]).strip() if row1[i] else None
        if current_date and metric:
            col_map[i] = (current_date, metric)

    # Agrupar por (familia, producto, date) => métricas
    brand_aggregates: dict[tuple[str, date], dict] = {}

    for row_vals in rows[2:]:
        if all(v is None for v in row_vals):
            continue
        familia = str(row_vals[0]).strip() if row_vals[0] else None
        producto = str(row_vals[1]).strip() if row_vals[1] else None

        if not familia and not producto:
            continue
        if producto and producto.lower() == "totales":
            # Fila de totales por familia -> agregar a stock_brand
            if familia:
                brand = get_or_create_brand(db, familia, is_siegfried=True)
                for i, (month_date, metric) in col_map.items():
                    val = row_vals[i] if i < len(row_vals) else None
                    key = (familia, month_date)
                    if key not in brand_aggregates:
                        brand_aggregates[key] = {"brand": brand}
                    norm_metric = metric.lower().replace("dias de stock", "days").replace("días de stock", "days")
                    if "stock final" in norm_metric:
                        brand_aggregates[key]["stock_units"] = safe_int(val)
                    elif "ventas" in norm_metric:
                        brand_aggregates[key]["sales"] = safe_int(val)
                    elif "facturación" in norm_metric or "facturacion" in norm_metric:
                        brand_aggregates[key]["billing"] = safe_float(val)
                    elif "dias" in norm_metric or "days" in norm_metric:
                        brand_aggregates[key]["days_cover"] = safe_int(val)
            continue

        # Fila de producto individual -> stock_presentation
        if familia and producto:
            try:
                brand = get_or_create_brand(db, familia, is_siegfried=True)
                pres = get_or_create_presentation(db, brand.id, producto, familia=familia)

                for i, (month_date, metric) in col_map.items():
                    val = row_vals[i] if i < len(row_vals) else None
                    if val is None:
                        continue

                    norm_metric = metric.lower()
                    existing = (
                        db.query(StockPresentation)
                        .filter_by(presentation_id=pres.id, month=month_date)
                        .first()
                    )
                    if not existing:
                        existing = StockPresentation(
                            presentation_id=pres.id, month=month_date,
                        )
                        db.add(existing)
                        db.flush()

                    if "stock final" in norm_metric:
                        existing.stock_units = safe_int(val)
                    elif "ventas" in norm_metric:
                        existing.sales = safe_int(val)
                    elif "facturación" in norm_metric or "facturacion" in norm_metric:
                        existing.billing = safe_float(val)
                    elif "dias" in norm_metric or "días" in norm_metric:
                        existing.days_cover = safe_int(val)

                stats["presentations"] += 1
            except Exception as e:
                errors.append(f"Stock pivot pres: {producto} - {e}")

    # Insertar agregados de brand
    for (familia_name, month_date), agg in brand_aggregates.items():
        try:
            brand = agg["brand"]
            existing = db.query(StockBrand).filter_by(brand_id=brand.id, month=month_date).first()
            if existing:
                if agg.get("stock_units") is not None:
                    existing.stock_units = agg["stock_units"]
                if agg.get("sales") is not None:
                    existing.sales = agg["sales"]
                if agg.get("billing") is not None:
                    existing.billing = agg["billing"]
                if agg.get("days_cover") is not None:
                    existing.days_cover = agg["days_cover"]
            else:
                db.add(StockBrand(
                    brand_id=brand.id, month=month_date,
                    stock_units=agg.get("stock_units"),
                    sales=agg.get("sales"),
                    billing=agg.get("billing"),
                    days_cover=agg.get("days_cover"),
                ))
            stats["brands"] += 1
        except Exception as e:
            errors.append(f"Stock pivot brand: {familia_name} - {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_mercado_regional(db: Session, wb, dry_run: bool = False, filepath: str = "") -> dict:
    """Archivo 3: Mercado regional (Producto-Molécula-ATC-provincia).
    Estrategia: DELETE + bulk INSERT por eficiencia con archivos grandes.
    Primera pasada: resolver dimensiones (mol, brand, region).
    Segunda pasada: bulk INSERT con IDs resueltos.
    """
    stats = {"records": 0, "duplicates": 0}
    errors = []

    ws = wb[wb.sheetnames[0]]

    # Caches de dimensiones
    mol_cache: dict[str, int] = {}
    brand_cache: dict[str, int] = {}
    region_cache: dict[str, int] = {}

    typer.echo("  Fase 1: Resolviendo dimensiones...")
    # Primera pasada rápida: resolver todas las dimensiones únicas
    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter)
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_raw)]

    unique_mols: set[tuple[str, str | None, str | None]] = set()
    unique_brands: set[str] = set()
    unique_regions: set[str] = set()

    for row_vals in rows_iter:
        row = dict(zip(headers, row_vals))
        droga = str(row.get("Droga", "")).strip()
        producto = str(row.get("Producto", "")).strip()
        region_name = str(row.get("RegionCUP", "")).strip()
        clase = str(row.get("Clase Terapeutica", "")).strip() if row.get("Clase Terapeutica") else None
        cod_clase = str(row.get("Codigo Clase Terapeutica", "")).strip() if row.get("Codigo Clase Terapeutica") else None
        if droga:
            unique_mols.add((droga, clase, cod_clase))
        if producto:
            unique_brands.add(producto)
        if region_name:
            unique_regions.add(region_name)

    for mol_name, clase, cod_clase in unique_mols:
        mol = get_or_create_molecule(db, mol_name, clase=clase)
        if cod_clase and not mol.atc_code:
            mol.atc_code = cod_clase
        mol_cache[mol_name] = mol.id

    for brand_name in unique_brands:
        brand = get_or_create_brand(db, brand_name)
        brand_cache[brand_name] = brand.id

    for region_name in unique_regions:
        region = get_or_create_region(db, region_name)
        region_cache[region_name] = region.id

    db.flush()
    typer.echo(f"  {len(unique_mols)} moléculas, {len(unique_brands)} marcas, {len(unique_regions)} regiones")

    # DELETE existentes
    if not dry_run:
        deleted = db.query(MarketPerformanceRegional).delete()
        typer.echo(f"  Borrados {deleted} registros previos")
        db.flush()

    # Segunda pasada: bulk INSERT
    typer.echo("  Fase 2: Inserción masiva...")
    seen_keys: set[tuple] = set()
    batch: list[MarketPerformanceRegional] = []
    BATCH_SIZE = 5000

    rows_iter2 = ws.iter_rows(values_only=True)
    next(rows_iter2)  # skip header

    for row_vals in rows_iter2:
        row = dict(zip(headers, row_vals))
        if all(v is None for v in row_vals):
            continue

        region_name = str(row.get("RegionCUP", "")).strip()
        droga = str(row.get("Droga", "")).strip()
        producto = str(row.get("Producto", "")).strip()
        unidades = safe_int(row.get("Unidades"))

        if not producto or not region_name or not droga:
            continue

        anio_mes = str(row.get("AñoMes", "")).strip()
        month_date = None
        if anio_mes and "-" in anio_mes:
            parts = anio_mes.split("-")
            month_num = MONTH_NAMES_SHORT_ES.get(parts[0].lower()[:3])
            if month_num and len(parts) > 1:
                month_date = date(int(parts[1]), month_num, 1)

        if not month_date:
            continue

        brand_id = brand_cache.get(producto)
        mol_id = mol_cache.get(droga)
        region_id = region_cache.get(region_name)
        if not brand_id or not mol_id or not region_id:
            continue

        key = (brand_id, mol_id, region_id, month_date)
        if key in seen_keys:
            stats["duplicates"] += 1
            continue
        seen_keys.add(key)

        batch.append(MarketPerformanceRegional(
            brand_id=brand_id, molecule_id=mol_id,
            region_id=region_id, month=month_date,
            units=unidades,
        ))
        stats["records"] += 1

        if len(batch) >= BATCH_SIZE:
            if not dry_run:
                db.bulk_save_objects(batch)
                db.flush()
            batch.clear()
            if stats["records"] % 50000 == 0:
                typer.echo(f"  ... {stats['records']} registros")

    if batch and not dry_run:
        db.bulk_save_objects(batch)

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_ateneo_csv(db: Session, filepath: str, dry_run: bool = False) -> dict:
    """Archivo 4: CSV ATENEO (mercado regional, denormalizado).
    Estrategia: DELETE por ref_date + bulk INSERT.
    """
    stats = {"records": 0}
    errors = []

    # MESACTUAL = mes anterior al corriente (Enero 2026)
    today = date.today()
    ref_month = today.month - 1 if today.month > 1 else 12
    ref_year = today.year if today.month > 1 else today.year - 1
    ref_date_val = date(ref_year, ref_month, 1)

    # Borrar registros previos con misma ref_date
    if not dry_run:
        deleted = db.query(MarketAteneo).filter(MarketAteneo.ref_date == ref_date_val).delete()
        typer.echo(f"  Borrados {deleted} registros previos con ref_date={ref_date_val}")
        db.flush()

    batch: list[MarketAteneo] = []
    BATCH_SIZE = 5000

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                record = MarketAteneo(
                    atc_code=row.get("CLASE", "").strip() or None,
                    region=row.get("REGION", "").strip() or None,
                    brand_name=row.get("MARCA", "").strip() or None,
                    product_name=row.get("PRODUCTO", "").strip() or None,
                    laboratory=row.get("LABORATORIO", "").strip() or None,
                    product_code=row.get("CODIGOPROD", "").strip() or None,
                    mat_5=safe_int(row.get("MAT5")),
                    mat_4=safe_int(row.get("MAT4")),
                    mat_3=safe_int(row.get("MAT3")),
                    mat_2=safe_int(row.get("MAT2")),
                    mat_current=safe_int(row.get("MATACTUAL")),
                    mes_12=safe_int(row.get("MES12")),
                    mes_11=safe_int(row.get("MES11")),
                    mes_10=safe_int(row.get("MES10")),
                    mes_9=safe_int(row.get("MES9")),
                    mes_8=safe_int(row.get("MES8")),
                    mes_7=safe_int(row.get("MES7")),
                    mes_6=safe_int(row.get("MES6")),
                    mes_5=safe_int(row.get("MES5")),
                    mes_4=safe_int(row.get("MES4")),
                    mes_3=safe_int(row.get("MES3")),
                    mes_2=safe_int(row.get("MES2")),
                    mes_current=safe_int(row.get("MESACTUAL")),
                    ytd_5=safe_int(row.get("YTD5")),
                    ytd_4=safe_int(row.get("YTD4")),
                    ytd_3=safe_int(row.get("YTD3")),
                    ytd_2=safe_int(row.get("YTD2")),
                    ytd_current=safe_int(row.get("YTDACTUAL")),
                    ref_date=ref_date_val,
                )
                batch.append(record)
                stats["records"] += 1

                if len(batch) >= BATCH_SIZE:
                    if not dry_run:
                        db.bulk_save_objects(batch)
                        db.flush()
                    batch.clear()

            except Exception as e:
                errors.append(f"ATENEO CSV: fila {stats['records']} - {e}")

    if batch and not dry_run:
        db.bulk_save_objects(batch)

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_ateneo_national(db: Session, wb, dry_run: bool = False, filepath: str = "") -> dict:
    """Archivo 5: MAT ATENEO nacional (wide → long).
    Parsea headers de 86 columnas, normaliza a 1 fila por pack × period × date.
    """
    stats = {"records": 0}
    errors = []

    # Procesar ambas sheets
    target_sheets = [s for s in wb.sheetnames if s.startswith("DS_AR_PM_FV_Standard")]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        raw_headers = list(rows[0])

        # Parsear las columnas de métricas (después de las 10 primeras dimensiones)
        dim_count = 10
        metric_cols: list[tuple[int, str, str, date]] = []
        # metric_cols = [(col_index, metric_group, period_type, period_date)]

        for i in range(dim_count, len(raw_headers)):
            h = str(raw_headers[i]).strip() if raw_headers[i] else ""
            if not h:
                continue

            # Headers tienen formato "US Dollars\nMAT Dec 2021" o "Units\nFeb 2025"
            parts = h.split("\n")
            if len(parts) != 2:
                continue

            metric_group = parts[0].strip()  # US Dollars, Local Currency, Units
            period_str = parts[1].strip()  # MAT Dec 2021, Feb 2025, YTD Dec 2025

            # Determinar period_type y fecha
            period_type = "MES"
            date_str = period_str
            if period_str.startswith("MAT "):
                period_type = "MAT"
                date_str = period_str[4:]
            elif period_str.startswith("YTD "):
                period_type = "YTD"
                date_str = period_str[4:]

            # Parsear fecha "Dec 2021", "Jan 2025"
            date_parts = date_str.split()
            if len(date_parts) == 2:
                month_num = MONTH_NAMES_EN.get(date_parts[0].lower()[:3])
                if month_num:
                    try:
                        period_date = date(int(date_parts[1]), month_num, 1)
                        metric_cols.append((i, metric_group, period_type, period_date))
                    except ValueError:
                        pass

        # Agrupar por (pack_code, period_type, period_date) => {usd, local, units}
        for row_vals in rows[1:]:
            if all(v is None for v in row_vals[:dim_count]):
                continue

            pack_code = str(row_vals[0]).strip() if row_vals[0] else None
            pack_name = str(row_vals[1]).strip() if row_vals[1] else None
            product_name = str(row_vals[2]).strip() if row_vals[2] else None
            corporation = str(row_vals[3]).strip() if row_vals[3] else None
            manufacturer = str(row_vals[4]).strip() if row_vals[4] else None
            atc_iv_raw = str(row_vals[5]).strip() if row_vals[5] else None
            # ATC IV: extraer solo el código antes de " - "
            atc_iv = atc_iv_raw.split(" - ")[0].strip() if atc_iv_raw else None
            molecule_val = str(row_vals[6]).strip() if row_vals[6] else None
            pharma_form = str(row_vals[7]).strip() if row_vals[7] else None
            launch_raw = row_vals[8]
            launch_date = None
            if launch_raw:
                if isinstance(launch_raw, (datetime, date)):
                    launch_date = str(launch_raw.date() if isinstance(launch_raw, datetime) else launch_raw)
                else:
                    launch_date = str(launch_raw).strip()
            market_type_raw = str(row_vals[9]).strip() if row_vals[9] else None
            market_type = "E" if market_type_raw and "ETICO" in market_type_raw.upper() else "OTC" if market_type_raw else None

            if not pack_code:
                continue

            # Agrupar métricas por (period_type, period_date)
            period_data: dict[tuple[str, date], dict] = {}
            for col_idx, metric_group, pt, pd in metric_cols:
                val = row_vals[col_idx] if col_idx < len(row_vals) else None
                if val is None:
                    continue
                key = (pt, pd)
                if key not in period_data:
                    period_data[key] = {}
                mg_lower = metric_group.lower()
                if "us dollar" in mg_lower:
                    period_data[key]["usd"] = safe_float(val)
                elif "local" in mg_lower:
                    period_data[key]["local"] = safe_float(val)
                elif "unit" in mg_lower:
                    period_data[key]["units"] = safe_int(val)

            for (pt, pd), vals in period_data.items():
                try:
                    existing = (
                        db.query(MarketAteneoNational)
                        .filter_by(pack_code=pack_code, period_type=pt, period_date=pd)
                        .first()
                    )
                    data = dict(
                        pack_name=pack_name, product_name=product_name,
                        corporation=corporation, manufacturer=manufacturer,
                        atc_iv=atc_iv, molecule=molecule_val,
                        pharma_form=pharma_form, launch_date=launch_date,
                        market_type=market_type,
                        usd_amount=vals.get("usd"),
                        local_amount=vals.get("local"),
                        units=vals.get("units"),
                    )
                    if existing:
                        for k, v in data.items():
                            setattr(existing, k, v)
                    else:
                        db.add(MarketAteneoNational(
                            pack_code=pack_code,
                            period_type=pt,
                            period_date=pd,
                            **data,
                        ))
                    stats["records"] += 1
                except Exception as e:
                    errors.append(f"ATENEO nacional: {pack_code}/{pt}/{pd} - {e}")

        if not dry_run:
            db.flush()

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


def ingest_price_catalog(db: Session, wb, dry_run: bool = False, filepath: str = "") -> dict:
    """Archivo 6: Catálogo de precios ANMAT."""
    stats = {"records": 0}
    errors = []

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"stats": stats, "errors": ["Hoja vacía"]}

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    # Detectar fechas de PVP del header
    pvp_prev_date = pvp_curr_date = None
    for h in headers:
        m = re.search(r"PVP al (\d{2}/\d{2}/\d{4})", h)
        if m:
            d = datetime.strptime(m.group(1), "%d/%m/%Y").date()
            if pvp_prev_date is None:
                pvp_prev_date = d
            else:
                pvp_curr_date = d

    for row_vals in rows[1:]:
        row = dict(zip(headers, row_vals))
        if all(v is None for v in row_vals):
            continue

        registro = str(row.get("Registro", "")).strip() if row.get("Registro") else None
        presentation = str(row.get("Presentacion", "")).strip() if row.get("Presentacion") else None

        if not registro and not presentation:
            continue

        try:
            troquel = str(row.get("Troquel", "")).strip() if row.get("Troquel") else None

            # Parsear fecha de vigencia
            eff_date_raw = row.get("Fecha Vigencia")
            eff_date = None
            if eff_date_raw:
                if isinstance(eff_date_raw, (datetime, date)):
                    eff_date = eff_date_raw if isinstance(eff_date_raw, date) else eff_date_raw.date()
                elif isinstance(eff_date_raw, str):
                    try:
                        eff_date = datetime.strptime(eff_date_raw.strip(), "%Y-%m-%d").date()
                    except ValueError:
                        pass

            existing = (
                db.query(PriceCatalog)
                .filter_by(registro=registro, presentation=presentation)
                .first()
            )

            # Detectar columnas de PVP dinámicamente
            pvp_prev_val = pvp_curr_val = None
            for h in headers:
                if h.startswith("PVP al"):
                    if pvp_prev_val is None:
                        pvp_prev_val = safe_float(row.get(h))
                    else:
                        pvp_curr_val = safe_float(row.get(h))

            data = dict(
                troquel=troquel,
                product_name=str(row.get("Producto", "")).strip() if row.get("Producto") else None,
                drug_name=str(row.get("Droga", "")).strip() if row.get("Droga") else None,
                pharma_action=str(row.get("Acción Farm", row.get("Accion Farm", ""))).strip() or None,
                laboratory=str(row.get("Laboratorio", "")).strip() if row.get("Laboratorio") else None,
                barcode=str(row.get("Cod Barras", "")).strip() if row.get("Cod Barras") else None,
                product_type=str(row.get("Tipo", "")).strip() if row.get("Tipo") else None,
                pami_category=str(row.get("PAMI", "")).strip() if row.get("PAMI") else None,
                status=str(row.get("Estado", "")).strip() if row.get("Estado") else None,
                qty_presentations=safe_int(row.get("Q Pres")),
                pvp_previous=pvp_prev_val,
                pvp_current=pvp_curr_val,
                variation_pct=safe_float(row.get("% Var")),
                effective_date=eff_date,
                pvp_previous_date=pvp_prev_date,
                pvp_current_date=pvp_curr_date,
            )

            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
            else:
                db.add(PriceCatalog(
                    registro=registro,
                    presentation=presentation,
                    **data,
                ))
            stats["records"] += 1
        except Exception as e:
            errors.append(f"Precio catálogo: {registro} - {e}")

    if not dry_run:
        db.commit()
    return {"stats": stats, "errors": errors}


INGESTORS = {
    "cardio_ventas": ingest_cardio_ventas,
    "cardio_mercado": ingest_cardio_mercado,
    "cardio_recetas_stock": ingest_cardio_recetas_stock,
    "ddd_data": ingest_ddd_data,
    "convenios_detalle": ingest_convenios_detalle,
    "stock_pivot": ingest_stock_pivot,
    "mercado_regional": ingest_mercado_regional,
    "ateneo_national": ingest_ateneo_national,
    "price_catalog": ingest_price_catalog,
}


def detect_csv_file_type(filepath: str) -> str | None:
    """Detecta tipo de CSV por sus headers."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
    header_set = {h.strip() for h in headers}
    if "CLASE" in header_set and "REGION" in header_set and "MARCA" in header_set:
        return "ateneo_csv"
    return None


def process_file(path: Path, dry_run: bool = False, strict: bool = False) -> None:
    typer.echo(f"\n{'='*60}")
    typer.echo(f"Procesando: {path.name}")

    db = get_session()
    try:
        # CSV: procesamiento especial
        if path.suffix.lower() == ".csv":
            file_type = detect_csv_file_type(str(path))
            if not file_type:
                typer.echo("No se pudo detectar el tipo de CSV")
                return
            typer.echo(f"Tipo detectado: {file_type}")
            result = ingest_ateneo_csv(db, str(path), dry_run=dry_run)
        else:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                file_type = detect_file_type_xlsx(wb)
                typer.echo(f"Tipo detectado: {file_type}")

                ingestor = INGESTORS[file_type]
                # Funciones nuevas reciben filepath como kwarg
                if file_type in ("convenios_detalle", "stock_pivot", "mercado_regional", "ateneo_national", "price_catalog"):
                    result = ingestor(db, wb, dry_run=dry_run, filepath=str(path))
                else:
                    result = ingestor(db, wb, dry_run=dry_run)
            finally:
                wb.close()

        typer.echo(f"\nResultados {'(DRY RUN)' if dry_run else ''}:")
        for key, val in result["stats"].items():
            typer.echo(f"  {key}: {val} filas")

        if result["errors"]:
            typer.echo(f"\nErrores ({len(result['errors'])}):")
            for err in result["errors"][:20]:
                typer.echo(f"  ! {err}")
            if strict:
                db.rollback()
                typer.echo("STRICT MODE: rollback realizado")
                raise typer.Exit(1)
    finally:
        db.close()


@app.command()
def main(
    path: str = typer.Argument(help="Ruta al archivo Excel o directorio"),
    all: bool = typer.Option(False, "--all", help="Procesar todos los xlsx del directorio"),
    dry_run: bool = typer.Option(False, "--dry-run", help="Solo validar, no escribir"),
    strict: bool = typer.Option(False, "--strict", help="Abortar en cualquier error"),
):
    """Ingesta de datos Excel/CSV a PostgreSQL."""
    p = Path(path)

    if all and p.is_dir():
        files = sorted(list(p.glob("*.xlsx")) + list(p.glob("*.csv")))
        if not files:
            typer.echo(f"No se encontraron archivos .xlsx/.csv en {p}")
            raise typer.Exit(1)
        for f in files:
            process_file(f, dry_run=dry_run, strict=strict)
    elif p.is_file():
        process_file(p, dry_run=dry_run, strict=strict)
    else:
        typer.echo(f"Ruta inválida: {path}")
        raise typer.Exit(1)

    typer.echo("\nIngesta completada.")


if __name__ == "__main__":
    app()
