"""Extrae datos de los HTML estáticos y genera los Excel para ingest.py.

Lee cardio/index.html y cardio/DDD/index.html, parsea el objeto `const D`
y genera 4 archivos Excel en data/ con las hojas que ingest.py espera.

Uso:
    python3 scripts/extract_from_html.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook

# ── Rutas ──────────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
CARDIO_HTML = ROOT / "cardio" / "index.html"
DDD_HTML = ROOT / "cardio" / "DDD" / "index.html"
DATA_DIR = ROOT / "data"

# ── Mapeo de meses ingleses → numérico ────────────────────────────────────

MONTH_MAP: dict[str, str] = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}

MONTH_NAMES_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                  "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def parse_month_en(label: str) -> str:
    """Convierte 'Jan 2024' → '2024-01'."""
    parts = label.split()
    if len(parts) != 2:
        raise ValueError(f"Formato de mes no reconocido: {label}")
    return f"{parts[1]}-{MONTH_MAP[parts[0]]}"


# ── Extracción del JSON embebido ──────────────────────────────────────────

def extract_const_d(html_path: Path) -> dict:
    """Lee un HTML y extrae el objeto JSON de `const D = {...};`."""
    content = html_path.read_text(encoding="utf-8")
    match = re.search(r"const\s+D\s*=\s*(\{.+?\});", content, re.DOTALL)
    if not match:
        print(f"ERROR: No se encontró 'const D' en {html_path}")
        sys.exit(1)
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError as e:
        print(f"ERROR: JSON inválido en {html_path}: {e}")
        sys.exit(1)


# ── Helpers para Excel ────────────────────────────────────────────────────

def write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]):
    """Crea una hoja con headers y filas."""
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    print(f"  {name}: {len(rows)} filas")


# ── Generador 1: cardio_ventas.xlsx ──────────────────────────────────────

def gen_cardio_ventas(D: dict, wb: Workbook) -> None:
    # Budget
    headers = ["Marca", "Año"]
    for m in MONTH_NAMES_ES:
        headers += [f"{m}_Budget", f"{m}_Real"]

    rows = []
    for brand, years in D["budget"].items():
        for year_str, data in years.items():
            row = [brand, int(year_str)]
            budget_vals = data.get("budget", [None] * 12)
            real_vals = data.get("real", [None] * 12)
            for i in range(12):
                row.append(budget_vals[i] if i < len(budget_vals) else None)
                row.append(real_vals[i] if i < len(real_vals) else None)
            rows.append(row)
    write_sheet(wb, "Budget", headers, rows)

    # Canales
    headers = ["Marca", "Unidades", "Convenios_%", "Mostrador_%"]
    rows = []
    for brand, data in D["canales"].items():
        rows.append([brand, data["unid"], data.get("conv"), data.get("most")])
    write_sheet(wb, "Canales", headers, rows)

    # KPIs - combina kpiStrip (global) y kpiByBrand (por marca)
    # Determinar todas las columnas de KPI
    kpi_keys_global = sorted(D["kpiStrip"].keys())
    kpi_keys_brand = set()
    for brand_data in D["kpiByBrand"].values():
        kpi_keys_brand.update(brand_data.keys())
    # Unir todas las claves KPI, priorizando el orden de kpiStrip
    all_kpi_keys = list(kpi_keys_global)
    for k in sorted(kpi_keys_brand):
        if k not in all_kpi_keys:
            all_kpi_keys.append(k)

    headers = ["Tipo", "Marca"] + all_kpi_keys
    rows = []

    # Fila global
    row = ["KPI_Global", None]
    for k in all_kpi_keys:
        row.append(D["kpiStrip"].get(k))
    rows.append(row)

    # Filas por marca
    for brand, data in D["kpiByBrand"].items():
        row = ["KPI_Marca", brand]
        for k in all_kpi_keys:
            row.append(data.get(k))
        rows.append(row)

    write_sheet(wb, "KPIs", headers, rows)


# ── Generador 2: cardio_mercado.xlsx ─────────────────────────────────────

def gen_cardio_mercado(D: dict, wb: Workbook) -> None:
    # Productos - catálogo de productos con molécula
    headers = ["Molécula", "Producto", "Laboratorio", "Es_Siegfried"]
    rows = []
    seen = set()
    for mol_name, mol_data in D["mol_perf"].items():
        for prod_info in mol_data["products"]:
            prod_name = prod_info["prod"]
            if (mol_name, prod_name) in seen:
                continue
            seen.add((mol_name, prod_name))
            rows.append([
                mol_name,
                prod_name,
                prod_info.get("manuf", ""),
                prod_info.get("is_sie", False),
            ])
    write_sheet(wb, "Productos", headers, rows)

    # Performance - datos mensuales por producto/molécula
    headers = ["Molécula", "Producto", "Mes", "Unidades", "MS_%"]
    rows = []
    for mol_name, mol_data in D["mol_perf"].items():
        for prod_info in mol_data["products"]:
            prod_name = prod_info["prod"]
            monthly_vals = prod_info.get("monthly_vals", {})
            ms_monthly = prod_info.get("ms_monthly", {})
            for month_label, units in monthly_vals.items():
                mes_fmt = parse_month_en(month_label)
                ms_val = ms_monthly.get(month_label)
                rows.append([mol_name, prod_name, mes_fmt, units, ms_val])
    write_sheet(wb, "Performance", headers, rows)


# ── Generador 3: cardio_recetas_stock.xlsx ───────────────────────────────

def gen_cardio_recetas_stock(D: dict, wb: Workbook) -> None:
    # Recetas
    headers = ["Marca", "Mes", "Recetas", "Médicos"]
    rows = []
    for brand, months_data in D["recetas"].items():
        for month_label, data in months_data.items():
            mes_fmt = parse_month_en(month_label)
            rows.append([brand, mes_fmt, data["recetas"], data["medicos"]])
    write_sheet(wb, "Recetas", headers, rows)

    # Recetas_MS
    headers = ["Marca", "Mes", "Recetas_SIE", "Recetas_Mercado", "MS_%"]
    rows = []
    for brand, data in D["rec_ms"].items():
        sie_data = data.get("sie", {})
        ms_data = data.get("ms", {})
        mkt_data = data.get("mkt", {})
        for month_label in sie_data:
            mes_fmt = parse_month_en(month_label)
            rows.append([
                brand,
                mes_fmt,
                sie_data.get(month_label),
                mkt_data.get(month_label),
                ms_data.get(month_label),
            ])
    write_sheet(wb, "Recetas_MS", headers, rows)

    # Convenios
    headers = ["Marca", "Obra_Social", "Unid_2025", "Unid_2024", "Delta_%", "Neto_$"]
    rows = []
    for brand, items in D["convenios"].items():
        for item in items:
            rows.append([
                brand,
                item["os"],
                item.get("unid"),
                item.get("unid24"),
                item.get("delta"),
                item.get("neto"),
            ])
    write_sheet(wb, "Convenios", headers, rows)

    # Stock_Marca - de D.stock (tiene dias, ventas, stock por mes)
    headers = ["Marca", "Mes", "Dias", "Ventas", "Stock"]
    rows = []
    for brand, months_data in D["stock"].items():
        for month_label, data in months_data.items():
            mes_fmt = parse_month_en(month_label)
            rows.append([
                brand,
                mes_fmt,
                data.get("dias"),
                data.get("ventas"),
                data.get("stock"),
            ])
    write_sheet(wb, "Stock_Marca", headers, rows)

    # Stock_Presentación - de D.stock_pres con stock_pres_months
    month_labels = D.get("stock_pres_months", [])
    sie_prods = D.get("sieProds", [])

    headers = ["Presentación", "Marca", "Familia", "Mes", "Ventas", "Dias", "Estado"]
    rows = []
    for pres_name, pres_data in D["stock_pres"].items():
        # Derivar marca: primero por prefijo en sieProds, fallback a familia
        familia = pres_data.get("familia", "")
        marca = _derive_brand(pres_name, sie_prods) or familia
        ventas = pres_data.get("ventas", [])
        dias = pres_data.get("dias", [])
        statuses = pres_data.get("statuses", [])

        for i, month_label in enumerate(month_labels):
            if i >= len(ventas):
                break
            mes_fmt = parse_month_en(month_label)
            rows.append([
                pres_name,
                marca,
                familia,
                mes_fmt,
                ventas[i] if i < len(ventas) else None,
                dias[i] if i < len(dias) else None,
                statuses[i] if i < len(statuses) else None,
            ])
    write_sheet(wb, "Stock_Presentación", headers, rows)

    # Precios - de D.precios
    # Los campos pvp tienen sufijos dinámicos (pvp_dic25, pvp_feb26)
    # Detectamos los nombres reales de los campos
    pvp_ant_key, pvp_act_key = _detect_pvp_keys(D["precios"])

    headers = ["Marca_Ref", "Presentación", "Laboratorio", "Producto",
               "PVP_Ant", "PVP_Act", "Var_%", "Es_SIE"]
    rows = []
    for brand_ref, presentations in D["precios"].items():
        for pres_name, items in presentations.items():
            for item in items:
                rows.append([
                    brand_ref,
                    pres_name,
                    item.get("lab", ""),
                    item.get("prod", ""),
                    item.get(pvp_ant_key),
                    item.get(pvp_act_key),
                    item.get("var"),
                    item.get("is_sie", False),
                ])
    write_sheet(wb, "Precios", headers, rows)


def _derive_brand(pres_name: str, sie_prods: list[str]) -> str:
    """Deriva la marca de una presentación buscando qué marca de sieProds es prefijo.

    Ordena por largo descendente para matchear primero los nombres más específicos
    (ej: 'DILATREND AP' antes que 'DILATREND').
    """
    pres_upper = pres_name.upper()
    # Ordenar por largo descendente para matchear prefijos más largos primero
    for prod in sorted(sie_prods, key=len, reverse=True):
        if pres_upper.startswith(prod.upper()):
            return prod
    return ""


def _detect_pvp_keys(precios: dict) -> tuple[str, str]:
    """Detecta las claves pvp_* dinámicas del primer item de precios.

    Retorna (pvp_anterior, pvp_actual) basándose en que la key con fecha
    menor es la anterior.
    """
    for presentations in precios.values():
        for items in presentations.values():
            if items:
                keys = [k for k in items[0].keys() if k.startswith("pvp_")]
                if len(keys) >= 2:
                    # Ordenar por la parte de fecha para determinar anterior/actual
                    keys.sort()
                    return keys[0], keys[1]
                elif len(keys) == 1:
                    return keys[0], keys[0]
    return "pvp_ant", "pvp_act"


# ── Generador 4: ddd_data.xlsx ───────────────────────────────────────────

def gen_ddd_data(D: dict, wb: Workbook) -> None:
    months = D["months"]  # ['Ene-2025', 'Feb-2025', ...]

    # Mercados
    headers = ["Mercado", "Clase", "Total_Unid", "Unid_SIE", "MS_%"]
    rows = []
    for mkt_name, mkt_data in D["markets"].items():
        rows.append([
            mkt_name,
            mkt_data.get("clase"),
            mkt_data.get("total_units"),
            mkt_data.get("sie_units"),
            mkt_data.get("global_ms"),
        ])
    write_sheet(wb, "Mercados", headers, rows)

    # Marcas
    headers = ["Mercado", "Marca", "Es_SIE"]
    rows = []
    for mkt_name, mkt_data in D["markets"].items():
        brand_meta = mkt_data.get("brand_meta", {})
        for brand_name in mkt_data.get("brands", []):
            meta = brand_meta.get(brand_name, {})
            rows.append([
                mkt_name,
                brand_name,
                meta.get("sie", False),
            ])
    write_sheet(wb, "Marcas", headers, rows)

    # Datos_Mensuales - formato ancho con meses como columnas
    # Usamos los nombres de meses en español (Ene, Feb, ...) que ingest.py espera
    headers = ["Mercado", "Marca", "Región"] + MONTH_NAMES_ES
    rows = []
    for mkt_name, mkt_data in D["markets"].items():
        # Datos por marca
        brand_monthly = mkt_data.get("brand_monthly", {})
        for brand_name, regions_data in brand_monthly.items():
            for region_name, values in regions_data.items():
                row = [mkt_name, brand_name, region_name]
                for i in range(12):
                    row.append(values[i] if i < len(values) else None)
                rows.append(row)

        # Totales del mercado (marca = __TOTAL__)
        total_monthly = mkt_data.get("total_monthly", {})
        for region_name, values in total_monthly.items():
            row = [mkt_name, "__TOTAL__", region_name]
            for i in range(12):
                row.append(values[i] if i < len(values) else None)
            rows.append(row)

    write_sheet(wb, "Datos_Mensuales", headers, rows)


# ── Main ──────────────────────────────────────────────────────────────────

def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # ── Cardio ────────────────────────────────────────────────────────────
    print(f"Leyendo {CARDIO_HTML.relative_to(ROOT)} ...")
    D_cardio = extract_const_d(CARDIO_HTML)

    # 1) cardio_ventas.xlsx
    print("\nGenerando cardio_ventas.xlsx ...")
    wb = Workbook()
    wb.remove(wb.active)
    gen_cardio_ventas(D_cardio, wb)
    out = DATA_DIR / "cardio_ventas.xlsx"
    wb.save(out)
    print(f"  → {out.relative_to(ROOT)}")

    # 2) cardio_mercado.xlsx
    print("\nGenerando cardio_mercado.xlsx ...")
    wb = Workbook()
    wb.remove(wb.active)
    gen_cardio_mercado(D_cardio, wb)
    out = DATA_DIR / "cardio_mercado.xlsx"
    wb.save(out)
    print(f"  → {out.relative_to(ROOT)}")

    # 3) cardio_recetas_stock.xlsx
    print("\nGenerando cardio_recetas_stock.xlsx ...")
    wb = Workbook()
    wb.remove(wb.active)
    gen_cardio_recetas_stock(D_cardio, wb)
    out = DATA_DIR / "cardio_recetas_stock.xlsx"
    wb.save(out)
    print(f"  → {out.relative_to(ROOT)}")

    # ── DDD ───────────────────────────────────────────────────────────────
    print(f"\nLeyendo {DDD_HTML.relative_to(ROOT)} ...")
    D_ddd = extract_const_d(DDD_HTML)

    # 4) ddd_data.xlsx
    print("\nGenerando ddd_data.xlsx ...")
    wb = Workbook()
    wb.remove(wb.active)
    gen_ddd_data(D_ddd, wb)
    out = DATA_DIR / "ddd_data.xlsx"
    wb.save(out)
    print(f"  → {out.relative_to(ROOT)}")

    print("\nExtracción completada.")


if __name__ == "__main__":
    main()
