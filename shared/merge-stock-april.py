#!/usr/bin/env python3
"""
shared/merge-stock-april.py

Actualiza datos de Stock/Cobertura para TODAS las lineas con la data
del SAP pivot 'Laboratorio - Familia - Producto - 11 de mayo de 2026.xlsx'
(meses Ene/Feb/Mar/Abr 2026).

Aplicado por linea:
1. D.stock[familia][month_key] = {stock, ventas, facturacion, dias}
   para los 4 meses del pivot (overwrite si existe, agrega si no).
2. D.coverage_labels (o D.stock_pres_months en SNC): agrega "Abr 26" si
   no esta presente. Para lineas data.js mantiene rolling 12-month
   shifteando si crece >12.
3. D.stock_alerts[familia]: appends 1 entry nueva por familia
   (ventas + dias + status del mes nuevo), recomputa alert_indices,
   worst_status, n_alerts. Shift left si len > target.
4. D.stock_pres[producto]: igual que stock_alerts pero match por
   nombre de producto (fuzzy por familia + similitud nombre).

NO toca otras secciones (mol_perf, recetas, budget, canales, precios).
"""
from __future__ import annotations
import re, json, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
PIVOT = Path(r'C:\Users\camarinaro\Downloads\Laboratorio - Familia - Producto - 11 de mayo de 2026 (1).xlsx')

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
MES_ES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
          7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
MES_ES_TO_NUM = {v:k for k,v in MES_ES.items()}
MES_EN_TO_NUM = {v:k for k,v in MES_EN.items()}

LINES_DATAJS = ['cardio', 'ATB', 'OTC', 'respiratorio']
LINES_INLINE = {
    'mujer': 'mujer/index.html',
    'SNC': 'SNC/index.html',
    'dermatologia': 'dermatologia/dermato_dashboard.html',
}


def cov_label_es(year, month):
    """e.g. (2026, 4) -> 'Abr 26'"""
    return f'{MES_ES[month]} {str(year)[-2:]}'


def month_key_en(year, month):
    """e.g. (2026, 4) -> 'Apr 2026'"""
    return f'{MES_EN[month]} {year}'


def cov_classify(d):
    if d is None: return 'nd'
    if d == 0: return 'quiebre'
    if d < 14: return 'bajo'
    if d < 20: return 'alerta'
    return 'ok'


def parse_pivot(path):
    """Returns (family_data, product_data, months_in_pivot)
       family_data[familia_upper][mk_en] = {stock, ventas, facturacion, dias}
       product_data[producto_norm][mk_en] = {...} + 'familia' attr"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    col_map = {}  # col_idx -> (mk_en, metric)
    months_in_pivot = []
    for i, h1 in enumerate(row1):
        if not isinstance(h1, datetime): continue
        mk = f'{MES_EN[h1.month]} {h1.year}'
        if mk not in months_in_pivot: months_in_pivot.append(mk)
        h2 = (str(row2[i] or '')).strip().lower()
        metric = None
        if 'stock final' in h2: metric = 'stock'
        elif 'ventas' in h2: metric = 'ventas'
        elif 'facturaci' in h2: metric = 'facturacion'
        elif 'dias' in h2 or 'días' in h2: metric = 'dias'
        if metric:
            col_map[i] = (mk, metric)

    family_data = defaultdict(lambda: defaultdict(dict))
    product_data = defaultdict(lambda: defaultdict(dict))
    product_to_fam = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row: continue
        lab = row[0] if len(row) > 0 else None
        familia = row[1] if len(row) > 1 else None
        producto = row[2] if len(row) > 2 else None
        if not lab or str(lab).strip() == 'Totales': continue
        if not familia or str(familia).strip() == 'Totales': continue
        fam_key = str(familia).strip().upper()
        is_family_row = (str(producto or '').strip() == 'Totales')

        if is_family_row:
            for col_idx, (mk, metric) in col_map.items():
                if col_idx >= len(row): continue
                v = row[col_idx]
                try: vf = float(v) if v is not None else None
                except: vf = None
                if vf is None: continue
                family_data[fam_key][mk][metric] = int(round(vf))
            continue

        if not producto: continue
        prod_key = str(producto).strip()
        product_to_fam[prod_key] = fam_key
        for col_idx, (mk, metric) in col_map.items():
            if col_idx >= len(row): continue
            v = row[col_idx]
            try: vf = float(v) if v is not None else None
            except: vf = None
            if vf is None: continue
            product_data[prod_key][mk][metric] = int(round(vf))

    wb.close()
    return dict(family_data), dict(product_data), product_to_fam, months_in_pivot


def normalize_prod_name(s):
    """Normaliza nombre de producto para fuzzy match."""
    if not s: return ''
    s = str(s).upper()
    s = s.replace('\xa0', ' ')  # non-breaking space
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def fuzzy_match_product(dashboard_name, xlsx_products, familia_hint=None):
    """Find xlsx product key best matching dashboard_name. Prefers same familia."""
    dn = normalize_prod_name(dashboard_name)
    dn_tokens = set(dn.split())
    best_match = None
    best_score = -1
    for xp in xlsx_products:
        xn = normalize_prod_name(xp)
        xn_tokens = set(xn.split())
        if not xn_tokens: continue
        common = dn_tokens & xn_tokens
        score = len(common) - 0.1 * abs(len(xn_tokens) - len(dn_tokens))
        if score > best_score:
            best_score = score
            best_match = xp
    # Require at least 50% token overlap
    if best_match:
        bn = normalize_prod_name(best_match)
        bn_tokens = set(bn.split())
        overlap = len(dn_tokens & bn_tokens) / max(1, len(dn_tokens))
        if overlap < 0.5: return None
    return best_match


def find_d_block(text, is_datajs):
    if is_datajs:
        m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    else:
        m = re.search(r'const\s+D\s*=\s*', text)
    if not m: return None
    obj_start = text.index('{', m.end())
    D, end_off = json.JSONDecoder().raw_decode(text[obj_start:])
    return obj_start, D, obj_start + end_off


def recompute_alerts(entry, target_window_len=None):
    """Recomputa alert_indices, worst_status, n_alerts a partir de statuses[]."""
    statuses = entry.get('statuses', [])
    # Shift to target window length if needed
    if target_window_len and len(statuses) > target_window_len:
        excess = len(statuses) - target_window_len
        for k in ('ventas','dias','statuses'):
            if k in entry: entry[k] = entry[k][excess:]
        statuses = entry['statuses']
    alerts = [i for i, s in enumerate(statuses) if s in ('quiebre','critico','bajo','alerta')]
    entry['alert_indices'] = alerts
    entry['n_alerts'] = len(alerts)
    order = {'quiebre':5,'critico':4,'bajo':3,'alerta':2,'ok':1,'nd':0}
    worst = 'nd'
    for s in statuses:
        if order.get(s,0) > order.get(worst,0): worst = s
    entry['worst_status'] = worst


def update_line(line_name, path_rel, is_datajs, fam_data, prod_data, product_to_fam, target_cov_label):
    p = REPO / path_rel
    text = p.read_text(encoding='utf-8-sig' if is_datajs else 'utf-8', errors='replace')
    pos = find_d_block(text, is_datajs)
    if not pos:
        print(f'  [{line_name}] SKIP: no D block'); return
    obj_start, D, abs_end = pos

    # ===== STEP 1: Update D.stock with all 4 months from pivot =====
    stock = D.setdefault('stock', {})
    stock_brands = set(stock.keys())
    n_stock_updated = 0
    for fam in stock_brands:
        fam_upper = fam.strip().upper()
        pdata = fam_data.get(fam_upper)
        if not pdata:
            # try whitespace-normalized
            for k in fam_data:
                if k.replace(' ','') == fam_upper.replace(' ',''):
                    pdata = fam_data[k]; break
        if not pdata: continue
        for mk_en, vals in pdata.items():
            if 'stock' not in vals and 'ventas' not in vals: continue
            stock[fam][mk_en] = {
                'stock': vals.get('stock', 0) or 0,
                'ventas': vals.get('ventas', 0) or 0,
                'facturacion': vals.get('facturacion', 0) or 0,
                'dias': vals.get('dias', 0) or 0,
            }
            n_stock_updated += 1

    # ===== STEP 2: Update coverage_labels (or stock_pres_months for SNC) =====
    use_cov_key = 'coverage_labels'
    if 'stock_pres_months' in D and D.get('stock_pres_months'):
        use_cov_key = 'stock_pres_months'
    cov_labels = D.get(use_cov_key, [])

    target_label = target_cov_label  # 'Abr 26' for ES, 'Apr 2026' for EN
    if use_cov_key == 'stock_pres_months':
        target_label = 'Apr 2026'

    cov_was_appended = False
    target_idx = None  # index of target month in cov_labels (for refresh or append)
    if cov_labels:
        if target_label in cov_labels:
            target_idx = cov_labels.index(target_label)
        else:
            cov_labels.append(target_label)
            cov_was_appended = True
            target_idx = len(cov_labels) - 1
        D[use_cov_key] = cov_labels

    # Determine target window length: keep 12 for data.js lines if already 12+
    target_len = None
    if is_datajs and len(cov_labels) > 12:
        excess = len(cov_labels) - 12
        cov_labels = cov_labels[excess:]
        D[use_cov_key] = cov_labels
        target_len = 12
    elif is_datajs:
        target_len = len(cov_labels)
    else:
        target_len = len(cov_labels) if cov_labels else None

    # ===== STEP 3: stock_alerts (per familia) =====
    n_pres_updated = 0
    n_alerts_updated = 0
    sa = D.get('stock_alerts', {})
    for fam, entry in sa.items():
        if not isinstance(entry, dict): continue
        if target_idx is None: continue  # no place to put new data
        fam_upper = fam.strip().upper()
        pdata = fam_data.get(fam_upper)
        if not pdata:
            for k in fam_data:
                if k.replace(' ','') == fam_upper.replace(' ',''):
                    pdata = fam_data[k]; break
        new_mk = 'Apr 2026'
        if pdata and new_mk in pdata:
            v = pdata[new_mk]
            ventas_v = v.get('ventas', 0) or 0
            dias_v = v.get('dias')
            if dias_v is None or dias_v == 0:
                if (v.get('stock', 0) or 0) == 0 and ventas_v == 0:
                    ventas_v = None; dias_v = None
            status = cov_classify(dias_v) if dias_v is not None else 'nd'
        else:
            ventas_v = None; dias_v = None; status = 'nd'

        ventas_arr = entry.setdefault('ventas', [])
        dias_arr = entry.setdefault('dias', [])
        statuses_arr = entry.setdefault('statuses', [])
        if cov_was_appended:
            ventas_arr.append(ventas_v); dias_arr.append(dias_v); statuses_arr.append(status)
        else:
            # Refresh existing slot at target_idx
            while len(ventas_arr) <= target_idx: ventas_arr.append(None)
            while len(dias_arr) <= target_idx: dias_arr.append(None)
            while len(statuses_arr) <= target_idx: statuses_arr.append('nd')
            ventas_arr[target_idx] = ventas_v
            dias_arr[target_idx] = dias_v
            statuses_arr[target_idx] = status
        recompute_alerts(entry, target_len)
        n_alerts_updated += 1

    # ===== STEP 4: stock_pres (per producto) =====
    sp = D.get('stock_pres', {})
    xlsx_products = list(prod_data.keys())
    for prod_name, entry in sp.items():
        if not isinstance(entry, dict): continue
        if target_idx is None: continue
        fam_hint = entry.get('familia', '')
        # Restrict xlsx products to those with matching familia
        candidates = [xp for xp in xlsx_products
                      if product_to_fam.get(xp, '').upper() == fam_hint.upper()]
        if not candidates:
            candidates = xlsx_products  # fallback
        match = fuzzy_match_product(prod_name, candidates, fam_hint)
        new_mk = 'Apr 2026'
        if match and new_mk in prod_data[match]:
            v = prod_data[match][new_mk]
            ventas_v = v.get('ventas', 0) or 0
            dias_v = v.get('dias')
            if dias_v is None or dias_v == 0:
                if (v.get('stock', 0) or 0) == 0 and ventas_v == 0:
                    ventas_v = None; dias_v = None
            status = cov_classify(dias_v) if dias_v is not None else 'nd'
        else:
            ventas_v = None; dias_v = None; status = 'nd'

        ventas_arr = entry.setdefault('ventas', [])
        dias_arr = entry.setdefault('dias', [])
        statuses_arr = entry.setdefault('statuses', [])
        if cov_was_appended:
            ventas_arr.append(ventas_v); dias_arr.append(dias_v); statuses_arr.append(status)
        else:
            while len(ventas_arr) <= target_idx: ventas_arr.append(None)
            while len(dias_arr) <= target_idx: dias_arr.append(None)
            while len(statuses_arr) <= target_idx: statuses_arr.append('nd')
            ventas_arr[target_idx] = ventas_v
            dias_arr[target_idx] = dias_v
            statuses_arr[target_idx] = status
        recompute_alerts(entry, target_len)
        n_pres_updated += 1

    # ===== Write back =====
    new_text = text[:obj_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    p.write_text(new_text, encoding='utf-8', newline='')
    print(f'  [{line_name}] stock={n_stock_updated} entries, cov_appended={cov_was_appended}, alerts={n_alerts_updated}, pres={n_pres_updated} ({p.stat().st_size:,} bytes)')


def main():
    if not PIVOT.is_file():
        print(f'ERROR: pivot no existe: {PIVOT}', file=sys.stderr); return 2

    print(f'Leyendo pivot: {PIVOT.name}')
    fam_data, prod_data, prod_to_fam, months = parse_pivot(PIVOT)
    print(f'  Meses: {months}')
    print(f'  Familias: {len(fam_data)}, productos: {len(prod_data)}')
    print()

    for line in LINES_DATAJS:
        update_line(line, f'{line}/data.js', True, fam_data, prod_data, prod_to_fam, 'Abr 26')
    for line, rel in LINES_INLINE.items():
        update_line(line, rel, False, fam_data, prod_data, prod_to_fam, 'Abr 26')

    return 0


if __name__ == '__main__':
    sys.exit(main())
