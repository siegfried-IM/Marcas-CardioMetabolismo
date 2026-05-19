"""Reconstruye respPerf.ACEMUK.atc.all en respiratorio/data.js con TODOS los
productos de ATC R05C0 (Expectorantes) desde el master IQVIA, NO solo los
8 que estaban antes.

ATC class destino: R05C0 - EXPECTORANTES
SIE products (mantienen flag is_sie=true): ACEMUK, ACEMUK VL.

Match por columna 'Product' del IQVIA — ya viene con sufijo (MANUF), p.ej.
'TOFLUX (CSO)', 'ATHOS (RMM)'.

Recompute por producto: monthly_vals, quarterly_vals, ytd[cierre], mat[cierre]
y ms_* (share vs total ATC R05C0).
"""
from __future__ import annotations
import json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA_FILE = REPO / 'respiratorio' / 'data.js'
MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\Ateneo Total - MAT Movil_May-19-2026 (3).xlsx')

TARGET_ATC = 'R05C0'  # Expectorantes
MARKET = 'ACEMUK'

# Productos SIE en R05C0 (preservar is_sie=true)
SIE_PRODUCTS = {'ACEMUK (SIE)', 'ACEMUK VL (SIE)'}

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM_TO_MES = {v:k for k,v in MES_INV.items()}


def msort(mk):
    p = mk.split()
    if len(p) != 2: return 0
    try: return int(p[1]) * 100 + MES_INV.get(p[0], 0)
    except: return 0


def quarter_key(mk):
    parts = mk.split()
    if len(parts) != 2: return ''
    m = MES_INV.get(parts[0])
    if not m: return ''
    q = (m - 1) // 3 + 1
    return f'Q{q} {parts[1]}'


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk:
            try: out[qk] += int(round(float(v or 0)))
            except: pass
    return dict(out)


def aggregate_ytd(monthly, cierre=4):
    out = defaultdict(int)
    cierre_lbl = NUM_TO_MES[cierre]
    for mk, v in monthly.items():
        parts = mk.split()
        if len(parts) != 2: continue
        m_num = MES_INV.get(parts[0])
        if not m_num: continue
        if m_num <= cierre:
            try: out[parts[1]] += int(round(float(v or 0)))
            except: pass
    return {f'{cierre_lbl} {y}': v for y, v in out.items()}


def aggregate_mat(monthly, cierre=4):
    cierre_lbl = NUM_TO_MES[cierre]
    years = set()
    for mk in monthly:
        parts = mk.split()
        if len(parts) == 2 and parts[0] in MES_INV:
            try: years.add(int(parts[1]))
            except: pass
    out = {}
    for y in sorted(years):
        total = 0
        for back in range(11, -1, -1):
            tot_idx = (y * 12 + (cierre - 1)) - back
            yy, mm = divmod(tot_idx, 12)
            mk = f'{NUM_TO_MES[mm + 1]} {yy}'
            v = monthly.get(mk)
            if v is not None:
                try: total += int(round(float(v or 0)))
                except: pass
        out[f'{cierre_lbl} {y}'] = total
    return out


def main():
    if not MASTER.is_file():
        print(f'ERROR: master not found at {MASTER}', file=sys.stderr); return 2

    print(f'Reading {MASTER.name}...')
    wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb.active
    hdrs = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    # cols
    col_mfr = col_prod = col_atc = col_mol = None
    month_cols = []  # (col_idx, 'Mes YYYY')
    for i, h in enumerate(hdrs):
        if not h: continue
        s = str(h).strip()
        s_norm = s.replace('\n', ' ').strip().lower()
        if s_norm.startswith('manufacturer'): col_mfr = i
        elif s_norm.startswith('product'): col_prod = i
        elif s_norm.startswith('atc'): col_atc = i
        elif s_norm.startswith('molecules'): col_mol = i
        if s.startswith('Units') and ('\n' in s or len(s.split()) >= 2):
            after = s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):].strip()
            after = after.strip()
            if after.upper().startswith('MAT') or after.upper().startswith('YTD'):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV:
                month_cols.append((i, f'{m.group(1)} {m.group(2)}'))

    if col_atc is None: col_atc = 2
    if col_prod is None: col_prod = 1
    if col_mfr is None: col_mfr = 0
    print(f'cols: mfr={col_mfr} prod={col_prod} atc={col_atc} mol={col_mol}')
    print(f'monthly columns: {len(month_cols)}, from {month_cols[0][1]} to {month_cols[-1][1]}')

    # Filter R05C0 rows + group by product
    by_prod = defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float)})
    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        atc = row[col_atc] if col_atc < len(row) else None
        prod = row[col_prod] if col_prod < len(row) else None
        if not atc or not prod: continue
        atc_str = str(atc)
        if TARGET_ATC not in atc_str: continue
        mfr = row[col_mfr] if col_mfr < len(row) else None
        bucket = by_prod[str(prod).strip()]
        if not bucket['manuf']:
            bucket['manuf'] = str(mfr or '').strip()
        for ci, mk in month_cols:
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try: bucket['monthly'][mk] += float(v)
            except: pass
        n_rows += 1
    wb.close()
    print(f'R05C0 rows: {n_rows}, unique products: {len(by_prod)}')

    # Build product list (Apr 2026 cierre)
    CIERRE = 4
    APR = 'Apr 2026'

    # Family-level aggregates
    fam_monthly = defaultdict(int)
    products_list = []
    for prod_name, info in by_prod.items():
        monthly = {mk: int(round(v)) for mk, v in info['monthly'].items()}
        for mk, v in monthly.items():
            fam_monthly[mk] += v
        is_sie = prod_name in SIE_PRODUCTS
        products_list.append({
            'prod': prod_name,
            'manuf': info['manuf'],
            'is_sie': is_sie,
            'monthly_vals': monthly,
            'quarterly_vals': aggregate_quarterly(monthly),
            'ytd': aggregate_ytd(monthly, CIERRE),
            'mat': aggregate_mat(monthly, CIERRE),
        })

    fam_monthly = dict(fam_monthly)
    fam_quarterly = aggregate_quarterly(fam_monthly)
    fam_ytd = aggregate_ytd(fam_monthly, CIERRE)
    fam_mat = aggregate_mat(fam_monthly, CIERRE)

    # ms_* per product
    def safe(num, den): return round((num or 0)/den*100, 2) if den else 0
    for p in products_list:
        mv = p['monthly_vals']; qv = p['quarterly_vals']; yv = p['ytd']; mtv = p['mat']
        p['ms_monthly']   = {mk: safe(mv.get(mk,0), fam_monthly.get(mk,0))   for mk in fam_monthly}
        p['ms_quarterly'] = {qk: safe(qv.get(qk,0), fam_quarterly.get(qk,0)) for qk in fam_quarterly}
        p['ms_ytd']       = {yk: safe(yv.get(yk,0), fam_ytd.get(yk,0))       for yk in fam_ytd}
        p['ms_mat']       = {mk: safe(mtv.get(mk,0), fam_mat.get(mk,0))      for mk in fam_mat}

    # Sort: SIE first, then by Apr 2026 units desc
    def sort_key(p):
        apr = (p.get('monthly_vals') or {}).get(APR, 0) or 0
        return (not p.get('is_sie'), -apr)
    products_list.sort(key=sort_key)

    n_sie = sum(1 for p in products_list if p.get('is_sie'))
    apr_top = sum((p['monthly_vals'].get(APR,0) or 0) for p in products_list[:5])
    print(f'Total products: {len(products_list)}, SIE: {n_sie}')
    print(f'Top 5 SIE+rank by Apr 2026:')
    for p in products_list[:5]:
        apr = p['monthly_vals'].get(APR, 0)
        ms_apr = p['ms_monthly'].get(APR, 0)
        print(f'  {p["prod"]}: Apr 2026={apr:,}u, MS%={ms_apr}%')

    # Patch into respiratorio/data.js
    t = DATA_FILE.read_text(encoding='utf-8-sig')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    D, end_idx = json.JSONDecoder().raw_decode(t[ob:])
    abs_end = ob + end_idx

    # Replace respPerf.ACEMUK.atc.all
    rp = D.setdefault('respPerf', {}).setdefault(MARKET, {}).setdefault('atc', {})
    new_all = {
        'family': MARKET,
        'products': products_list,
        'monthly': fam_monthly,
        'quarterly': fam_quarterly,
        'ytd': fam_ytd,
        'mat': fam_mat,
    }
    rp['all'] = new_all

    new_t = t[:ob] + json.dumps(D, ensure_ascii=False) + t[abs_end:]
    DATA_FILE.write_text(new_t, encoding='utf-8', newline='')
    print(f'\nrespiratorio/data.js written ({len(new_t):,} bytes)')


if __name__ == '__main__':
    sys.exit(main() or 0)
