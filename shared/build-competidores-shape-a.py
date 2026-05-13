#!/usr/bin/env python3
"""
shared/build-competidores-shape-a.py

Generaliza el builder Shape A de competidores DDD para cualquier linea.
Lee un xlsx en formato 'Producto-Mol-ATC-provincia' (cols:
RegionCUP, Mercado, Droga, Clase Terapeutica, AñoMes, Codigo Clase,
Codigo Producto, Producto, Unidades) y produce un competidores-data.js
con shape A (brand_monthly[brand][region][N months]).

Aplica a:
  - ATB       <- Hub-Marcas-Inputs/ATB/2026-04/fuentes-originales/DDD ATB.xlsx
  - respi     <- Hub-Marcas-Inputs/respiratorio/2026-04/ddd/Producto-Mol...xlsx
  - mujer     <- ya hecho separadamente (build-mujer-competidores-data.py)
  - OTC, cardio, SNC, dermato: NO aplica (cardio/SNC/dermato ya son
    Shape A nativos; OTC no tiene xlsx fuente)
"""
from __future__ import annotations
import re, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
INPUTS = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')

# (line, xlsx_path, out_data_js)
LINES = [
    ('ATB',          INPUTS / 'ATB/2026-04/fuentes-originales/DDD ATB.xlsx',                                  REPO / 'ATB/DDD/competidores-data.js'),
    ('respiratorio', INPUTS / 'respiratorio/2026-04/fuentes-originales/Producto-Molécula-ATC-provincia - 13 de mayo de 2026 (1).xlsx',
                                                                                                              REPO / 'respiratorio/DDD/competidores-data.js'),
    ('OTC',          INPUTS / 'OTC/2026-04/fuentes-originales/Producto-Molécula-ATC-provincia - 12 de mayo de 2026.xlsx',
                                                                                                              REPO / 'OTC/DDD/competidores-data.js'),
    ('dermato',      INPUTS / 'dermato/2026-04/fuentes-originales/Producto-Molécula-ATC-provincia - 12 de mayo de 2026 (2).xlsx',
                                                                                                              REPO / 'dermatologia/competidores-data.js'),
    ('cardio',       INPUTS / 'cardio/2026-04/fuentes-originales/Producto-Molécula-ATC-provincia - 12 de mayo de 2026 (4).xlsx',
                                                                                                              REPO / 'cardio/DDD/competidores-data.js'),
    ('mujer',        INPUTS / 'linea-mujer/2026-04/fuentes-originales/Producto-Molécula-ATC-provincia - 13 de mayo de 2026.xlsx',
                                                                                                              REPO / 'mujer/DDD/competidores-data.js'),
    ('SNC',          INPUTS / 'PSQ/2026-04/fuentes-originales/Producto-Molécula-ATC-provincia - 13 de mayo de 2026 (3).xlsx',
                                                                                                              REPO / 'SNC/DDD/competidores-data.js'),
]

MONTH_ORDER = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
               'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

PACKAGE_TOKENS = {
    'TABL','TABLE','TABLET','TABLETA','TABLETAS',
    'COMP','COMPR','COMPS','COMPRIMIDO','COMPRIMIDOS',
    'CAP','CAPS','CAPSULA','CAPSULAS',
    'AMP','AMPOLLA','AMPOLLAS',
    'JBE','JARABE',
    'INY','INYECTABLE','V.IM','V.IV',
    'SUSP','SUSPENSION',
    'POL','POLVO',
    'LIQ','LIQUIDO',
    'GTAS','GOTAS',
    'SOL','SOLUC','SOLUCION','SOLUCIONES',
    'CR','CREMA','CREMAS',
    'UNG','UNGUENTO',
    'SOB','SOBRE','SOBRES',
    'BOLS','BOLSA',
    'EFER','EFERV','EFERVES','EFERVESCENTE','EFERVESCENTES',
    'DESLEI','DESLEIBLES','DESLEÍBLES',
    'REC','RECUB','RECUBIE','RECUBIERTO','RECUBIERTOS','RECUBIERTAS',
    'IM','PED','AD','ADULT','INF',
    'SPRAY','PVO','PLV',
    'DISP','DISPERS','DISPERSABLE','DISPERSABLES',
    'MAST','GRAG','GRAGEAS',
    'INHAL','INHALACION',
    'OFT','GINEC','VAGINAL','OVUL','OVULO','OVULOS','PESARIO',
    'GEL','EMUL','LOC','LOCION',
    'COLIRIO','COLIR','PARCHE','APOSITO',
    'AERO','AEROSOL',
    'ANILLO','ANILLOS',
}


def extract_brand(producto: str) -> str:
    if not producto: return 'UNKNOWN'
    s = str(producto).strip().upper().replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s)
    parts = s.split(' ')
    brand = []
    for i, p in enumerate(parts):
        clean = re.sub(r'^[.,;:()\[\]/\\-]+|[.,;:()\[\]/\\-]+$', '', p)
        if not clean: continue
        if clean == 'X' and i > 0: break
        if clean in PACKAGE_TOKENS: break
        if re.match(r'^[\d.,]+([A-Z%]+)?$', clean):
            if len(brand) >= 1 and re.match(r'^\d{1,2}$', clean):
                brand.append(clean); continue
            break
        if re.match(r'^\d+(\.\d+)?(MG|ML|G|MCG|UI|%).*', clean): break
        brand.append(clean)
    return ' '.join(brand) if brand else (parts[0] if parts else 'UNKNOWN')


# SIE detection patterns per linea (regex case-insensitive)
SIE_PATTERNS_BY_LINE = {
    'ATB': [r'^ACANTEX\b', r'^BACTRIM\b', r'^CEFALEXINA ARG', r'^MACROMAX\b'],
    'respiratorio': [r'^ACEMUK\b', r'^AIREAL\b', r'^ALIDIAL\b', r'^DECADRON\b',
                     r'^DUO-DECADRON\b', r'^HEXALER\b'],
    'OTC': [r'^ACERPES\b', r'^ACI-TIP\b', r'^ALUMPAK\b', r'^ARTRO RED\b',
            r'^FLEXINA\b', r'^MAGNUS\b', r'^TETRALGIN\b'],
    'dermato': [r'^ACNECLIN\b', r'^CLOBESOL\b', r'^MICOMAZOL\b',
                r'^MICROSONA\b', r'^MOMETAX\b', r'^PALDAR\b', r'^ROACCUTAN\b'],
    'cardio': [r'^DAURAN\b', r'^DILATREND\b', r'^DIOVAN\b', r'^EMPAX\b',
               r'^ENTRESTO\b', r'^EXFORGE\b', r'^METGLUCON\b', r'^PIXABAN\b',
               r'^ROXOLAN\b', r'^SILTRAN\b', r'^SINTROM\b', r'^TELPRES\b',
               r'^TERLOC\b'],
    'mujer': [r'^ISIS\b', r'^SIDERBLUT\b', r'^SIDER\b', r'^TRIP\b',
              r'^CALCIO BASE\b', r'^CALCIO CITRATO\b', r'^CLIMATIX\b',
              r'^DELTROX\b', r'^GYNODERM\b', r'^ROXOLAN\b', r'^ALUMPAK\b'],
    'SNC': [r'^VALIUM\b', r'^MADOPAR\b', r'^QTP\b', r'^PGB\b', r'^EMERAL\b',
            r'^LURAP\b', r'^VALQUIR\b', r'^MELERIL\b', r'^LEVITAL\b', r'^VISDON\b'],
}


def is_sie_brand(brand: str, line: str) -> bool:
    pats = SIE_PATTERNS_BY_LINE.get(line, [])
    s = str(brand).upper()
    return any(re.match(p, s) for p in pats)


def month_sort_key(mk: str) -> int:
    parts = mk.split('-')
    return int(parts[1]) * 100 + MONTH_ORDER.get(parts[0], 0)


def build_one(line: str, xlsx: Path, out: Path) -> str:
    if not xlsx.is_file():
        return f'  [{line}] SKIP: xlsx no existe ({xlsx})'

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {str(h or '').strip(): i for i, h in enumerate(hdr)}
    c_region = cols.get('RegionCUP', 0)
    c_market = cols.get('Mercado', 1)
    c_mes    = cols.get('AñoMes', 4)
    c_prod   = cols.get('Producto', 7)
    c_unid   = cols.get('Unidades', 8)

    months_set = set()
    regions_set = set()
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    mkt_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    sie_flag = defaultdict(dict)

    n_rows = n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        if not row: continue
        region = row[c_region]
        market = row[c_market]
        mes    = row[c_mes]
        prod   = row[c_prod]
        unid   = row[c_unid]
        if not region or not market or not mes or not prod: continue
        region = str(region).strip()
        if region in ('Totales', '-'): continue
        market = str(market).strip()
        mes = str(mes).strip()
        try: u = int(round(float(unid or 0)))
        except (TypeError, ValueError): u = 0
        if u <= 0: continue
        brand = extract_brand(prod)
        if not brand or brand == 'UNKNOWN': continue
        data[market][brand][region][mes] += u
        mkt_totals[market][region][mes] += u
        if brand not in sie_flag[market]:
            sie_flag[market][brand] = is_sie_brand(brand, line)
        regions_set.add(region)
        months_set.add(mes)
        n_kept += 1
    wb.close()

    months = sorted(months_set, key=month_sort_key)
    regions = sorted(regions_set, key=lambda r: (r.startswith('_'), r))

    out_obj = {
        'months': months,
        'regions': regions,
        'markets': {},
    }
    for market in sorted(data.keys()):
        bm = {}
        for brand in sorted(data[market].keys()):
            bm[brand] = {}
            for region in regions:
                arr = [data[market][brand].get(region, {}).get(mk, 0) for mk in months]
                if any(arr):
                    bm[brand][region] = arr
        tm = {}
        for region in regions:
            arr = [mkt_totals[market].get(region, {}).get(mk, 0) for mk in months]
            if any(arr):
                tm[region] = arr
        sie_brands = sorted([b for b, is_s in sie_flag[market].items() if is_s])
        out_obj['markets'][market] = {
            'brands': sie_brands,
            'brand_monthly': bm,
            'total_monthly': tm,
        }

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        f'window.SFG_COMP_DATA = {json.dumps(out_obj, ensure_ascii=False)};\n',
        encoding='utf-8', newline=''
    )

    sie_total = sum(1 for mk in out_obj['markets'].values() for b in mk['brand_monthly'] if mk['brands'])
    return (f'  [{line}] OK rows={n_rows:,} kept={n_kept:,} '
            f'months={len(months)} ({months[0]}..{months[-1]}) '
            f'regions={len(regions)} markets={len(data)} '
            f'-> {out.name} ({out.stat().st_size:,} bytes)')


def patch_html_loader(line: str, html_path: Path):
    """Make the competidores.html load competidores-data.js instead of ../data.js."""
    if not html_path.is_file():
        return f'  [{line}/html] SKIP (no existe)'
    text = html_path.read_text(encoding='utf-8', errors='replace')
    if '<script src="./competidores-data.js"></script>' in text:
        return f'  [{line}/html] ya carga competidores-data.js'
    new_text = text.replace('<script src="../data.js"></script>',
                             '<script src="./competidores-data.js"></script>', 1)
    if new_text == text:
        return f'  [{line}/html] WARN: no match for ../data.js'
    html_path.write_text(new_text, encoding='utf-8', newline='')
    return f'  [{line}/html] OK (loader updated)'


def main():
    print('Build Shape A competidores-data.js...\n')
    for line, xlsx, out in LINES:
        print(build_one(line, xlsx, out))

    print('\nPatch competidores.html loaders...\n')
    for line, _, out in LINES:
        html = out.parent / 'competidores.html'
        print(patch_html_loader(line, html))

    print('\nListo.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
