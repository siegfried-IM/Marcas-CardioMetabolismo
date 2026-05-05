#!/usr/bin/env python3
"""
shared/merge-recetas-respi.py

Mergea data de recetas de RESPI desde el pivot CloseUp completo
(15 meses Ene 2025 - Mar 2026, 10 mercados):

  ANTIGRIPALES (ACEMUK DN), ANTIHISTA SIST (HEXALER),
  CETIRIZINA (ALIDIAL), CORTICO. NASAL (H NASAL),
  CORTICOST ASOC (H CORT), DECADRON, DECADRON INY,
  EXPECTORANTES (ACEMUK), MONTE LEVO (AIREAL PLUS),
  MONTELUKAST (AIREAL).

Para respiratorio/data.js (window.OTC_DASHBOARD):
  - rec_ms[fam]:       sie (monthly), ms (monthly), ms_quarterly
  - rec_comp[fam][brand]: is_sie, monthly/quarterly/total +
                           monthly/quarterly/total_medicos
  - recetas[fam]:      total mercado (recetas + medicos por mes)

Uso:
    py shared/merge-recetas-respi.py [--pivot <path>] [--dry-run]
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA_JS = REPO / 'respiratorio' / 'data.js'
DEFAULT_PIVOT = Path(r'C:\Users\camarinaro\Downloads\Sin título - Tabla dinámica - 4 de mayo de 2026 (4).xlsx')

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
MES_INV = {v:k for k,v in MES_EN.items()}

# Market name (col A en pivot) -> family key en respi data
MARKET_KEY_MAP = {
    'ANTIGRIPALES (ACEMUK DN)': 'ACEMUK DIA Y NOCHE',
    'ANTIHISTA SIST (HEXALER)': 'ANTIHISTA SIST',
    'CETIRIZINA (ALIDIAL)':     'ALIDIAL',
    'CORTICO. NASAL (H NASAL)': 'HEXALER NASAL',
    'CORTICOST ASOC (H CORT)':  'HEXALER CORT',
    'DECADRON':                 'DECADRON',
    'DECADRON INY':             'DECADRON INY',
    'EXPECTORANTES (ACEMUK)':   'ACEMUK',
    'MONTE LEVO (AIREAL PLUS)': 'AIREAL PLUS',
    'MONTELUKAST (AIREAL)':     'AIREAL',
}


def msort(mk):
    p = mk.split()
    if len(p) != 2: return 0
    return int(p[1]) * 100 + MES_INV.get(p[0], 0)


def quarter_key(mk):
    parts = mk.split()
    if len(parts) != 2: return ''
    m = MES_INV.get(parts[0])
    if not m: return ''
    return f'Q{(m-1)//3+1} {parts[1]}'


def to_int(v):
    if v is None or v == '-' or v == '': return 0
    try: return int(round(float(v)))
    except: return 0


def parse_pivot(pivot_path):
    """Devuelve dict: market -> {months_seen, brands: [{droga, marca, monthly_rec, monthly_med}], totales: monthly_rec/med}"""
    wb = openpyxl.load_workbook(pivot_path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    # Build col map: idx -> (month_key, metric)
    col_map = {}
    months = []
    for i, h1 in enumerate(row1):
        if not isinstance(h1, datetime): continue
        mk = f'{MES_EN[h1.month]} {h1.year}'
        if mk not in months: months.append(mk)
        h2 = (str(row2[i]) if i < len(row2) and row2[i] else '').strip().lower()
        if 'recetas' in h2:
            col_map[i] = (mk, 'rec')
        elif 'm' in h2 and ('dico' in h2 or 'edico' in h2):
            col_map[i] = (mk, 'med')

    markets = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3: continue
        merc = (row[0] or '').strip()
        droga = (row[1] or '').strip()
        marca = (row[2] or '').strip()
        if not merc: continue

        mk_data = markets.setdefault(merc, {
            'totales': {'rec': {}, 'med': {}},
            'brands': [],
        })

        # Three row types:
        # 1) (merc, 'Totales', '') = market totals
        # 2) (merc, droga, 'Totales') = droga totals (skip)
        # 3) (merc, droga, brand) = brand row
        if droga == 'Totales' and not marca:
            for ci, (mkk, metric) in col_map.items():
                if ci >= len(row): continue
                v = to_int(row[ci])
                mk_data['totales'][metric][mkk] = v
            continue
        if marca == 'Totales':
            continue
        if not marca: continue

        brand_obj = {'droga': droga, 'marca': marca,
                     'rec': {}, 'med': {}}
        for ci, (mkk, metric) in col_map.items():
            if ci >= len(row): continue
            v = to_int(row[ci])
            brand_obj[metric][mkk] = v
        mk_data['brands'].append(brand_obj)

    wb.close()
    return months, markets


def is_sie_brand(marca):
    return marca.upper().endswith(' SIE')


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk: out[qk] += v
    return dict(out)


def merge_into_dict(existing, new_dict):
    """Merge new month entries into existing dict (overwrite if conflict)."""
    if not isinstance(existing, dict): existing = {}
    for k, v in new_dict.items():
        existing[k] = v
    return existing


def load_data_js(p):
    text = p.read_text(encoding='utf-8-sig', errors='replace')
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text)
    if not m1: raise ValueError('OTC_DATA not found')
    obj_start1 = text.index('{', m1.end())
    d1, end1 = json.JSONDecoder().raw_decode(text[obj_start1:])
    abs_end1 = obj_start1 + end1
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text[abs_end1:])
    if not m2: return text, d1, None, abs_end1, None
    obj_start2 = abs_end1 + text[abs_end1:].index('{', m2.end())
    d2, end2 = json.JSONDecoder().raw_decode(text[obj_start2:])
    abs_end2 = obj_start2 + end2
    return text, d1, d2, abs_end1, (obj_start1, abs_end1, obj_start2, abs_end2)


def write_data_js(text_orig, d1, d2, bounds):
    obj_start1, abs_end1, obj_start2, abs_end2 = bounds
    prefix = text_orig[:obj_start1]
    middle = text_orig[abs_end1:obj_start2]
    suffix = text_orig[abs_end2:]
    return prefix + json.dumps(d1, ensure_ascii=False) + middle + json.dumps(d2, ensure_ascii=False) + suffix


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', default=str(DEFAULT_PIVOT))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pp = Path(args.pivot)
    if not pp.is_file():
        print(f'ERROR: pivot no existe: {pp}', file=sys.stderr); return 2

    print(f'Leyendo: {pp}')
    months, markets = parse_pivot(pp)
    months_sorted = sorted(months, key=msort)
    print(f'  {len(months_sorted)} meses ({months_sorted[0]} .. {months_sorted[-1]})')
    print(f'  {len(markets)} mercados encontrados:')
    for m in markets: print(f'    - {m}'.encode('ascii','replace').decode())

    # Verify mapping
    unmapped = [m for m in markets if m not in MARKET_KEY_MAP]
    if unmapped:
        print(f'  WARN: mercados sin mapeo: {unmapped}'.encode('ascii','replace').decode())

    # Load respi data.js
    text_orig, d1, d2, _, bounds = load_data_js(DATA_JS)
    if d2 is None:
        print('ERROR: respi no tiene OTC_DASHBOARD'); return 3

    rec_ms   = d2.setdefault('rec_ms', {})
    rec_comp = d2.setdefault('rec_comp', {})
    recetas  = d2.setdefault('recetas', {})

    n_added_keys = 0
    n_updated_keys = 0

    for market_name, mkt_data in markets.items():
        fam_key = MARKET_KEY_MAP.get(market_name)
        if not fam_key:
            print(f'  SKIP {market_name} (sin mapeo)')
            continue

        # 1) recetas[fam] = total mercado por mes
        recetas_obj = recetas.setdefault(fam_key, {})
        if recetas_obj == {}: n_added_keys += 1
        else: n_updated_keys += 1
        for mk in months_sorted:
            r = mkt_data['totales']['rec'].get(mk)
            md = mkt_data['totales']['med'].get(mk)
            if r is None and md is None: continue
            recetas_obj[mk] = {'recetas': r or 0, 'medicos': md or 0}

        # 2) rec_ms[fam].sie = suma de SIE rows por mes
        sie_monthly = defaultdict(int)
        sie_brands_in_mkt = []
        for b in mkt_data['brands']:
            if is_sie_brand(b['marca']):
                sie_brands_in_mkt.append(b['marca'])
                for mk, v in b['rec'].items():
                    sie_monthly[mk] += v

        rms_obj = rec_ms.setdefault(fam_key, {'sie': {}, 'ms': {}, 'ms_quarterly': {}})
        sie_dict = rms_obj.setdefault('sie', {})
        ms_dict  = rms_obj.setdefault('ms', {})
        msq_dict = rms_obj.setdefault('ms_quarterly', {})

        for mk in months_sorted:
            sie_v = sie_monthly.get(mk, 0)
            sie_dict[mk] = sie_v
            tot = mkt_data['totales']['rec'].get(mk, 0) or 0
            ms_dict[mk] = round(sie_v / tot * 100, 2) if tot > 0 else 0

        # ms_quarterly: agregar por cuatrimestre desde sie y mercado total
        sie_q = aggregate_quarterly({mk: sie_dict[mk] for mk in months_sorted if mk in sie_dict})
        tot_q = aggregate_quarterly({mk: mkt_data['totales']['rec'].get(mk, 0) for mk in months_sorted})
        for qk, sv in sie_q.items():
            tv = tot_q.get(qk, 0)
            msq_dict[qk] = round(sv / tv * 100, 2) if tv > 0 else 0

        # 3) rec_comp[fam][brand] = todos los brands del mercado
        rc_obj = rec_comp.setdefault(fam_key, {})
        # Reemplazamos contenido completo del fam (datos del pivot son mas completos)
        rc_new = {}
        for b in mkt_data['brands']:
            brand_name = b['marca']
            monthly_rec = {mk: v for mk, v in b['rec'].items() if v}
            monthly_med = {mk: v for mk, v in b['med'].items() if v}
            quarterly = aggregate_quarterly(monthly_rec)
            quarterly_med = aggregate_quarterly(monthly_med)
            rc_new[brand_name] = {
                'is_sie': is_sie_brand(brand_name),
                'monthly':           monthly_rec,
                'quarterly':         quarterly,
                'total':             sum(monthly_rec.values()),
                'monthly_medicos':   monthly_med,
                'quarterly_medicos': quarterly_med,
                'total_medicos':     sum(monthly_med.values()),
            }
        # Sort: SIE primero, luego por total recetas desc
        rc_new = dict(sorted(rc_new.items(),
                              key=lambda kv: (not kv[1]['is_sie'], -kv[1]['total'])))
        rec_comp[fam_key] = rc_new

        print(f'  [{fam_key}] OK: {len(rc_new)} brands ({len(sie_brands_in_mkt)} SIE: {sie_brands_in_mkt})'.encode('ascii','replace').decode())

    # Update meta
    meta = d2.setdefault('meta', {})
    meta['rxCut'] = months_sorted[-1]

    if args.dry_run:
        print('\nDRY RUN: no se escribio.')
        return 0

    new_text = write_data_js(text_orig, d1, d2, bounds)
    DATA_JS.write_text(new_text, encoding='utf-8', newline='')
    print(f'\n-> {DATA_JS} reescrito ({DATA_JS.stat().st_size:,} bytes)')
    print(f'  rec_ms keys ahora: {sorted(rec_ms.keys())}'.encode('ascii','replace').decode())
    return 0


if __name__ == '__main__':
    sys.exit(main())
