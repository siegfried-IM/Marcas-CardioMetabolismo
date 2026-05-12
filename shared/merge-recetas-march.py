#!/usr/bin/env python3
"""
shared/merge-recetas-march.py

Merge ad-hoc para agregar UN mes de recetas (sin re-correr build-data.ps1)
a partir del pivot de CloseUp.

Input: el pivot Excel "Sin titulo - Tabla dinamica - <fecha>.xlsx"
       con columnas [Mercado, Droga, Marca, <fecha>] y valores de recetas.

Output: cada <linea>/data.js modificado in-place agregando el mes nuevo en:
  - OTC_DATA.prescriptions.months -> append "<MMM-YYYY>"
  - OTC_DATA.prescriptions.families[F].prescriptions -> append family total
  - OTC_DATA.prescriptions.families[F].doctors -> append 0 (no en pivot)
  - OTC_DATA.prescriptions.families[F].latestMonth -> "<MMM-YYYY>"
  - OTC_DATA.meta.rxCut -> "<MMM-YYYY>"
  - OTC_DASHBOARD.recetas[F]["<MMM YYYY>"] = {recetas, medicos:0}
  - OTC_DASHBOARD.rec_ms[F].sie/ms["<MMM YYYY>"] = ...
  - OTC_DASHBOARD.rec_comp[F][P].monthly["<MMM YYYY>"] = product_count
  - OTC_DASHBOARD.meta.rec_label = "<MMM>'YY"

Lineas soportadas: cardio, ATB, OTC, respiratorio (estructura OTC_DATA + OTC_DASHBOARD).
NO toca mujer/SNC (estructura distinta, requiere su propio extractor).

Uso:
    py shared/merge-recetas-march.py \\
        --pivot "C:\\Users\\.../Sin titulo - Tabla dinamica - 30 de abril de 2026.xlsx" \\
        --month "2026-03" \\
        [--lines cardio ATB OTC respiratorio]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

LINES_DEFAULT = ['cardio', 'ATB', 'OTC', 'respiratorio']

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
MES_ES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}


def normalize_brand(name: str | None) -> str:
    if not name: return ''
    s = str(name).upper().strip()
    # Quitar puntos y espacios extras pero mantener forma
    s = re.sub(r'\s+', ' ', s)
    return s


# Mercados que son roll-ups de SIE u otros agregados; sus productos aparecen
# tambien en el mercado especifico, asi que incluirlos genera double-counting.
EXCLUDED_MARKETS = {
    'MIX SIEGFRIED',
}


def load_pivot(pivot_path: Path):
    """Auto-detecta columnas de meses (Cant. Recetas + opcional Cant. Medicos)
    y devuelve para CADA mes detectado:

       - months: lista ordenada de month_key_en ej. ["Jan 2026","Feb 2026","Mar 2026"]
       - brand_recetas[month][brand_normalized] = max count visto en mercados validos
       - brand_medicos[month][brand_normalized] = max count visto

    Si el pivot solo tiene 1 mes (formato viejo sin medicos), funciona igual.
    """
    wb = openpyxl.load_workbook(pivot_path, read_only=True, data_only=True)
    ws = wb.active

    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    # Detectar columnas: cada fecha en row1 puede tener subheader "Cant. Recetas"
    # o "Cant. Medicos" en row2. Si row2 es vacio (formato viejo), tratamos esa
    # columna como recetas.
    col_map = {}  # col_idx -> (month_key_en, 'recetas'|'medicos')
    cur_month = None
    months = []
    for i, h1 in enumerate(row1):
        if isinstance(h1, datetime):
            cur_month = f"{MES_EN[h1.month]} {h1.year}"
            if cur_month not in months:
                months.append(cur_month)
        sub = (str(row2[i]) if i < len(row2) and row2[i] else '').lower()
        if not cur_month:
            continue
        if 'medico' in sub or 'médico' in sub:
            col_map[i] = (cur_month, 'medicos')
        elif 'receta' in sub:
            col_map[i] = (cur_month, 'recetas')
        elif sub == '' and i >= 3 and i not in col_map:
            # Formato viejo: una sola columna por mes, sin subheaders
            col_map[i] = (cur_month, 'recetas')

    brand_recetas = {m: {} for m in months}
    brand_medicos = {m: {} for m in months}
    # Family-level totals from the (market, 'Totales', '') row — UNIQUE doctor count
    fam_recetas  = {m: {} for m in months}  # fam_recetas[month][market] = recetas
    fam_medicos  = {m: {} for m in months}  # fam_medicos[month][market] = medicos UNIQUE

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row: continue
        merc = row[0] if len(row) > 0 else None
        droga = row[1] if len(row) > 1 else None
        marca = row[2] if len(row) > 2 else None
        if not merc: continue
        cur_market = str(merc).strip()
        if cur_market.upper() in EXCLUDED_MARKETS:
            continue
        # (market, 'Totales', '') -> MARKET total (medicos unique)
        is_market_total = (str(droga or '').strip().lower() == 'totales' and not marca)
        if is_market_total:
            for col_idx, (month_key, kind) in col_map.items():
                if col_idx >= len(row): continue
                val = row[col_idx]
                try: v = int(val) if val is not None else 0
                except (TypeError, ValueError): v = 0
                target = fam_recetas if kind == 'recetas' else fam_medicos
                # Solo guardar el primero (no agregar drogas internas que tambien tienen Totales)
                if cur_market not in target[month_key]:
                    target[month_key][cur_market] = v
            continue
        # Skip droga-level "Totales" and rows without marca
        if (str(marca or '').strip().lower() == 'totales' or not marca):
            continue
        b = normalize_brand(marca)

        for col_idx, (month_key, kind) in col_map.items():
            if col_idx >= len(row): continue
            val = row[col_idx]
            try:
                v = int(val) if val is not None else 0
            except (TypeError, ValueError):
                v = 0
            target = brand_recetas if kind == 'recetas' else brand_medicos
            prev = target[month_key].get(b, 0)
            if v > prev:
                target[month_key][b] = v

    wb.close()
    return months, brand_recetas, brand_medicos, fam_recetas, fam_medicos


def parse_data_js(text: str):
    """Extrae los dos objetos window.OTC_DATA = {...}; window.OTC_DASHBOARD = {...};
    Devuelve (otc_data, otc_dashboard, prefix1, suffix1, prefix2, suffix2) para
    poder reconstruir el archivo despues."""
    # Find first window.X = {...};
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text)
    if not m1:
        raise ValueError('OTC_DATA not found')
    obj_start1 = text.index('{', m1.end())
    d1, end1 = json.JSONDecoder().raw_decode(text[obj_start1:])
    abs_end1 = obj_start1 + end1
    # The semicolon and newlines after
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text[abs_end1:])
    if not m2:
        return d1, None, text[:obj_start1], text[abs_end1:], None, None
    obj_start2 = abs_end1 + text[abs_end1:].index('{', m2.end())
    d2, end2 = json.JSONDecoder().raw_decode(text[obj_start2:])
    abs_end2 = obj_start2 + end2

    prefix1 = text[:obj_start1]
    middle = text[abs_end1:obj_start2]
    suffix2 = text[abs_end2:]
    return d1, d2, prefix1, middle, obj_start2, suffix2


def serialize_data_js(text_orig: str, d1: dict, d2: dict | None) -> str:
    """Reconstruye data.js preservando prefijos y sufijos textuales."""
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text_orig)
    obj_start1 = text_orig.index('{', m1.end())
    _, end1 = json.JSONDecoder().raw_decode(text_orig[obj_start1:])
    abs_end1 = obj_start1 + end1
    prefix1 = text_orig[:obj_start1]

    if d2 is None:
        return prefix1 + json.dumps(d1, ensure_ascii=False) + text_orig[abs_end1:]

    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text_orig[abs_end1:])
    obj_start2 = abs_end1 + text_orig[abs_end1:].index('{', m2.end())
    _, end2 = json.JSONDecoder().raw_decode(text_orig[obj_start2:])
    abs_end2 = obj_start2 + end2

    middle = text_orig[abs_end1:obj_start2]
    suffix = text_orig[abs_end2:]
    return (prefix1
            + json.dumps(d1, ensure_ascii=False)
            + middle
            + json.dumps(d2, ensure_ascii=False)
            + suffix)


def merge_line(data_js_path: Path, months_to_merge: list,
               brand_recetas: dict, brand_medicos: dict,
               fam_recetas: dict = None, fam_medicos: dict = None,
               dry_run: bool = False):
    """Patcha data.js de UNA linea agregando UNO O MAS meses de recetas."""
    text = data_js_path.read_text(encoding='utf-8-sig', errors='replace')
    d1, d2, *_ = parse_data_js(text)
    if d2 is None:
        return {'skipped': 'no OTC_DASHBOARD'}

    rec_comp = d2.get('rec_comp') or {}
    if not rec_comp:
        return {'skipped': 'no rec_comp'}

    families = list(rec_comp.keys())
    fam_summary = {f: {'months': {}} for f in families}

    last_month_label_es = None
    last_month_key_short = None

    for month_key_en in months_to_merge:
        # Convertir month_key_en ("Mar 2026") a labels en espanol
        en_mon, year = month_key_en.split()
        mo_idx = list(MES_EN.values()).index(en_mon) + 1
        month_label_es = f"{MES_ES[mo_idx]}-{year}"
        month_key_short = f"{en_mon}'{str(year)[-2:]}"
        last_month_label_es = month_label_es
        last_month_key_short = month_key_short

        recetas_lookup = brand_recetas.get(month_key_en, {})
        medicos_lookup = brand_medicos.get(month_key_en, {})

        for fam in families:
            comp = rec_comp[fam]
            sie_total = 0
            sie_med_total = 0
            market_total = 0
            market_med_total = 0
            unmatched = []
            for prod_name, prod_data in comp.items():
                if not isinstance(prod_data, dict): continue
                monthly = prod_data.setdefault('monthly', {})
                b = normalize_brand(prod_name)
                count = recetas_lookup.get(b, 0)
                medicos = medicos_lookup.get(b, 0)
                if count == 0 and b not in recetas_lookup:
                    unmatched.append(prod_name)
                monthly[month_key_en] = count
                if isinstance(prod_data.get('total'), (int, float)):
                    prod_data['total'] = (prod_data['total'] or 0) + count
                market_total += count
                market_med_total += medicos
                if b.endswith(' SIE') or b.endswith('(SIE)'):
                    sie_total += count
                    sie_med_total += medicos

            # rec_ms
            rms = d2.setdefault('rec_ms', {}).setdefault(fam, {})
            sie_dict = rms.setdefault('sie', {})
            ms_dict = rms.setdefault('ms', {})
            mkt_dict = rms.setdefault('mkt', {})
            sie_dict[month_key_en] = sie_total
            mkt_dict[month_key_en] = market_total
            ms_pct = round((sie_total / market_total) * 100, 1) if market_total else 0
            ms_dict[month_key_en] = ms_pct

            # recetas (per family monthly aggregate).
            # Medicos: usar el count UNICO del pivot row (mercado, 'Totales', '')
            # — NO la suma de medicos por marca (CloseUp NO es sumable: un mismo
            # medico que prescribe varias marcas se contaria multiple veces).
            rec_fam = d2.setdefault('recetas', {}).setdefault(fam, {})
            uniq_med = None
            uniq_rec = None
            if fam_medicos is not None:
                uniq_med = (fam_medicos.get(month_key_en) or {}).get(fam)
            if fam_recetas is not None:
                uniq_rec = (fam_recetas.get(month_key_en) or {}).get(fam)
            rec_fam[month_key_en] = {
                'recetas': uniq_rec if uniq_rec is not None else market_total,
                'medicos': uniq_med if uniq_med is not None else market_med_total,
            }

            # OTC_DATA.prescriptions: upsert (no append duplicado si el mes ya existe)
            pres_root = d1.setdefault('prescriptions', {})
            months_arr = pres_root.setdefault('months', [])
            if month_label_es in months_arr:
                idx = months_arr.index(month_label_es)
            else:
                months_arr.append(month_label_es)
                idx = len(months_arr) - 1
            pres = pres_root.setdefault('families', {}).setdefault(fam, {
                'prescriptions': [], 'doctors': [], 'latestMonth': '', 'topBrands': []
            })
            # Pad arrays si quedaron cortos
            while len(pres.get('prescriptions', [])) < idx:
                pres.setdefault('prescriptions', []).append(0)
            while len(pres.get('doctors', [])) < idx:
                pres.setdefault('doctors', []).append(0)
            # Upsert en idx
            pr_arr = pres.setdefault('prescriptions', [])
            dr_arr = pres.setdefault('doctors', [])
            if idx < len(pr_arr):
                pr_arr[idx] = market_total
            else:
                pr_arr.append(market_total)
            if idx < len(dr_arr):
                dr_arr[idx] = market_med_total
            else:
                dr_arr.append(market_med_total)
            pres['latestMonth'] = month_label_es

            fam_summary[fam]['months'][month_key_en] = {
                'sie': sie_total, 'mkt': market_total, 'ms': ms_pct,
                'medicos': market_med_total,
                'unmatched_count': len(unmatched)
            }

    # Pad final: aseguramos que prescriptions.families[F].prescriptions/doctors
    # tengan la misma longitud que months_arr para todas las familias (incluyendo
    # las que no estan en rec_comp y no se tocaron en el loop). Si una familia
    # no tenia data en algun mes, llenamos con 0 (no asumimos data).
    pres_root = d1.get('prescriptions', {})
    months_arr = pres_root.get('months', [])
    target_len = len(months_arr)
    for fam_name, fam_obj in pres_root.get('families', {}).items():
        if not isinstance(fam_obj, dict): continue
        for arr_key in ('prescriptions', 'doctors'):
            arr = fam_obj.get(arr_key)
            if not isinstance(arr, list): continue
            while len(arr) < target_len:
                arr.append(0)

    # Meta updates apuntan al ULTIMO mes mergeado
    if last_month_label_es:
        d1.setdefault('meta', {})['rxCut'] = last_month_label_es
        d2.setdefault('meta', {})['rec_label'] = last_month_key_short

    if not dry_run:
        new_text = serialize_data_js(text, d1, d2)
        data_js_path.write_text(new_text, encoding='utf-8', newline='')

    return {'families': fam_summary}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', required=True, help='Ruta al pivot Excel.')
    ap.add_argument('--months', nargs='*', default=None,
                    help='Filtrar a meses YYYY-MM. Si no se pasa, mergea TODOS los meses detectados en el pivot.')
    ap.add_argument('--repo', default=str(Path(__file__).resolve().parent.parent))
    ap.add_argument('--lines', nargs='+', default=LINES_DEFAULT)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pivot_path = Path(args.pivot)
    if not pivot_path.is_file():
        print(f'ERROR: pivot no existe: {pivot_path}', file=sys.stderr)
        return 2

    print(f'Loading pivot: {pivot_path}')
    months_in_pivot, brand_recetas, brand_medicos, fam_recetas, fam_medicos = load_pivot(pivot_path)
    print(f'  Meses detectados en pivot: {months_in_pivot}')
    for m in months_in_pivot:
        print(f'    {m}: brands={len(brand_recetas[m])} (con medicos: {len(brand_medicos[m])})')

    # Filtrar a --months si se paso
    if args.months:
        wanted = []
        for ym in args.months:
            mm = re.match(r'^(\d{4})-(\d{2})$', ym)
            if not mm: continue
            yr = int(mm.group(1)); mo = int(mm.group(2))
            wanted.append(f"{MES_EN[mo]} {yr}")
        months_to_merge = [m for m in months_in_pivot if m in wanted]
        if not months_to_merge:
            print(f'ERROR: ninguno de --months matchea con los meses del pivot.', file=sys.stderr)
            return 2
    else:
        months_to_merge = months_in_pivot

    print(f'\n  Mergeando meses: {months_to_merge}')

    repo = Path(args.repo)
    print(f'\nMerging en {len(args.lines)} lineas...')
    for line in args.lines:
        data_js = repo / line / 'data.js'
        if not data_js.is_file():
            print(f'  [{line}] SKIP: no data.js')
            continue
        try:
            res = merge_line(data_js, months_to_merge, brand_recetas, brand_medicos,
                             fam_recetas=fam_recetas, fam_medicos=fam_medicos, dry_run=args.dry_run)
        except Exception as e:
            print(f'  [{line}] ERROR: {e}')
            continue
        if 'skipped' in res:
            print(f'  [{line}] SKIP: {res["skipped"]}')
            continue
        fams = res['families']
        print(f'  [{line}] OK: {len(fams)} familias x {len(months_to_merge)} meses')
        for fam, s in list(fams.items())[:3]:
            ms_data = s.get('months', {})
            for m in months_to_merge:
                d = ms_data.get(m, {})
                if d:
                    print(f'    - {fam:18} {m}: SIE={d["sie"]:>6} mkt={d["mkt"]:>7} ms={d["ms"]:>5}%')

    if args.dry_run:
        print('\nDRY RUN: nada se escribio.')

    return 0


if __name__ == '__main__':
    sys.exit(main())
