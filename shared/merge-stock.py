#!/usr/bin/env python3
"""
shared/merge-stock.py

Mergea stock + ventas + facturacion + dias desde el pivot
"Laboratorio - Familia - Producto - <fecha>.xlsx" del SAP/sistema interno.

Estructura del pivot:
  R1: vacio | vacio | "Mes-Año" | "Totales" x4 | <fecha1> x4 | <fecha2> x4 | ...
  R2: "Laboratorio" | "Familia" | "Producto" | "Stock final" | "Ventas" | "Facturacion" | "Dias de Stock" | (repetido por mes)
  R3+: data, con filas "Totales" en columnas 1 y 2

Para cada linea (cardio/ATB/OTC/respi) actualiza in-place:
  - stock[F][<MMM YYYY>] = {stock, ventas, facturacion, dias} para meses
    NUEVOS (que no estaban en stock[F]).

Solo agrega meses que no existian. No sobreescribe meses presentes.
NO toca stock_alerts ni stock_pres (arrays derivados de longitud fija
que se usan para el panel de alerta; cualquier cambio ahi requeriria
recomputar statuses).
NO toca otras secciones: precios, recetas, mol_perf, etc.
NO toca meta.stockCut (si querés actualizarlo, pasa --update-meta).
NO toca SNC ni mujer (estructuras inline distintas).

Uso:
    py shared/merge-stock.py --pivot "<archivo.xlsx>"
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

LINES_DEFAULT = ['cardio', 'ATB', 'OTC', 'respiratorio']

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
MES_ES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}

def month_sort_value(month_key_en):
    parts = str(month_key_en).split()
    if len(parts) != 2: return 0
    en = parts[0]
    yr = int(parts[1])
    inv = {v:k for k,v in MES_EN.items()}
    return yr * 100 + inv.get(en, 0)


def parse_pivot(pivot_path):
    """Devuelve:
       - data: dict[family_norm][month_key_en] = {stock, ventas, facturacion, dias}
              (totales por familia)
              + dict[family_norm + '|' + producto_norm][month_key_en] = {...}
                para per-producto
       - months_in_pivot: lista ordenada de month_key_en"""
    wb = openpyxl.load_workbook(pivot_path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    # Map columns: { col_idx: (month_key_en, metric) }
    col_map = {}
    months_in_pivot = []
    for i, h1 in enumerate(row1):
        if not isinstance(h1, datetime): continue
        mk = f"{MES_EN[h1.month]} {h1.year}"
        if mk not in months_in_pivot:
            months_in_pivot.append(mk)
        h2 = (str(row2[i]) if i < len(row2) and row2[i] else '').strip().lower()
        metric = None
        if 'stock final' in h2: metric = 'stock'
        elif 'ventas' in h2: metric = 'ventas'
        elif 'facturaci' in h2: metric = 'facturacion'
        elif 'dias' in h2 or 'días' in h2: metric = 'dias'
        if metric:
            col_map[i] = (mk, metric)

    family_data = defaultdict(lambda: defaultdict(dict))   # family -> month -> {stock, ventas, ...}
    product_data = defaultdict(lambda: defaultdict(dict))  # producto -> month -> {...}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row: continue
        lab = row[0] if len(row) > 0 else None
        familia = row[1] if len(row) > 1 else None
        producto = row[2] if len(row) > 2 else None
        if not lab or str(lab).strip() == 'Totales': continue
        if not familia or str(familia).strip() == 'Totales': continue
        # Familia row (producto = "Totales")
        is_family_row = (str(producto or '').strip() == 'Totales')
        if not producto and not is_family_row: continue

        fam_key = str(familia).strip().upper()

        if is_family_row:
            for col_idx, (mk, metric) in col_map.items():
                if col_idx >= len(row): continue
                val = row[col_idx]
                try: v = float(val) if val is not None else None
                except: v = None
                if v is None: continue
                family_data[fam_key][mk][metric] = int(round(v)) if metric != 'dias' else int(round(v))
            continue

        # Producto row
        prod_key = str(producto).strip()
        for col_idx, (mk, metric) in col_map.items():
            if col_idx >= len(row): continue
            val = row[col_idx]
            try: v = float(val) if val is not None else None
            except: v = None
            if v is None: continue
            product_data[prod_key][mk][metric] = int(round(v))

    wb.close()
    return family_data, product_data, months_in_pivot


def parse_data_js(text):
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text)
    if not m1: raise ValueError('OTC_DATA not found')
    obj_start1 = text.index('{', m1.end())
    d1, end1 = json.JSONDecoder().raw_decode(text[obj_start1:])
    abs_end1 = obj_start1 + end1
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text[abs_end1:])
    if not m2: return d1, None
    obj_start2 = abs_end1 + text[abs_end1:].index('{', m2.end())
    d2, _ = json.JSONDecoder().raw_decode(text[obj_start2:])
    return d1, d2


def serialize_data_js(text_orig, d1, d2):
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text_orig)
    obj_start1 = text_orig.index('{', m1.end())
    _, end1 = json.JSONDecoder().raw_decode(text_orig[obj_start1:])
    abs_end1 = obj_start1 + end1
    prefix1 = text_orig[:obj_start1]
    if d2 is None:
        return prefix1 + json.dumps(d1, ensure_ascii=False) + text_orig[abs_end1:]
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text_orig[abs_end1:])
    obj_start2 = abs_end1 + text_orig[abs_end1:].index('{', m2.end())
    _, end2 = json.JSONDecoder().raw_decode(text_orig[obj_start2:])
    abs_end2 = obj_start2 + end2
    middle = text_orig[abs_end1:obj_start2]
    suffix = text_orig[abs_end2:]
    return prefix1 + json.dumps(d1, ensure_ascii=False) + middle + json.dumps(d2, ensure_ascii=False) + suffix


def merge_line(data_js_path, family_data, months_in_pivot, dry_run=False):
    text = data_js_path.read_text(encoding='utf-8-sig', errors='replace')
    d1, d2 = parse_data_js(text)
    if d2 is None: return {'skipped': 'no OTC_DASHBOARD'}
    stock = d2.get('stock')
    if not isinstance(stock, dict): return {'skipped': 'no stock'}

    summary = {'families': {}, 'months_added_total': 0, 'matched': 0, 'unmatched_families': []}

    for fam in stock.keys():
        fam_dict = stock[fam]
        if not isinstance(fam_dict, dict): continue
        existing_months = set(fam_dict.keys())
        # Encontrar la familia en el pivot. Match upper case.
        fam_upper = str(fam).strip().upper()
        pivot_fam_data = family_data.get(fam_upper)
        if not pivot_fam_data:
            # Probar match parcial (algunas familias tienen sub-spaces)
            for k in family_data.keys():
                if k.replace(' ', '') == fam_upper.replace(' ', ''):
                    pivot_fam_data = family_data[k]; break
        if not pivot_fam_data:
            summary['unmatched_families'].append(fam)
            continue
        summary['matched'] += 1

        # Solo agregar meses NUEVOS y CERRADOS. Un mes "en curso" tiene
        # stock=0 + ventas=0 (SAP todavia no consolido el cierre); incluirlo
        # contamina el dashboard con dias=0/alertas falsas. Skip esos meses.
        added = []
        skipped_partial = []
        for mk in months_in_pivot:
            if mk in existing_months: continue
            if mk not in pivot_fam_data: continue
            entry = pivot_fam_data[mk]
            stock_v = entry.get('stock', 0) or 0
            ventas_v = entry.get('ventas', 0) or 0
            if stock_v == 0 and ventas_v == 0:
                # Mes en curso, datos parciales -> skip
                skipped_partial.append(mk)
                continue
            fam_dict[mk] = {
                'stock': stock_v,
                'ventas': ventas_v,
                'facturacion': entry.get('facturacion', 0) or 0,
                'dias': entry.get('dias', 0) or 0,
            }
            added.append(mk)
            summary['months_added_total'] += 1
        if added:
            summary['families'][fam] = added
        if skipped_partial:
            summary.setdefault('skipped_partial', {})[fam] = skipped_partial

    if not dry_run:
        new_text = serialize_data_js(text, d1, d2)
        data_js_path.write_text(new_text, encoding='utf-8', newline='')

    return summary


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', required=True)
    ap.add_argument('--repo', default=str(Path(__file__).resolve().parent.parent))
    ap.add_argument('--lines', nargs='+', default=LINES_DEFAULT)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pf = Path(args.pivot)
    if not pf.is_file():
        print(f'ERROR: pivot no existe: {pf}', file=sys.stderr); return 2

    print(f'Leyendo: {pf}')
    family_data, product_data, months_in_pivot = parse_pivot(pf)
    print(f'  Meses en pivot: {months_in_pivot}')
    print(f'  Familias en pivot: {len(family_data)}')

    repo = Path(args.repo)
    print(f'\nMergeando stock en {len(args.lines)} lineas (solo meses NUEVOS)...')
    for line in args.lines:
        data_js = repo / line / 'data.js'
        if not data_js.is_file():
            print(f'  [{line}] SKIP: no data.js'); continue
        try:
            res = merge_line(data_js, family_data, months_in_pivot, dry_run=args.dry_run)
        except Exception as e:
            print(f'  [{line}] ERROR: {e}'); continue
        if 'skipped' in res:
            print(f'  [{line}] SKIP: {res["skipped"]}'); continue
        added_count = res['months_added_total']
        if added_count == 0:
            print(f'  [{line}] OK: nada nuevo cerrado para mergear ({res["matched"]} familias matched)')
        else:
            print(f'  [{line}] OK: {added_count} meses agregados, {res["matched"]} familias matched')
            for f, ms in res['families'].items():
                print(f'    + {f}: {ms}')
        if res.get('skipped_partial'):
            sample = list(res['skipped_partial'].items())[:3]
            print(f'    Meses parciales (en curso) skipeados:')
            for f, ms in sample:
                print(f'      {f}: {ms}')
        if res['unmatched_families']:
            print(f'    Familias sin match en pivot: {res["unmatched_families"][:5]}')

    if args.dry_run:
        print('\nDRY RUN: nada se escribio.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
