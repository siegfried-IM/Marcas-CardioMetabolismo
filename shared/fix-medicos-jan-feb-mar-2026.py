#!/usr/bin/env python3
"""
shared/fix-medicos-jan-feb-mar-2026.py

One-off fix: re-lee los pivots CloseUp y reemplaza el campo `medicos` en
D.recetas[fam][month] para Ene/Feb/Mar 2026 con el COUNT ÚNICO (la fila
(mercado,'Totales','') del pivot), en vez de la SUMA por marca que estaba
inflada (un mismo médico que prescribe N marcas se contaba N veces).

Líneas afectadas: cardio, ATB, respiratorio.
OTC no tiene pivot CloseUp disponible — se deja como está.

Uso: py shared/fix-medicos-jan-feb-mar-2026.py
"""
from __future__ import annotations
import re, json, sys
from pathlib import Path
import openpyxl
from datetime import datetime

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

INPUTS = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')
REPO = Path(__file__).resolve().parent.parent

PIVOTS = [
    ('cardio',       INPUTS / 'cardio/2026-04/fuentes-originales/RECETAS.xlsx'),
    ('ATB',          INPUTS / 'ATB/2026-04/fuentes-originales/recetas.xlsx'),
    ('respiratorio', INPUTS / 'respiratorio/2026-04/dashboard/RECETAS.xlsx'),
]
MONTHS_FIX = ['Jan 2026', 'Feb 2026', 'Mar 2026']


def load_fam_medicos(pivot_path: Path):
    """Devuelve {month_key_en: {market: medicos_unique}} desde row (mercado,'Totales','')."""
    wb = openpyxl.load_workbook(pivot_path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    col_map = {}
    cur_month = None
    months = []
    for i, h1 in enumerate(row1):
        if isinstance(h1, datetime):
            cur_month = f"{MES_EN[h1.month]} {h1.year}"
            if cur_month not in months: months.append(cur_month)
        sub = (str(row2[i]) if i < len(row2) and row2[i] else '').lower()
        if not cur_month: continue
        if 'medico' in sub or 'médico' in sub:
            col_map[i] = (cur_month, 'medicos')
        elif 'receta' in sub:
            col_map[i] = (cur_month, 'recetas')
    fam_med = {m: {} for m in months}
    fam_rec = {m: {} for m in months}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row: continue
        merc = row[0] if len(row) > 0 else None
        droga = row[1] if len(row) > 1 else None
        marca = row[2] if len(row) > 2 else None
        if not merc: continue
        is_market_total = (str(droga or '').strip().lower() == 'totales' and not marca)
        if not is_market_total: continue
        market = str(merc).strip()
        for col_idx, (mk, kind) in col_map.items():
            if col_idx >= len(row): continue
            v = row[col_idx]
            try: v = int(v) if v is not None else 0
            except (TypeError, ValueError): v = 0
            target = fam_med if kind == 'medicos' else fam_rec
            if market not in target[mk]:
                target[mk][market] = v
    wb.close()
    return fam_rec, fam_med


def parse_data_js(text):
    pat = re.compile(r'window\.OTC_DASHBOARD\s*=\s*', re.S)
    m = pat.search(text)
    if not m:
        return None, None, None
    ob = text.index('{', m.end())
    d2, end = json.JSONDecoder().raw_decode(text[ob:])
    return text[:ob], d2, text[ob+end:]


def main():
    for line_key, pivot in PIVOTS:
        if not pivot.is_file():
            print(f'  [{line_key}] SKIP: pivot no existe: {pivot}')
            continue
        print(f'\n[{line_key}] pivot: {pivot.name}')
        fam_rec, fam_med = load_fam_medicos(pivot)
        for mk in MONTHS_FIX:
            print(f'  {mk}: {len(fam_med.get(mk, {}))} familias con medicos unique')

        data_js = REPO / line_key / 'data.js'
        if not data_js.is_file():
            print(f'  data.js no existe en {data_js}')
            continue
        text = data_js.read_text(encoding='utf-8-sig', errors='replace')
        prefix, d2, suffix = parse_data_js(text)
        if d2 is None:
            print(f'  ERROR parsing data.js'); continue

        recetas = d2.get('recetas', {})
        changed = 0
        for fam, months_data in recetas.items():
            if not isinstance(months_data, dict): continue
            # 1) Aplicar correccion para meses que estan en el pivot
            for mk in MONTHS_FIX:
                if mk not in months_data: continue
                new_med = fam_med.get(mk, {}).get(fam)
                if new_med is None: continue
                old = months_data[mk].get('medicos', 0)
                if old != new_med:
                    months_data[mk]['medicos'] = new_med
                    changed += 1
                    if changed <= 10:
                        print(f'    {fam} {mk}: medicos {old} -> {new_med}')
            # 2) Mar 2026 NO esta en pivot (corte 2026-04 va hasta Feb).
            # Si Feb 2026 tiene unique-count y Mar 2026 sigue con la inflada,
            # usar Feb como proxy para Mar (mejor que la inflada).
            feb_med = (fam_med.get('Feb 2026') or {}).get(fam)
            if feb_med is not None and 'Mar 2026' in months_data:
                mar_old = months_data['Mar 2026'].get('medicos', 0)
                if mar_old > feb_med * 1.5:  # inflada >50% sobre Feb -> sospechosa
                    months_data['Mar 2026']['medicos'] = feb_med
                    changed += 1
                    print(f'    {fam} Mar 2026: medicos {mar_old} -> {feb_med} (proxy Feb, no hay pivot Mar)')
        print(f'  Total entradas actualizadas: {changed}')

        new_text = prefix + json.dumps(d2, ensure_ascii=False) + suffix
        data_js.write_text(new_text, encoding='utf-8', newline='')
        print(f'  -> escrito {data_js} ({data_js.stat().st_size:,} bytes)')

    return 0


if __name__ == '__main__':
    sys.exit(main())
