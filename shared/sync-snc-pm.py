#!/usr/bin/env python3
"""
shared/sync-snc-pm.py

Sincroniza la inline D de SNC/index.html con la data IQVIA Premium del
AR_PM master. Reconstruye mol_perf SEGMENTADO POR MOLECULA con todos
los productos del mercado por molecula.

Para cada molecula tracked en SNC:
  - ALPRAZOLAM, BENSERAZIDE_LEVODOPA, DIAZEPAM, LEVETIRACETAM,
    LURASIDONE, PREGABALIN, QUETIAPINE, THIORIDAZINE, VILAZODONE,
    VORTIOXETINE
  -> Toma todas las rows del AR_PM con esa Molecules Long
  -> Agrupa por Product (col 1)
  -> Marca is_sie cuando manuf == 'SIEGFRIED' o el Product viene del set
     conocido de SIE (extraido de inline D actual)
  -> Construye monthly_vals / quarterly_vals / ytd / mat / ms_*

Preserva todos los demas campos de inline D (rec_ms, rec_comp, recetas,
stock, precios, etc.). Solo reemplaza mol_perf.

Uso:
    py shared/sync-snc-pm.py [--master <path>] [--dry-run]
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'SNC' / 'index.html'
DEFAULT_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Apr-27-2026.xlsx')

# Moleculas que SNC trackea (deben matchear AR_PM "Molecules Long")
SNC_MOLECULES = [
    'ALPRAZOLAM', 'BENSERAZIDE_LEVODOPA', 'DIAZEPAM',
    'LEVETIRACETAM', 'LURASIDONE', 'PREGABALIN',
    'QUETIAPINE', 'THIORIDAZINE', 'VILAZODONE', 'VORTIOXETINE',
]

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}


def msort(mk):
    p = mk.split()
    if len(p) != 2: return 0
    return int(p[1]) * 100 + MES_INV.get(p[0], 0)


def quarter_key(mk):
    parts = mk.split()
    if len(parts) != 2: return ''
    m = MES_INV.get(parts[0])
    if not m: return ''
    q = (m - 1) // 3 + 1
    return f'Q{q} {parts[1]}'


def load_pm(path, molecule_filter):
    """Devuelve (months_asc, by_molecule_then_product).
    by_molecule[molecule_key][product] = {'manuf':..., 'monthly':{mk:v}}.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    col_manuf = col_product = col_pack = col_atc = col_mol = None
    month_cols = []
    for i, h in enumerate(row1):
        if not h: continue
        s = str(h).strip()
        s_norm = s.replace('\n', ' ').strip()
        if s_norm.lower().startswith('manufacturer'): col_manuf = i
        elif s_norm.lower().startswith('product'):    col_product = i
        elif s_norm.lower().startswith('pack'):       col_pack = i
        elif s_norm.lower().startswith('atc'):        col_atc = i
        elif s_norm.lower().startswith('molecules'):  col_mol = i
        if s.startswith('Units') and ('\n' in s or len(s.split()) >= 2):
            after = s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):].strip()
            after = after.strip()
            if after.upper().startswith('MAT') or after.upper().startswith('YTD'):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV:
                month_cols.append((i, f'{m.group(1)} {m.group(2)}'))

    if col_manuf is None: col_manuf = 0
    if col_product is None: col_product = 1
    if col_pack is None: col_pack = 2
    if col_atc is None: col_atc = 3
    if col_mol is None: col_mol = 5

    target_set = set(molecule_filter)
    by_mol_prod = defaultdict(lambda: defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float)}))

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        manuf   = row[col_manuf]   if col_manuf < len(row)   else None
        product = row[col_product] if col_product < len(row) else None
        mol     = row[col_mol]     if col_mol < len(row)     else None
        if not product or not mol: continue
        mol_str = str(mol).strip()
        if mol_str not in target_set:
            continue
        bucket = by_mol_prod[mol_str][product]
        bucket['manuf'] = manuf
        for ci, mk in month_cols:
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try:
                bucket['monthly'][mk] += float(v)
            except (ValueError, TypeError):
                pass
    wb.close()

    months_asc = [mk for _, mk in sorted(month_cols, key=lambda x: msort(x[1]))]
    # Convert defaultdict-of-float to plain dict
    out = {}
    for mol_key, prods in by_mol_prod.items():
        out[mol_key] = {p: {'manuf': info['manuf'],
                            'monthly': {mk: int(round(v)) for mk, v in info['monthly'].items()}}
                        for p, info in prods.items()}
    return months_asc, out


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk: out[qk] += v
    return dict(out)


def aggregate_ytd_per_year(monthly, cierre_month=12):
    """YTD por año = sum Jan..<cierre_month>. Key = '<mes_cierre> <YYYY>'."""
    if not monthly: return {}
    num_to_label = {v: k for k, v in MES_INV.items()}
    mes_label = num_to_label.get(cierre_month, 'Dec')
    by_year = defaultdict(int)
    for mk, v in monthly.items():
        parts = mk.split()
        if len(parts) != 2: continue
        m_num = MES_INV.get(parts[0])
        if not m_num: continue
        if m_num <= cierre_month:
            by_year[parts[1]] += v
    return {f'{mes_label} {y}': v for y, v in by_year.items()}


def aggregate_mat(monthly, cierre_month=12):
    """MAT por año = rolling 12 meses terminando en <cierre_month>.
    Key = '<mes_cierre> <YYYY>'."""
    if not monthly: return {}
    num_to_label = {v: k for k, v in MES_INV.items()}
    mes_label = num_to_label.get(cierre_month, 'Dec')
    years_with = set()
    for mk in monthly:
        parts = mk.split()
        if len(parts) == 2 and parts[0] in MES_INV:
            years_with.add(int(parts[1]))
    out = {}
    for y in sorted(years_with):
        total = 0
        for back in range(11, -1, -1):
            total_idx = (y * 12 + (cierre_month - 1)) - back
            yy, mm = divmod(total_idx, 12)
            mk = f'{num_to_label[mm + 1]} {yy}'
            total += int(monthly.get(mk, 0) or 0)
        out[f'{mes_label} {y}'] = total
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--html', default=str(HTML))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pm_path = Path(args.master)
    if not pm_path.is_file():
        print(f'ERROR: master no existe: {pm_path}', file=sys.stderr); return 2

    print(f'Leyendo: {pm_path}')
    print(f'Filtrando moleculas: {SNC_MOLECULES}')
    all_months, by_mol = load_pm(pm_path, SNC_MOLECULES)
    print(f'  {len(all_months)} meses ({all_months[0]} .. {all_months[-1]})')
    # Cierre month como int (1-12) para aggregates ytd/mat
    cierre_m = MES_INV.get(all_months[-1].split()[0], 12) if all_months else 12
    for mol, prods in by_mol.items():
        print(f'  [{mol}]: {len(prods)} productos')

    # Read inline D
    html_path = Path(args.html)
    text = html_path.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    if not m:
        print('ERROR: const D no encontrado'); return 3
    abs_start = m.start() + len('const D = ')
    abs_start = text.index('{', abs_start)
    D, end = json.JSONDecoder().raw_decode(text[abs_start:])
    abs_end = abs_start + end

    old_mol_perf = D.get('mol_perf', {})

    # Detectar SIE por molecule: si el viejo inline D tenia is_sie=true, mantener
    sie_set_per_mol = {}
    for mol_key, fam_obj in old_mol_perf.items():
        if isinstance(fam_obj, dict):
            sie_set_per_mol[mol_key] = {p['prod'] for p in fam_obj.get('products', []) if p.get('is_sie')}

    # Construir nuevo mol_perf
    new_mol_perf = {}
    for mol_key in SNC_MOLECULES:
        prods = by_mol.get(mol_key, {})
        old_fam = old_mol_perf.get(mol_key, {})
        # Family-level monthly
        fam_monthly = defaultdict(int)
        product_list = []
        sie_set = sie_set_per_mol.get(mol_key, set())
        # Si no tenemos SIE set previo, detectar por manuf
        for prod_name, info in prods.items():
            monthly = info['monthly']
            for mk, v in monthly.items():
                fam_monthly[mk] += v
            is_sie = (prod_name in sie_set) or (str(info.get('manuf') or '').strip().upper() == 'SIEGFRIED')
            quarterly = aggregate_quarterly(monthly)
            ytd = aggregate_ytd_per_year(monthly, cierre_month=cierre_m)
            mat = aggregate_mat(monthly, cierre_month=cierre_m)
            product_list.append({
                'prod': prod_name,
                'manuf': info.get('manuf') or '',
                'is_sie': is_sie,
                'monthly_vals': monthly,
                'quarterly_vals': quarterly,
                'ytd': ytd,
                'mat': mat,
                'ms_monthly': {},   # se llena despues
                'ms_quarterly': {},
                'ms_ytd': {},
                'ms_mat': {},
            })

        fam_monthly = dict(fam_monthly)
        fam_quarterly = aggregate_quarterly(fam_monthly)
        fam_ytd = aggregate_ytd_per_year(fam_monthly, cierre_month=cierre_m)
        fam_mat = aggregate_mat(fam_monthly, cierre_month=cierre_m)

        # Compute MS por producto vs total molecula
        for p in product_list:
            mv = p['monthly_vals']
            p['ms_monthly'] = {mk: round(mv.get(mk, 0)/fv*100, 2) if fv > 0 else 0
                                for mk, fv in fam_monthly.items()}
            qv = p['quarterly_vals']
            p['ms_quarterly'] = {qk: round(qv.get(qk, 0)/fv*100, 2) if fv > 0 else 0
                                 for qk, fv in fam_quarterly.items()}
            yv = p['ytd']
            p['ms_ytd'] = {y: round(yv.get(y, 0)/fv*100, 2) if fv > 0 else 0
                            for y, fv in fam_ytd.items()}
            mtv = p['mat']
            p['ms_mat'] = {y: round(mtv.get(y, 0)/fv*100, 2) if fv > 0 else 0
                            for y, fv in fam_mat.items()}

        # Ordenar productos: SIE primero, luego por monthly_vals max desc
        def sort_key(p):
            last_mk = max(p['monthly_vals'].keys(), key=msort, default='')
            last_v = p['monthly_vals'].get(last_mk, 0)
            return (not p['is_sie'], -last_v)
        product_list.sort(key=sort_key)

        new_mol_perf[mol_key] = {
            'family': old_fam.get('family', '') if isinstance(old_fam, dict) else '',
            'products': product_list,
            'monthly': fam_monthly,
            'quarterly': fam_quarterly,
            'ytd': fam_ytd,
            'mat': fam_mat,
        }

    D['mol_perf'] = new_mol_perf

    # Diagnostico
    sample_mol = SNC_MOLECULES[0]
    sample = new_mol_perf[sample_mol]
    sie_prods = [p for p in sample['products'] if p['is_sie']]
    print(f'\nSample [{sample_mol}]:')
    print(f'  Total products: {len(sample["products"])}, SIE: {len(sie_prods)}')
    if sie_prods:
        s = sie_prods[0]
        last_mks = sorted(s['monthly_vals'].keys(), key=msort)[-3:]
        print(f'  SIE {s["prod"]}: {[(mk, s["monthly_vals"][mk]) for mk in last_mks]}'.encode('ascii','replace').decode())

    if args.dry_run:
        print('\nDRY RUN: no se escribio.')
        return 0

    new_text = text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    html_path.write_text(new_text, encoding='utf-8', newline='')
    print(f'\n-> {html_path} reescrito ({html_path.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
