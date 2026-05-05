#!/usr/bin/env python3
"""
shared/merge-carbidopa-into-benserazide.py

Mergea la molecula CARBIDOPA_LEVODOPA dentro de BENSERAZIDE_LEVODOPA en
mol_perf de SNC. En IQVIA estan separadas (benserazida vs carbidopa
como inhibidores periféricos diferentes), pero clinicamente compiten en
Parkinson, asi que el dashboard agrupa ambas para que MADOPAR/MADOPAR HBS
muestren a PARKINEL (BAG), LEBOCAR (PFZ), STALEVO (BU-) como competidores.

Pasos:
1. Lee AR_PM master, extrae todos los productos de CARBIDOPA_LEVODOPA con
   monthly_vals
2. Los agrega como products no-SIE a mol_perf.BENSERAZIDE_LEVODOPA
3. Recomputa monthly/quarterly/ytd/mat de la familia
4. Recomputa ms_* de TODOS los productos contra los nuevos totales
5. Actualiza MOL_LABELS['BENSERAZIDE_LEVODOPA'] para reflejar la fusion

NO toca otras moleculas, ni rec_ms, ni recetas, ni stock, ni nada mas.
"""
from __future__ import annotations
import re, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

IQVIA_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Apr-27-2026.xlsx')
HTML_PATH = Path(r'C:\Users\camarinaro\Marcas-CardioMetabolismo\SNC\index.html')

ADD_MOL_FROM_MASTER = 'CARBIDOPA_LEVODOPA'
TARGET_MOL_KEY = 'BENSERAZIDE_LEVODOPA'

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
MES_EN = {v:k for k,v in MES_INV.items()}
CIERRE_M = 3


def msort(mk):
    p = mk.split()
    return int(p[1]) * 100 + MES_INV.get(p[0], 0) if len(p) == 2 else 0


def agg_ytd(mv, cm):
    by_y = defaultdict(int)
    for mk, v in mv.items():
        p = mk.split()
        m = MES_INV.get(p[0]) if len(p) == 2 else None
        if m and m <= cm:
            by_y[p[1]] += int(v or 0)
    return {f'{MES_EN[cm]} {y}': v for y, v in by_y.items()}


def agg_mat_yearly(mv, cm):
    years = set(mk.split()[1] for mk in mv if len(mk.split()) == 2)
    out = {}
    for yr in sorted(years, key=int):
        y = int(yr); total = 0
        for back in range(12):
            idx = y * 12 + (cm - 1) - back
            yy, mm = divmod(idx, 12)
            total += int(mv.get(f'{MES_EN[mm + 1]} {yy}', 0) or 0)
        out[f'{MES_EN[cm]} {yr}'] = total
    return out


def agg_mat_monthly(mv):
    sk = sorted(mv.keys(), key=msort); out = {}
    for i, mk in enumerate(sk):
        if i < 11: continue
        out[mk] = sum(int(mv.get(m, 0) or 0) for m in sk[i - 11:i + 1])
    return out


def agg_quarterly(mv):
    out = defaultdict(int)
    for mk, v in mv.items():
        p = mk.split()
        m = MES_INV.get(p[0]) if len(p) == 2 else None
        if m: out[f'Q{(m - 1) // 3 + 1} {p[1]}'] += int(v or 0)
    return dict(out)


def recompute_ms(p, fm, fq, fy, fmat):
    mv = p.get('monthly_vals', {}); qv = p.get('quarterly_vals', {})
    yv = p.get('ytd', {}); mtv = p.get('mat', {})
    p['ms_monthly']   = {mk: round(mv.get(mk, 0)/fv*100, 2) if fv > 0 else 0  for mk, fv in fm.items()}
    p['ms_quarterly'] = {qk: round(qv.get(qk, 0)/fv*100, 2) if fv > 0 else 0  for qk, fv in fq.items()}
    p['ms_ytd']       = {y:  round(yv.get(y,  0)/fv*100, 2) if fv > 0 else 0  for y,  fv in fy.items()}
    p['ms_mat']       = {mk: round(mtv.get(mk,0)/fv*100, 2) if fv > 0 else 0  for mk, fv in fmat.items()}


def main():
    if not IQVIA_MASTER.is_file():
        print(f'ERROR: {IQVIA_MASTER} no existe', file=sys.stderr); return 2

    # ── 1. Parse AR_PM, extract CARBIDOPA_LEVODOPA products ─────────────
    wb = openpyxl.load_workbook(IQVIA_MASTER, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_manuf=0; col_product=1; col_pack=2; col_mol=5
    month_cols = []
    for i, h in enumerate(row1):
        if not h: continue
        s = str(h).strip()
        if not s.startswith('Units'): continue
        after = s.split('\n', 1)[-1].strip() if '\n' in s else s[5:].strip()
        if after.upper().startswith('MAT') or after.upper().startswith('YTD'): continue
        m = re.match(r'(\w+)\s+(\d{4})', after)
        if m and m.group(1) in MES_INV:
            month_cols.append((i, f'{m.group(1)} {m.group(2)}'))

    by_prod = defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float)})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        prod = row[col_product]; mol = row[col_mol]; manuf = row[col_manuf]
        if not prod or not mol: continue
        if str(mol).strip() != ADD_MOL_FROM_MASTER: continue
        bucket = by_prod[str(prod).strip()]
        bucket['manuf'] = manuf
        for ci, mk in month_cols:
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try: bucket['monthly'][mk] += float(v)
            except (ValueError, TypeError): pass
    wb.close()

    print(f'Productos {ADD_MOL_FROM_MASTER} encontrados: {len(by_prod)}')
    for p, info in by_prod.items():
        ytd = agg_ytd({k:int(round(v)) for k,v in info['monthly'].items()}, CIERRE_M)
        print(f'  {p:30s} manuf={info["manuf"]:15s} ytd_mar26={ytd.get("Mar 2026")}')

    # ── 2. Load HTML + D ────────────────────────────────────────────────
    text = HTML_PATH.read_text(encoding='utf-8', errors='replace')
    m_obj = re.search(r'const D = (\{)', text)
    obj_start = m_obj.start() + len('const D = ')
    obj_start = text.index('{', obj_start)
    D, end_off = json.JSONDecoder().raw_decode(text[obj_start:])
    abs_end = obj_start + end_off

    benz = D['mol_perf'][TARGET_MOL_KEY]
    existing_prods = {p['prod'] for p in benz['products']}
    print(f'\nProductos actuales en {TARGET_MOL_KEY}: {existing_prods}')

    # ── 3. Append CARBIDOPA products as non-SIE products ────────────────
    # Use a template product structure based on existing non-SIE shape
    template = next((p for p in benz['products'] if not p.get('is_sie')), None)
    if template is None:
        # Build minimal template from a SIE product
        template = benz['products'][0] if benz['products'] else None

    for prod_name, info in by_prod.items():
        if prod_name in existing_prods:
            print(f'  SKIP {prod_name}: ya existe'); continue
        mv = {k: int(round(v)) for k, v in info['monthly'].items()}
        new_p = {}
        # Copy base structure from template (lab, color etc. fields)
        if template:
            for k in ('lab', 'color'):
                if k in template:
                    new_p[k] = template[k]
        new_p['prod'] = prod_name
        new_p['is_sie'] = False
        new_p['monthly_vals'] = mv
        new_p['quarterly_vals'] = agg_quarterly(mv)
        new_p['ytd'] = agg_ytd(mv, CIERRE_M)
        new_p['mat'] = {**agg_mat_yearly(mv, CIERRE_M), **agg_mat_monthly(mv)}
        benz['products'].append(new_p)

    # ── 4. Recompute family totals ──────────────────────────────────────
    fam_monthly = defaultdict(int)
    for p in benz['products']:
        for mk, v in p.get('monthly_vals', {}).items():
            fam_monthly[mk] += int(v or 0)
    fam_monthly = dict(fam_monthly)
    fam_quarterly = agg_quarterly(fam_monthly)
    fam_ytd = agg_ytd(fam_monthly, CIERRE_M)
    fam_mat = {**agg_mat_yearly(fam_monthly, CIERRE_M), **agg_mat_monthly(fam_monthly)}
    benz['monthly'] = fam_monthly
    benz['quarterly'] = fam_quarterly
    benz['ytd'] = fam_ytd
    benz['mat'] = fam_mat

    # ── 5. Recompute MS for all products vs new family ──────────────────
    for p in benz['products']:
        recompute_ms(p, fam_monthly, fam_quarterly, fam_ytd, fam_mat)

    print(f'\nNuevos totales familia (Mar 2026):')
    print(f'  YTD: {fam_ytd.get("Mar 2026")}')
    print(f'  MAT: {fam_mat.get("Mar 2026")}')
    print(f'\nMS% YTD Mar 2026 por producto:')
    for p in benz['products']:
        ms = p.get('ms_ytd', {}).get('Mar 2026')
        ytd_v = p.get('ytd', {}).get('Mar 2026')
        sie_tag = ' [SIE]' if p.get('is_sie') else ''
        print(f'  {p["prod"]:30s}{sie_tag}: ytd={ytd_v}, ms={ms}%')

    # ── 6. Serialize ────────────────────────────────────────────────────
    new_text = text[:obj_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]

    # Update MOL_LABELS to reflect merged molecule
    new_text = new_text.replace(
        "'BENSERAZIDE_LEVODOPA':'Benserazida+L-dopa',",
        "'BENSERAZIDE_LEVODOPA':'Benserazida/Carbidopa+L-dopa',"
    )

    HTML_PATH.write_text(new_text, encoding='utf-8', newline='')
    print(f'\nSaved: {HTML_PATH.stat().st_size:,} bytes')
    return 0


if __name__ == '__main__':
    sys.exit(main())
