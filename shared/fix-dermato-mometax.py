#!/usr/bin/env python3
"""
shared/fix-dermato-mometax.py

Restaura recetas de MOMETAX en dermato. Estaban en 0 desde Oct 2025
en adelante. La pivot CloseUp (3) tiene data fresca para Ene/Feb/Mar
2026 del mercado MOMETAX TOTAL droga MOMETASONA.

Actualiza:
  - rec_ms[MOMETAX].sie/ms para Ene/Feb/Mar 2026
  - rec_comp[MOMETAX][brand] para Ene/Feb/Mar 2026
  - recetas[MOMETAX] (total mercado) para Ene/Feb/Mar 2026

NO toca otras familias / lineas / data.
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'dermatologia' / 'dermato_dashboard.html'
DEFAULT_PIVOT = Path(r'C:\Users\camarinaro\Downloads\Sin título - Tabla dinámica - 4 de mayo de 2026 (3).xlsx')
MERCADO = 'MOMETAX TOTAL'
DROGA = 'MOMETASONA'
SIE_BRANDS = {'MOMETAX SIE'}

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
MES_INV = {v:k for k,v in MES_EN.items()}


def parse_pivot(p):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    col_map = {}
    months_in = []
    for i, h in enumerate(row1):
        if not isinstance(h, datetime): continue
        mk = f'{MES_EN[h.month]} {h.year}'
        if mk not in months_in: months_in.append(mk)
        h2 = (str(row2[i]) if i < len(row2) and row2[i] else '').strip().lower()
        if 'recetas' in h2: col_map[i] = (mk, 'rec')
        elif 'm' in h2 and ('dico' in h2 or 'edico' in h2): col_map[i] = (mk, 'med')
    out = {'totales_droga': {'rec':{}, 'med':{}}, 'brands': {}}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3: continue
        merc = (row[0] or '').strip()
        droga = (row[1] or '').strip()
        marca = (row[2] or '').strip()
        if merc != MERCADO: continue
        if droga != DROGA: continue
        if marca == 'Totales':
            for i, (mk, metric) in col_map.items():
                if i >= len(row): continue
                v = row[i]
                if v is None or v == '-': continue
                try: out['totales_droga'][metric][mk] = int(round(float(v)))
                except: pass
            continue
        if not marca: continue
        b = out['brands'].setdefault(marca, {'rec':{}, 'med':{}, 'is_sie': marca in SIE_BRANDS})
        for i, (mk, metric) in col_map.items():
            if i >= len(row): continue
            v = row[i]
            if v is None or v == '-': continue
            try: b[metric][mk] = int(round(float(v)))
            except: pass
    wb.close()
    return months_in, out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', default=str(DEFAULT_PIVOT))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    pf = Path(args.pivot)
    if not pf.is_file():
        print(f'ERROR: no existe {pf}', file=sys.stderr); return 2
    print(f'Leyendo: {pf}')
    months, data = parse_pivot(pf)
    print(f'  Meses: {months}')
    print(f'  Brands: {list(data["brands"].keys())}')

    # Mapeo brand pivot -> brand en dermato (sin sufijo SIE para SIE products)
    BRAND_MAP = {
        'MOMETAX SIE': 'MOMETAX',  # SIE name en dermato
        'SUAVICORT CSR': 'SUAVICORT CSR',
        'MOMEPLUS PAB': 'MOMEPLUS PAB',
        'ELOCON ORG': 'ELOCON ORG',
        'METASON F-B': 'METASON F-B',
        'MOMETASONA LNI LNI': 'MOMETASONA LNI LNI',
        'NOVASONE ORG': 'NOVASONE ORG',
    }

    text = HTML.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    abs_start = m.start() + len('const D = ')
    abs_start = text.index('{', abs_start)
    D, end = json.JSONDecoder().raw_decode(text[abs_start:])
    abs_end = abs_start + end

    # 1) rec_ms[MOMETAX].sie / .ms
    rms = D['rec_ms'].setdefault('MOMETAX', {'sie':{}, 'ms':{}, 'ms_quarterly':{}})
    sie_d = rms.setdefault('sie', {})
    ms_d = rms.setdefault('ms', {})
    sie_brand_data = data['brands'].get('MOMETAX SIE', {})
    sie_monthly = sie_brand_data.get('rec', {})
    for mk in months:
        sv = sie_monthly.get(mk, 0)
        sie_d[mk] = sv
        tot = data['totales_droga']['rec'].get(mk, 0)
        ms_d[mk] = round(sv / tot * 100, 2) if tot > 0 else 0
        print(f'  rec_ms.MOMETAX.sie[{mk}] = {sv}, ms = {ms_d[mk]}%')
    # ms_quarterly
    rms.setdefault('ms_quarterly', {})
    Q1 = ['Jan 2026','Feb 2026','Mar 2026']
    sie_q = sum(sie_monthly.get(m,0) or 0 for m in Q1)
    tot_q = sum(data['totales_droga']['rec'].get(m,0) or 0 for m in Q1)
    if tot_q > 0:
        rms['ms_quarterly']['Q1 2026'] = round(sie_q/tot_q*100, 2)

    # 2) rec_comp[MOMETAX][brand].monthly / quarterly / total
    rc = D['rec_comp'].setdefault('MOMETAX', {})
    for pivot_brand, dermato_brand in BRAND_MAP.items():
        b = data['brands'].get(pivot_brand, {})
        rec_m = b.get('rec', {})
        med_m = b.get('med', {})
        if not rec_m and not med_m:
            print(f'  WARN: no data for {pivot_brand}')
            continue
        existing = rc.setdefault(dermato_brand, {
            'is_sie': dermato_brand == 'MOMETAX',
            'monthly':{}, 'quarterly':{}, 'total':0,
            'monthly_medicos':{}, 'quarterly_medicos':{}, 'total_medicos':0
        })
        # Asegurar que sub-dicts existan (puede haberse creado sin ellos antes)
        existing.setdefault('monthly', {})
        existing.setdefault('quarterly', {})
        existing.setdefault('monthly_medicos', {})
        existing.setdefault('quarterly_medicos', {})
        # Update monthly with new months
        for mk in months:
            if mk in rec_m: existing['monthly'][mk] = rec_m[mk]
            if mk in med_m: existing['monthly_medicos'][mk] = med_m[mk]
        # Recompute totals
        existing['total'] = sum(v for v in existing['monthly'].values() if v)
        existing['total_medicos'] = sum(v for v in existing.get('monthly_medicos',{}).values() if v)
        # Quarterly Q1 2026
        q1_rec = sum(rec_m.get(m,0) or 0 for m in Q1)
        q1_med = sum(med_m.get(m,0) or 0 for m in Q1)
        existing.setdefault('quarterly', {})['Q1 2026'] = q1_rec
        existing.setdefault('quarterly_medicos', {})['Q1 2026'] = q1_med
        print(f'  rec_comp.MOMETAX.{dermato_brand}: total={existing["total"]}, monthly +={list(months)}')

    # 3) recetas[MOMETAX] = total mercado (droga MOMETASONA totales)
    recetas = D.setdefault('recetas', {}).setdefault('MOMETAX', {})
    for mk in months:
        recetas[mk] = {
            'recetas': data['totales_droga']['rec'].get(mk, 0),
            'medicos': data['totales_droga']['med'].get(mk, 0),
        }

    if args.dry_run:
        print('\nDRY RUN: no se escribio.')
        return 0
    new_text = text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    HTML.write_text(new_text, encoding='utf-8', newline='')
    print(f'\n-> {HTML} reescrito ({HTML.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
