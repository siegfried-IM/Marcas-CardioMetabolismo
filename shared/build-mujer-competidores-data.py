#!/usr/bin/env python3
"""
shared/build-mujer-competidores-data.py

Lee el xlsx 'Producto-Molécula-ATC-provincia - 11 de mayo de 2026.xlsx'
y construye mujer/DDD/competidores-data.js con shape A:

  window.SFG_COMP_DATA = {
    months: ['Ene-2025', ..., 'Mar-2026'],
    regions: [...43 regions...],
    markets: {
      <mercado_IQVIA>: {
        brands: [list of SIE brand names en este mercado],
        total_monthly: { region: [N monthly] },
        brand_monthly: {
          <brand>: { region: [N monthly] }
        }
      }
    }
  }

Esto reemplaza el shape B (que solo tenia SIE vs total) por shape A
con granularidad competidor x region, igual que cardio/SNC/dermato.
"""
from __future__ import annotations
import re, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

XLSX = Path(r'C:\Users\camarinaro\Downloads\Producto-Molécula-ATC-provincia - 11 de mayo de 2026.xlsx')
OUT = Path(r'C:\Users\camarinaro\Marcas-CardioMetabolismo\mujer\DDD\competidores-data.js')

MONTH_ORDER = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
               'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

# Words that signal end-of-brand (case-insensitive). After these, parsing stops.
PACKAGE_TOKENS = {
    'TABL','TABLE','TABLET','TABLETA','TABLETAS',
    'COMP','COMPR','COMPS','COMPRIMIDO','COMPRIMIDOS',
    'CAP','CAPS','CAPSULA','CAPSULAS',
    'AMP','AMPOLLA','AMPOLLAS',
    'JBE','JARABE',
    'INY','INYECTABLE','V.IM',
    'SUSP','SUSPENSION',
    'POL','POLVO',
    'LIQ','LIQUIDO',
    'GTAS','GOTAS',
    'SOL','SOLUC','SOLUCION','SOLUCIONES',
    'CR','CREMA','CREMAS',
    'UNG','UNGUENTO',
    'SOB','SOBRE','SOBRES',
    'BOLS','BOLSA',
    'EFER','EFERV','EFERVES','EFERVESCENTE','EFERVESCENTES',
    'DESLEI','DESLEIBLES','DESLEÍBLES',
    'REC','RECUB','RECUBIE','RECUBIERTO','RECUBIERTOS','RECUBIERTAS',
    'IM','PED','AD','ADULT','INF',
    'SPRAY','PVO','PLV',
    'DISP','DISPERS','DISPERSABLE','DISPERSABLES',
    'MAST','MAS','MAST.',
    'GRAG','GRAGEAS',
    'INHAL','INHALACION',
    'OFT','GINEC','VAGINAL','OVUL','OVULO','OVULOS','PESARIO',
    'GEL','EMUL','LOC','LOCION',
    'COLIRIO','COLIR','PARCHE','APOSITO',
    'AERO','AEROSOL',
    'ANILLO','ANILLOS',
    'CHEW','MAS','CHICLE',
}


def extract_brand(producto: str) -> str:
    """Heuristic: take prefix uppercase tokens before any packaging/dose marker."""
    if not producto: return 'UNKNOWN'
    s = str(producto).strip().upper().replace('\xa0', ' ')
    # Normalize multiple spaces
    s = re.sub(r'\s+', ' ', s)
    parts = s.split(' ')
    brand = []
    for i, p in enumerate(parts):
        clean = re.sub(r'^[.,;:()\[\]/\\-]+|[.,;:()\[\]/\\-]+$', '', p)
        if not clean: continue
        # Stop at "X" (pack indicator like "X 28")
        if clean == 'X' and i > 0: break
        # Stop at packaging token
        if clean in PACKAGE_TOKENS: break
        # Stop if pure dose (e.g. "100MG", "0.5%", "5ML", "2.5MG")
        if re.match(r'^[\d.,]+([A-Z%]+)?$', clean):
            # Allow small numerics as part of brand if 1-2 digits and we already have ≥1 alpha token
            if len(brand) >= 1 and re.match(r'^\d{1,2}$', clean):
                brand.append(clean)
                continue
            break
        # Stop if "MG/", "MCG", etc. (mid-pack indicators)
        if re.match(r'^\d+(\.\d+)?(MG|ML|G|MCG|UI|%|MG/ML|UI/ML).*', clean): break
        brand.append(clean)
    out = ' '.join(brand) if brand else parts[0] if parts else 'UNKNOWN'
    # Normalize known SIE brand variants
    norm_map = {
        'TRIP +45': 'TRIP +45',
        'SIDER': 'SIDERBLUT',
        'ISIS PROMOCION': 'ISIS',
        'ISIS PROMOCIÓN': 'ISIS',
    }
    return norm_map.get(out, out)


SIE_BRAND_PATTERNS = [
    r'^ISIS\b',
    r'^SIDERBLUT\b', r'^SIDER\b',
    r'^TRIP\b',
    r'^CALCIO BASE\b', r'^CALCIO CITRATO\b',
    r'^CLIMATIX\b',
    r'^DELTROX\b',
    r'^GYNODERM\b',
    r'^ROXOLAN\b',
    r'^ALUMPAK\b',
]


def is_sie_brand(brand: str) -> bool:
    s = str(brand).upper()
    return any(re.match(p, s) for p in SIE_BRAND_PATTERNS)


def month_sort_key(mk: str) -> int:
    parts = mk.split('-')
    return int(parts[1]) * 100 + MONTH_ORDER.get(parts[0], 0)


def main():
    if not XLSX.is_file():
        print(f'ERROR: {XLSX} no existe', file=sys.stderr); return 2

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {str(h or '').strip(): i for i, h in enumerate(hdr)}
    c_region = cols.get('RegionCUP', 0)
    c_market = cols.get('Mercado', 1)
    c_mes    = cols.get('AñoMes', 4)
    c_prod   = cols.get('Producto', 7)
    c_unid   = cols.get('Unidades', 8)

    months_set = set()
    regions_set = set()
    # data[market][brand][region][month] = units
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    # market_totals[market][region][month] = units (sum across all products)
    mkt_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    # market_brand_is_sie[market][brand] = bool
    sie_flag = defaultdict(dict)

    n_rows = 0
    n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        if not row: continue
        region = row[c_region]
        market = row[c_market]
        mes    = row[c_mes]
        prod   = row[c_prod]
        unid   = row[c_unid]
        if not region or not market or not mes or not prod: continue
        region = str(region).strip()
        if region in ('Totales', '-'): continue
        market = str(market).strip()
        mes = str(mes).strip()
        try: u = int(round(float(unid or 0)))
        except (TypeError, ValueError): u = 0
        if u <= 0: continue
        brand = extract_brand(prod)
        if not brand or brand == 'UNKNOWN': continue
        data[market][brand][region][mes] += u
        mkt_totals[market][region][mes] += u
        if brand not in sie_flag[market]:
            sie_flag[market][brand] = is_sie_brand(brand)
        regions_set.add(region)
        months_set.add(mes)
        n_kept += 1
    wb.close()
    print(f'  rows leidas: {n_rows:,}, kept: {n_kept:,}')

    months = sorted(months_set, key=month_sort_key)
    regions = sorted(regions_set, key=lambda r: (r.startswith('_'), r))
    print(f'  months ({len(months)}): {months}')
    print(f'  regions: {len(regions)}, markets: {len(data)}')

    # Build final structure
    out = {
        'months': months,
        'regions': regions,
        'markets': {},
    }
    for market in sorted(data.keys()):
        bm = {}  # brand_monthly[brand][region] = [N month values]
        for brand in sorted(data[market].keys()):
            bm[brand] = {}
            for region in regions:
                arr = [data[market][brand].get(region, {}).get(mk, 0) for mk in months]
                # Skip regions with all zeros to save space
                if any(arr):
                    bm[brand][region] = arr
        tm = {}  # total_monthly[region] = [N month values]
        for region in regions:
            arr = [mkt_totals[market].get(region, {}).get(mk, 0) for mk in months]
            if any(arr):
                tm[region] = arr
        sie_brands = sorted([b for b, is_s in sie_flag[market].items() if is_s])
        out['markets'][market] = {
            'brands': sie_brands,
            'brand_monthly': bm,
            'total_monthly': tm,
        }
        print(f'    [{market}]: {len(bm)} brands ({len(sie_brands)} SIE), {len(tm)} regs with data')

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        f'window.SFG_COMP_DATA = {json.dumps(out, ensure_ascii=False)};\n',
        encoding='utf-8', newline=''
    )
    print(f'\nSaved: {OUT} ({OUT.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
