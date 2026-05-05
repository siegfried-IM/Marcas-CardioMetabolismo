#!/usr/bin/env python3
"""
shared/update-dermato-precios.py

Actualiza precios de dermato desde Manual Farmaceutico.
Estructura de dermato es flat: precios[FAM][pres] = [entries],
distinta de cardio (precios[FAM].molecule[pres] = [entries]).
Por eso merge-precios.py no la maneja bien y se hace en script
dedicado.

Tambien actualiza meta.price_prev_label y price_curr_label.

Uso: py shared/update-dermato-precios.py [--pricefile <xlsx>] [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'dermatologia' / 'dermato_dashboard.html'
DEFAULT_FILE = Path(r'C:\Users\camarinaro\Downloads\Sin título - Tabla - 4 de mayo de 2026.xlsx')


def norm(s):
    """Normalizacion: extrae tokens significativos (forma_farma + numero+unidad)."""
    if s is None: return ''
    s = str(s).strip().lower()
    s = s.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
    # Primero quitar puntos y comas (los unimos a espacios)
    s = re.sub(r'[\.,/?¿]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    # Reemplazar variantes de forma farmaceutica a una forma canonica
    # Importante: usar \s+ para flexibilidad en espacios
    s = re.sub(r'\bcomp\s*rec\b', 'comp', s)
    s = re.sub(r'\bcomp\b', 'comp', s)
    s = re.sub(r'\bcaps?\b', 'cap', s)
    s = re.sub(r'\bung\b', 'ung', s)
    s = re.sub(r'\bcr\b', 'crema', s)
    s = re.sub(r'\bcrema\b', 'crema', s)
    s = re.sub(r'\bloc\b', 'loc', s)
    s = re.sub(r'\blocion\b', 'loc', s)
    s = re.sub(r'\bpvo\b', 'polvo', s)
    s = re.sub(r'\bpolvo\b', 'polvo', s)
    s = re.sub(r'\bgel\b', 'gel', s)
    s = re.sub(r'\bespuma\b', 'espuma', s)
    s = re.sub(r'\baer(osol)?\b', 'aerosol', s)
    s = re.sub(r'\blaca\b', 'laca', s)
    s = re.sub(r'\bpasta\b', 'pasta', s)
    s = re.sub(r'\bderm\b', '', s)
    s = re.sub(r'\bclotrimazol\b', '', s)  # nombres droga en SIE pres
    s = re.sub(r'\bciclopirox\b', '', s)
    # Tokenizar
    tokens = re.findall(r'[a-z]+|\d+(?:\.\d+)?(?:%)?', s)
    return ' '.join(sorted(tokens))


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_prod = col_pres = col_lab = col_prev = col_curr = None
    label_prev = label_curr = None
    for i, h in enumerate(row1):
        if not h: continue
        s = str(h).strip()
        sl = s.lower()
        if 'producto' in sl: col_prod = i
        elif 'presentacion' in sl: col_pres = i
        elif 'laboratorio' in sl: col_lab = i
        elif s.startswith('PVP') and 'al' in s:
            if col_prev is None:
                col_prev = i; label_prev = s
            else:
                col_curr = i; label_curr = s
    if not all([col_prod is not None, col_pres is not None, col_prev is not None, col_curr is not None]):
        raise RuntimeError(f'No se encontraron columnas: prod={col_prod}, pres={col_pres}, prev={col_prev}, curr={col_curr}')

    # lookup[(prod_norm, pres_norm)] = (pvp_prev, pvp_curr)
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        prod = row[col_prod] if col_prod < len(row) else None
        pres = row[col_pres] if col_pres < len(row) else None
        if not prod or not pres: continue
        pvp_prev = row[col_prev] if col_prev < len(row) else None
        pvp_curr = row[col_curr] if col_curr < len(row) else None
        try: pvp_prev = float(pvp_prev) if pvp_prev is not None else None
        except: pvp_prev = None
        try: pvp_curr = float(pvp_curr) if pvp_curr is not None else None
        except: pvp_curr = None
        if pvp_prev is None and pvp_curr is None: continue
        key = (norm(prod), norm(pres))
        lookup[key] = (pvp_prev, pvp_curr)
    wb.close()
    return lookup, label_prev, label_curr


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pricefile', default=str(DEFAULT_FILE))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pf = Path(args.pricefile)
    if not pf.is_file():
        print(f'ERROR: no existe {pf}', file=sys.stderr); return 2
    print(f'Leyendo: {pf}')
    lookup, label_prev, label_curr = parse_xlsx(pf)
    print(f'  {len(lookup)} entries en pricefile')
    print(f'  label_prev: {label_prev}')
    print(f'  label_curr: {label_curr}')

    text = HTML.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    abs_start = m.start() + len('const D = ')
    abs_start = text.index('{', abs_start)
    D, end = json.JSONDecoder().raw_decode(text[abs_start:])
    abs_end = abs_start + end

    matched = unmatched = 0
    for fam, fdata in D['precios'].items():
        for pres, entries in fdata.items():
            for e in entries:
                key = (norm(e.get('prod')), norm(pres))
                if key in lookup:
                    pp, pc = lookup[key]
                    if pp is not None: e['pvp_dic25'] = pp
                    if pc is not None: e['pvp_feb26'] = pc
                    if pp and pc and pp > 0:
                        e['var'] = (pc - pp) / pp
                    matched += 1
                else:
                    unmatched += 1

    print(f'\nMatched: {matched}, Unmatched: {unmatched}')

    # Update meta labels
    meta = D.setdefault('meta', {})
    meta['price_prev_label'] = label_prev
    meta['price_curr_label'] = label_curr
    print(f'meta updated: prev={label_prev}, curr={label_curr}')

    if args.dry_run:
        print('DRY RUN: no se escribio.')
        return 0
    new_text = text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    HTML.write_text(new_text, encoding='utf-8', newline='')
    print(f'-> {HTML} reescrito ({HTML.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
