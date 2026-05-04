#!/usr/bin/env python3
"""
shared/fix-mujer-d3-comp.py

Pobla mujer.rec_comp.D3 con los competidores de la familia D3 (VITAMINA D3 =
COLECALCIFEROL) que actualmente esta vacio, lo que provoca "Sin datos de
competidores" en el dashboard.

Lee la pivot de CloseUp (Mes-Año + Mercado/Droga/Marca + Cant. Recetas/Medicos)
y filtra:
  Mercado = "VITAMINA D (TRIP D3)"
  Droga   = "VITAMINA D3 = COLECALCIFEROL"
  Marca   != "Totales" ni vacio

Construye para cada brand: { is_sie, monthly, quarterly, total,
monthly_medicos, quarterly_medicos, total_medicos } igual que el resto de
las familias.

Tambien:
  - Actualiza rec_ms.D3.sie con Mar 2026 sumando los SIE en la familia
    (TRIP D3 SIE + CALCITOL D3 SIE).
  - Recomputa rec_ms.D3.ms para Ene/Feb/Mar 2026 = sie/total_droga * 100.
  - Recomputa rec_ms.D3.ms_quarterly Q1 2026.
  - NO toca las otras 9 familias.
  - NO toca historico (Jan 2025 - Dec 2025) — el pivot solo tiene 2026.

Uso:
    py shared/fix-mujer-d3-comp.py [--pivot <archivo.xlsx>] [--dry-run]
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

DEFAULT_PIVOT = r'C:\Users\camarinaro\Downloads\Sin título - Tabla dinámica - 4 de mayo de 2026 (3).xlsx'
DEFAULT_HTML = r'C:\Users\camarinaro\Marcas-CardioMetabolismo\mujer\index.html'

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

MERCADO = 'VITAMINA D (TRIP D3)'
DROGA   = 'VITAMINA D3 = COLECALCIFEROL'

# SIE products dentro de la familia D3 (colecalciferol)
SIE_BRANDS = {'TRIP D3 SIE', 'CALCITOL D3 SIE'}


def quarter_key(month_key: str) -> str:
    """`Jan 2026` -> `Q1 2026`."""
    parts = month_key.split()
    if len(parts) != 2: return ''
    mes, year = parts
    inv = {v: k for k, v in MES_EN.items()}
    m = inv.get(mes, 0)
    if not m: return ''
    q = (m - 1) // 3 + 1
    return f'Q{q} {year}'


def to_int(v):
    if v is None or v == '-' or v == '': return 0
    try: return int(round(float(v)))
    except (ValueError, TypeError): return 0


def parse_pivot(pivot_path):
    """Devuelve (months_in_order, brand_data) where:
       brand_data[brand] = {
         'monthly': {mk: rec},
         'monthly_medicos': {mk: med},
         'is_sie': bool,
       }
       Y tambien droga_totals[mk] = recetas (totales del droga colecalciferol)."""
    wb = openpyxl.load_workbook(pivot_path, read_only=True, data_only=True)
    ws = wb.active

    # Row 1: month headers (datetime); row 2: metric labels
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    # col_idx -> (mk, metric)  metric in {recetas, medicos}
    col_map = {}
    months_seen = []
    for i, h1 in enumerate(row1):
        if not isinstance(h1, datetime): continue
        mk = f'{MES_EN[h1.month]} {h1.year}'
        if mk not in months_seen:
            months_seen.append(mk)
        h2 = (str(row2[i]) if i < len(row2) and row2[i] else '').strip().lower()
        if 'recetas' in h2:
            col_map[i] = (mk, 'recetas')
        elif 'm' in h2 and ('dico' in h2 or 'edico' in h2):
            col_map[i] = (mk, 'medicos')

    brand_data = {}      # marca -> {monthly, monthly_medicos, is_sie}
    droga_totals = {}    # mk -> recetas totales del droga COLECALCIFEROL

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3: continue
        mercado = (row[0] or '').strip()
        droga   = (row[1] or '').strip()
        marca   = (row[2] or '').strip()
        if mercado != MERCADO: continue
        if droga != DROGA: continue

        if marca == 'Totales':
            # Total del droga (COLECALCIFEROL)
            for i, (mk, metric) in col_map.items():
                if metric != 'recetas': continue
                if i >= len(row): continue
                droga_totals[mk] = droga_totals.get(mk, 0) + to_int(row[i])
            continue

        if not marca: continue

        b = brand_data.setdefault(marca, {
            'monthly': {}, 'monthly_medicos': {},
            'is_sie': marca in SIE_BRANDS,
        })
        for i, (mk, metric) in col_map.items():
            if i >= len(row): continue
            v = to_int(row[i])
            if metric == 'recetas':
                b['monthly'][mk] = b['monthly'].get(mk, 0) + v
            else:
                b['monthly_medicos'][mk] = b['monthly_medicos'].get(mk, 0) + v

    wb.close()
    return months_seen, brand_data, droga_totals


def aggregate(monthly: dict) -> tuple[dict, int]:
    quarterly = defaultdict(int)
    total = 0
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk:
            quarterly[qk] += v
        total += v
    return dict(quarterly), total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', default=DEFAULT_PIVOT)
    ap.add_argument('--html', default=DEFAULT_HTML)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pf = Path(args.pivot)
    if not pf.is_file():
        print(f'ERROR: pivot no existe: {pf}', file=sys.stderr); return 2

    print(f'Leyendo: {pf}')
    months, brand_data, droga_totals = parse_pivot(pf)
    print(f'  Meses: {months}')
    print(f'  Brands extraidos: {len(brand_data)}')
    for mk in months:
        print(f'  Total droga {DROGA} {mk}: {droga_totals.get(mk, 0):,}')

    # Mostrar SIE detectados
    sie_found = [b for b, d in brand_data.items() if d['is_sie']]
    print(f'  SIE detectados en familia: {sie_found}')

    # Cargar HTML mujer
    hp = Path(args.html)
    text = hp.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    if not m:
        print('ERROR: no encontre const D = en mujer/index.html'); return 3
    obj_start = m.start() + len('const D = ')
    obj_start = text.index('{', obj_start)
    D, end = json.JSONDecoder().raw_decode(text[obj_start:])
    abs_end = obj_start + end

    # === 1) Poblar rec_comp.D3 ===
    rec_comp = D.setdefault('rec_comp', {})
    d3 = rec_comp.setdefault('D3', {})
    if d3:
        print(f'WARN: rec_comp.D3 ya tenia {len(d3)} brands -- se sobreescribe')
    rec_comp['D3'] = {}
    for marca, b in brand_data.items():
        monthly = b['monthly']
        monthly_med = b['monthly_medicos']
        quarterly, total = aggregate(monthly)
        quarterly_med, total_med = aggregate(monthly_med)
        rec_comp['D3'][marca] = {
            'is_sie': b['is_sie'],
            'monthly': monthly,
            'quarterly': quarterly,
            'total': total,
            'monthly_medicos': monthly_med,
            'quarterly_medicos': quarterly_med,
            'total_medicos': total_med,
        }

    # Ordenar por total desc para que aparezcan los mas grandes primero (cosmetico)
    rec_comp['D3'] = dict(sorted(
        rec_comp['D3'].items(), key=lambda kv: kv[1]['total'], reverse=True
    ))
    print(f'OK: rec_comp.D3 poblado con {len(rec_comp["D3"])} brands')

    # === 2) Recomputar rec_ms.D3 ===
    rec_ms = D.setdefault('rec_ms', {})
    d3_ms = rec_ms.setdefault('D3', {'sie': {}, 'ms': {}, 'ms_quarterly': {}})
    sie_dict = d3_ms.setdefault('sie', {})
    ms_dict  = d3_ms.setdefault('ms', {})
    msq_dict = d3_ms.setdefault('ms_quarterly', {})

    # Sumar SIE por mes desde brand_data
    sie_monthly_from_pivot = defaultdict(int)
    for marca, b in brand_data.items():
        if b['is_sie']:
            for mk, v in b['monthly'].items():
                sie_monthly_from_pivot[mk] += v

    # Solo agregar Mar 2026 (Jan/Feb ya estan con valores TRIP D3 SIE solo;
    # los conservamos para no romper continuidad historica)
    months_to_add = [mk for mk in months if mk not in sie_dict]
    for mk in months_to_add:
        sie_dict[mk] = sie_monthly_from_pivot.get(mk, 0)
        print(f'  rec_ms.D3.sie += {mk}: {sie_dict[mk]}')

    # Recomputar ms% para los meses del pivot (con droga_totals)
    for mk in months:
        sie_v = sie_dict.get(mk, 0)
        droga_v = droga_totals.get(mk, 0)
        if droga_v > 0:
            ms_pct = round(sie_v / droga_v * 100, 2)
            ms_dict[mk] = ms_pct
            print(f'  rec_ms.D3.ms[{mk}] = {sie_v}/{droga_v} = {ms_pct}%')

    # ms_quarterly: recomputar Q1 2026 desde monthly
    q1_sie = sum(sie_dict.get(mk, 0) for mk in ['Jan 2026', 'Feb 2026', 'Mar 2026'])
    q1_droga = sum(droga_totals.get(mk, 0) for mk in ['Jan 2026', 'Feb 2026', 'Mar 2026'])
    if q1_droga > 0:
        msq_dict['Q1 2026'] = round(q1_sie / q1_droga * 100, 2)
        print(f'  rec_ms.D3.ms_quarterly[Q1 2026] = {q1_sie}/{q1_droga} = {msq_dict["Q1 2026"]}%')

    # Re-ordenar las claves para mantener orden cronologico (importante para charts)
    def sort_months(d):
        items = list(d.items())
        items.sort(key=lambda kv: month_sort_value(kv[0]))
        d.clear(); d.update(items)

    def month_sort_value(mk):
        parts = str(mk).split()
        if len(parts) != 2: return 0
        inv = {v: k for k, v in MES_EN.items()}
        return int(parts[1]) * 100 + inv.get(parts[0], 0)

    sort_months(sie_dict)
    sort_months(ms_dict)

    # === 3) Serializar ===
    if args.dry_run:
        print('\nDRY RUN: nada se escribio.')
        return 0

    new_text = text[:obj_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    hp.write_text(new_text, encoding='utf-8', newline='')
    print(f'\n-> {hp} reescrito ({hp.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
