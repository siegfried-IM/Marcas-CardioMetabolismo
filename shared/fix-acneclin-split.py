#!/usr/bin/env python3
"""
shared/fix-acneclin-split.py

Separa ACNECLIN y ACNECLIN AP en mol_perf.MINOCYCLINE.
En IQVIA ambos estan bajo el mismo product name 'ACNECLIN (SIE)' pero
son dos packs distintos:
  - ACNECLIN CAPS A.P 100MG x 30  -> ACNECLIN AP (SIE)
  - ACNECLIN TABL RECUBIE 50mg x 30 -> ACNECLIN (SIE)

El script:
1. Lee AR_PM master, separa por pack, computa ytd/mat/ms por producto
2. Reemplaza la entrada 'ACNECLIN (SIE)' en mol_perf.MINOCYCLINE.products
   con dos entradas separadas
3. Actualiza brandKpis para ACNECLIN y ACNECLIN AP con los valores correctos
4. Actualiza BUD_IQVIA['ACNECLIN AP'] de 'ACNECLIN (SIE)' a 'ACNECLIN AP (SIE)'

NO toca ninguna otra linea ni seccion.
"""
from __future__ import annotations
import re, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

IQVIA_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Apr-27-2026.xlsx')
HTML_PATH = Path(r'C:\Users\camarinaro\Marcas-CardioMetabolismo\dermatologia\dermato_dashboard.html')

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
MES_EN = {v:k for k,v in MES_INV.items()}
CIERRE_M = 3
CIERRE_Y = 2026


def msort(mk):
    p = mk.split()
    return int(p[1]) * 100 + MES_INV.get(p[0], 0) if len(p) == 2 else 0


def agg_ytd(mv, cm):
    by_y = defaultdict(int)
    for mk, v in mv.items():
        p = mk.split()
        m = MES_INV.get(p[0]) if len(p) == 2 else None
        if m and m <= cm:
            by_y[p[1]] += int(v or 0)
    return {f'{MES_EN[cm]} {y}': v for y, v in by_y.items()}


def agg_mat_yearly(mv, cm):
    years = set(mk.split()[1] for mk in mv if len(mk.split()) == 2)
    out = {}
    for yr in sorted(years, key=int):
        y = int(yr)
        total = 0
        for back in range(12):
            idx = y * 12 + (cm - 1) - back
            yy, mm = divmod(idx, 12)
            total += int(mv.get(f'{MES_EN[mm + 1]} {yy}', 0) or 0)
        out[f'{MES_EN[cm]} {yr}'] = total
    return out


def agg_mat_monthly(mv):
    sk = sorted(mv.keys(), key=msort)
    out = {}
    for i, mk in enumerate(sk):
        if i < 11:
            continue
        out[mk] = sum(int(mv.get(m, 0) or 0) for m in sk[i - 11:i + 1])
    return out


def agg_quarterly(mv):
    out = defaultdict(int)
    for mk, v in mv.items():
        p = mk.split()
        m = MES_INV.get(p[0]) if len(p) == 2 else None
        if m:
            out[f'Q{(m - 1) // 3 + 1} {p[1]}'] += int(v or 0)
    return dict(out)


def recompute_ms(p, fam_monthly, fam_ytd, fam_mat, fam_quarterly):
    mv = p.get('monthly_vals', {})
    qv = p.get('quarterly_vals', {})
    yv = p.get('ytd', {})
    mtv = p.get('mat', {})
    p['ms_monthly'] = {mk: round(mv.get(mk, 0) / fv * 100, 2) if fv > 0 else 0
                       for mk, fv in fam_monthly.items()}
    p['ms_quarterly'] = {qk: round(qv.get(qk, 0) / fv * 100, 2) if fv > 0 else 0
                         for qk, fv in fam_quarterly.items()}
    p['ms_ytd'] = {y: round(yv.get(y, 0) / fv * 100, 2) if fv > 0 else 0
                   for y, fv in fam_ytd.items()}
    p['ms_mat'] = {mk: round(mtv.get(mk, 0) / fv * 100, 2) if fv > 0 else 0
                   for mk, fv in fam_mat.items()}


def main():
    # ── Step 1: parse IQVIA master ──────────────────────────────────────
    if not IQVIA_MASTER.is_file():
        print(f'ERROR: no existe {IQVIA_MASTER}', file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(IQVIA_MASTER, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_product = 1
    col_pack = 2
    month_cols = []
    for i, h in enumerate(row1):
        if not h:
            continue
        s = str(h).strip()
        if not s.startswith('Units'):
            continue
        after = s.split('\n', 1)[-1].strip() if '\n' in s else s[5:].strip()
        if after.upper().startswith('MAT') or after.upper().startswith('YTD'):
            continue
        m = re.match(r'(\w+)\s+(\d{4})', after)
        if m and m.group(1) in MES_INV:
            month_cols.append((i, f'{m.group(1)} {m.group(2)}'))

    data_by_prod = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        prod = str(row[col_product] or '').strip()
        pack = str(row[col_pack] or '').strip().upper()
        if prod != 'ACNECLIN (SIE)':
            continue
        key = ('ACNECLIN AP (SIE)'
               if ('A.P' in pack or '100MG' in pack or '100 MG' in pack)
               else 'ACNECLIN (SIE)')
        for ci, mk in month_cols:
            if ci >= len(row):
                continue
            v = row[ci]
            if v is None:
                continue
            try:
                data_by_prod[key][mk] += float(v)
            except (ValueError, TypeError):
                pass
    wb.close()

    split_prods = {}
    for key, raw_mv in data_by_prod.items():
        mv = {k: int(round(v)) for k, v in raw_mv.items()}
        ytd = agg_ytd(mv, CIERRE_M)
        mat = {**agg_mat_yearly(mv, CIERRE_M), **agg_mat_monthly(mv)}
        split_prods[key] = {'monthly_vals': mv, 'quarterly_vals': agg_quarterly(mv), 'ytd': ytd, 'mat': mat}
        print(f'  {key}: ytd_mar26={ytd.get("Mar 2026")}, mat_mar26={mat.get("Mar 2026")}')

    # ── Step 2: load HTML + D ────────────────────────────────────────────
    text = HTML_PATH.read_text(encoding='utf-8', errors='replace')
    m_obj = re.search(r'const D = (\{)', text)
    if not m_obj:
        print('ERROR: const D not found', file=sys.stderr)
        return 3
    obj_start = m_obj.start() + len('const D = ')
    obj_start = text.index('{', obj_start)
    D, end_off = json.JSONDecoder().raw_decode(text[obj_start:])
    abs_end = obj_start + end_off

    # ── Step 3: update mol_perf.MINOCYCLINE ─────────────────────────────
    mino = D['mol_perf']['MINOCYCLINE']

    # Family totals (unchanged: sum of all products)
    fam_monthly: dict = defaultdict(int)
    for p in mino['products']:
        for mk, v in p.get('monthly_vals', {}).items():
            fam_monthly[mk] += int(v or 0)
    fam_monthly = dict(fam_monthly)
    fam_ytd = agg_ytd(fam_monthly, CIERRE_M)
    fam_mat = {**agg_mat_yearly(fam_monthly, CIERRE_M), **agg_mat_monthly(fam_monthly)}
    fam_quarterly = agg_quarterly(fam_monthly)

    new_products = []
    for p in mino['products']:
        if p['prod'] == 'ACNECLIN (SIE)':
            # Replace with ACNECLIN AP (SIE) first, then ACNECLIN (SIE)
            for new_name in ['ACNECLIN AP (SIE)', 'ACNECLIN (SIE)']:
                new_p = {k: v for k, v in p.items()}
                new_p['prod'] = new_name
                new_p['is_sie'] = True
                sd = split_prods[new_name]
                new_p['monthly_vals'] = sd['monthly_vals']
                new_p['quarterly_vals'] = sd['quarterly_vals']
                new_p['ytd'] = sd['ytd']
                new_p['mat'] = sd['mat']
                recompute_ms(new_p, fam_monthly, fam_ytd, fam_mat, fam_quarterly)
                new_products.append(new_p)
        else:
            recompute_ms(p, fam_monthly, fam_ytd, fam_mat, fam_quarterly)
            new_products.append(p)

    mino['products'] = new_products
    print(f'mol_perf MINOCYCLINE products: {[p["prod"] for p in new_products]}')

    # ── Step 4: update brandKpis ─────────────────────────────────────────
    bk = D.get('brandKpis', {})
    fam_ytd_v = fam_ytd.get('Mar 2026')
    fam_mat_v = fam_mat.get('Mar 2026')
    fam_ytd_p = fam_ytd.get('Mar 2025')
    fam_mat_p = fam_mat.get('Mar 2025')

    for brand, sie_name in [('ACNECLIN AP', 'ACNECLIN AP (SIE)'), ('ACNECLIN', 'ACNECLIN (SIE)')]:
        if brand not in bk:
            continue
        sp = split_prods.get(sie_name, {})
        ytd_u = sp.get('ytd', {}).get('Mar 2026')
        ytd_prev = sp.get('ytd', {}).get('Mar 2025')
        mat_u = sp.get('mat', {}).get('Mar 2026')
        mat_prev = sp.get('mat', {}).get('Mar 2025')
        ie_ytd = round(ytd_u / ytd_prev * 100, 1) if ytd_u and ytd_prev else None
        ms_ytd = round(ytd_u / fam_ytd_v * 100, 1) if ytd_u and fam_ytd_v else None
        ie_mat = round(mat_u / mat_prev * 100, 1) if mat_u and mat_prev else None
        ms_mat = round(mat_u / fam_mat_v * 100, 1) if mat_u and fam_mat_v else None

        bk[brand]['ytd'].update({
            'ie': ie_ytd, 'ms': ms_ytd, 'units': ytd_u, 'units_prev': ytd_prev,
            'market_total': fam_ytd_v,
            'growth': round(ie_ytd - 100, 1) if ie_ytd else None
        })
        bk[brand]['mat'].update({
            'ie': ie_mat, 'ms': ms_mat, 'units': mat_u, 'units_prev': mat_prev,
            'market_total': fam_mat_v,
            'growth': round(ie_mat - 100, 1) if ie_mat else None
        })
        print(f'brandKpis {brand}: ytd ie={ie_ytd} ms={ms_ytd}%, mat ie={ie_mat} ms={ms_mat}%')

    # ── Step 5: serialize + fix BUD_IQVIA ────────────────────────────────
    new_text = text[:obj_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]

    # BUD_IQVIA: "ACNECLIN AP":"ACNECLIN (SIE)" -> "ACNECLIN AP (SIE)"
    new_text = new_text.replace(
        '"ACNECLIN AP":"ACNECLIN (SIE)"',
        '"ACNECLIN AP":"ACNECLIN AP (SIE)"'
    )
    if '"ACNECLIN AP":"ACNECLIN AP (SIE)"' in new_text:
        print('BUD_IQVIA["ACNECLIN AP"] actualizado OK')
    else:
        print('WARN: BUD_IQVIA pattern no encontrado')

    HTML_PATH.write_text(new_text, encoding='utf-8', newline='')
    print(f'Saved: {HTML_PATH.stat().st_size:,} bytes')
    return 0


if __name__ == '__main__':
    sys.exit(main())
