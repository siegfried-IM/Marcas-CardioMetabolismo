#!/usr/bin/env python3
"""
shared/merge-precios.py

Mergea actualizacion de precios desde el dump del Manual Farmaceutico
(formato Excel con columnas: Producto, Presentacion, Droga, Laboratorio,
PVP al <fecha_prev>, PVP al <fecha_curr>, % Var, Fecha Vigencia).

Para cada linea (cardio/ATB/OTC/respi) actualiza in-place:
  - precios[F][molecule|atc][pres][i].pvp_dic25  (slot "previo")
  - precios[F][molecule|atc][pres][i].pvp_feb26  (slot "actual")
  - precios[F][molecule|atc][pres][i].var
  - meta.price_prev_label = "PVP al <fecha_prev>"
  - meta.price_curr_label = "PVP al <fecha_curr>"

Match por (producto, presentacion). Lab solo informativo si hace falta
desempatar. Productos no encontrados en el archivo conservan los precios
viejos sin tocar.

NO toca recetas, mol_perf, budget, stock, convenios, canales, etc.
NO toca SNC ni mujer (estructuras distintas).

Uso:
    py shared/merge-precios.py --pricefile "<archivo.xlsx>"
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

LINES_DEFAULT = ['cardio', 'ATB', 'OTC', 'respiratorio']
INLINE_LINES = {'mujer', 'SNC'}  # tienen `const D = {...};` en lugar de window.OTC_*

# Para mujer, los productos en data.js tienen sufijos como "(SIE)", "(ELE)",
# "(GAD)" que no estan en el archivo de precios. Los stripeamos al matchear.
SUFFIX_STRIP_RE = re.compile(r'\s*\([A-Z\-+0-9]{2,5}\)\s*$')


def normalize_text(s):
    if s is None: return ''
    return re.sub(r'\s+', ' ', str(s).upper().strip())


def normalize_pres(s):
    """Normaliza presentacion: minuscula, espacios colapsados, sin acentos comunes."""
    if s is None: return ''
    s = str(s).strip().lower()
    # Colapsa whitespace
    s = re.sub(r'\s+', ' ', s)
    # Quita acentos comunes (caps. cáps. cps. → mismo)
    s = s.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
    return s


def parse_pricefile(path):
    """Devuelve:
       (lookup, prev_label, curr_label)
       lookup = { (prod_norm, pres_norm) : (pvp_prev, pvp_curr, var) }
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    # Find the two PVP columns and var
    col_prev = col_curr = col_var = None
    label_prev = label_curr = None
    for i, h in enumerate(headers):
        s = str(h or '').strip()
        if not s.startswith('PVP'): continue
        if col_prev is None:
            col_prev = i; label_prev = s
        else:
            col_curr = i; label_curr = s
    if col_prev is None or col_curr is None:
        raise ValueError(f'No encontre 2 columnas PVP en headers: {headers}')
    for i, h in enumerate(headers):
        if str(h or '').lower().startswith('% var') or str(h or '').lower() == 'var':
            col_var = i; break

    # Find col indexes for Producto / Presentacion / Laboratorio
    col_prod = col_pres = col_lab = None
    for i, h in enumerate(headers):
        s = str(h or '').strip().lower()
        if s == 'producto': col_prod = i
        elif s == 'presentacion' or s == 'presentación': col_pres = i
        elif s == 'laboratorio': col_lab = i

    if col_prod is None or col_pres is None:
        raise ValueError(f'Faltan cols Producto/Presentacion. Headers: {headers}')

    lookup = {}
    lookup_by_prod = {}  # prod_norm -> [(pres_norm, (prev,curr,var))]
    n_total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(col_prev, col_curr): continue
        prod = row[col_prod]
        pres = row[col_pres]
        if not prod or not pres: continue
        try: pvp_prev = float(row[col_prev]) if row[col_prev] is not None else None
        except: pvp_prev = None
        try: pvp_curr = float(row[col_curr]) if row[col_curr] is not None else None
        except: pvp_curr = None
        if pvp_curr is None: continue
        try: var = float(row[col_var]) if col_var is not None and row[col_var] is not None else None
        except: var = None
        if var is None and pvp_prev:
            var = (pvp_curr - pvp_prev) / pvp_prev
        prod_n = normalize_text(prod)
        pres_n = normalize_pres(pres)
        triple = (pvp_prev, pvp_curr, var)
        lookup[(prod_n, pres_n)] = triple
        lookup_by_prod.setdefault(prod_n, []).append((pres_n, triple))
        n_total += 1

    wb.close()
    return lookup, lookup_by_prod, label_prev, label_curr, n_total


def parse_data_js(text):
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text)
    if not m1: raise ValueError('OTC_DATA not found')
    obj_start1 = text.index('{', m1.end())
    d1, end1 = json.JSONDecoder().raw_decode(text[obj_start1:])
    abs_end1 = obj_start1 + end1
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text[abs_end1:])
    if not m2:
        return d1, None
    obj_start2 = abs_end1 + text[abs_end1:].index('{', m2.end())
    d2, _ = json.JSONDecoder().raw_decode(text[obj_start2:])
    return d1, d2


def serialize_data_js(text_orig, d1, d2):
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text_orig)
    obj_start1 = text_orig.index('{', m1.end())
    _, end1 = json.JSONDecoder().raw_decode(text_orig[obj_start1:])
    abs_end1 = obj_start1 + end1
    prefix1 = text_orig[:obj_start1]
    if d2 is None:
        return prefix1 + json.dumps(d1, ensure_ascii=False) + text_orig[abs_end1:]
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text_orig[abs_end1:])
    obj_start2 = abs_end1 + text_orig[abs_end1:].index('{', m2.end())
    _, end2 = json.JSONDecoder().raw_decode(text_orig[obj_start2:])
    abs_end2 = obj_start2 + end2
    middle = text_orig[abs_end1:obj_start2]
    suffix = text_orig[abs_end2:]
    return prefix1 + json.dumps(d1, ensure_ascii=False) + middle + json.dumps(d2, ensure_ascii=False) + suffix


def update_entry(entry, lookup, pres_key, lookup_by_prod=None):
    """Actualiza un entry de producto si lo encuentra en el lookup.
    Estrategia de match (en orden):
      1) (prod limpio + pres = entry.pres si existe)  ← SNC con pres explicito
      2) (prod limpio + pres = pres_key)              ← cardio/ATB/respi/OTC
      3) lookup_by_prod[prod_limpio]: filtra rows donde la pres del archivo
         CONTIENE el pres_key (loose match para mujer, ej. key='x 28' matches
         'comp.rec.x 28' en archivo).
    Devuelve True si matched."""
    prod_raw = entry.get('prod', '')
    if not prod_raw: return False
    # Strip suffixes "(SIE)", "(ELE)", "(GAD)", etc.
    prod_clean = SUFFIX_STRIP_RE.sub('', str(prod_raw)).strip()
    prod_norm = normalize_text(prod_clean)
    pvp_prev = pvp_curr = var = None

    # 1) Si entry tiene pres explicito (caso SNC)
    if entry.get('pres'):
        key = (prod_norm, normalize_pres(entry['pres']))
        if key in lookup:
            pvp_prev, pvp_curr, var = lookup[key]

    # 2) Match estandar por (prod, pres_key)
    if pvp_curr is None:
        key = (prod_norm, normalize_pres(pres_key))
        if key in lookup:
            pvp_prev, pvp_curr, var = lookup[key]

    # 3) Loose match para mujer: prod en lookup_by_prod, pres del archivo
    #    contiene el pres_key (ej. 'x 28' ⊂ 'comp.rec.x 28').
    #    Si el prod completo no esta, probamos truncando palabras del final
    #    (ej. "ISIS FREE S/ESTROG" -> "ISIS FREE" -> "ISIS").
    if pvp_curr is None and lookup_by_prod is not None:
        pres_norm = normalize_pres(pres_key)
        words = prod_norm.split()
        # Probamos del mas especifico (full) al mas generico (first word)
        for take in range(len(words), 0, -1):
            candidate_prod = ' '.join(words[:take])
            candidates = lookup_by_prod.get(candidate_prod, [])
            if not candidates: continue
            matches = [(p, val) for (p, val) in candidates if pres_norm in p]
            if len(matches) == 1:
                pvp_prev, pvp_curr, var = matches[0][1]
                break
            # Si hay multiples matches solo aceptamos si el prod es full
            # (no truncado) — para evitar matchear "ISIS" con multiples.
            if len(matches) > 1 and take == len(words):
                # Tomamos el primer match (ambiguous, pero al menos updatea uno)
                pvp_prev, pvp_curr, var = matches[0][1]
                break

    if pvp_curr is None:
        return False
    if pvp_prev is not None:
        entry['pvp_dic25'] = pvp_prev
    entry['pvp_feb26'] = pvp_curr
    if var is not None:
        entry['var'] = var
    return True


def parse_inline_d(text):
    """Para mujer/SNC: const D = {...};"""
    m = re.search(r'const\s+D\s*=\s*', text)
    if not m: return None
    obj_start = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[obj_start:])
    return obj_start, D, obj_start + end


def serialize_inline_d(text, obj_start, D, obj_end):
    return text[:obj_start] + json.dumps(D, ensure_ascii=False) + text[obj_end:]


def update_precios_obj(precios, lookup, lookup_by_prod):
    matched = 0
    not_matched = 0
    samples_unmatched = []
    if not isinstance(precios, dict): return matched, not_matched, samples_unmatched

    for fam, fam_obj in precios.items():
        if not isinstance(fam_obj, dict): continue
        keys = list(fam_obj.keys())
        if not keys: continue
        first_val = fam_obj[keys[0]]
        if isinstance(first_val, dict):
            for view_key in keys:
                view = fam_obj[view_key]
                if not isinstance(view, dict): continue
                for pres_key, items in view.items():
                    if not isinstance(items, list): continue
                    for entry in items:
                        if not isinstance(entry, dict): continue
                        if update_entry(entry, lookup, pres_key, lookup_by_prod):
                            matched += 1
                        else:
                            not_matched += 1
                            if len(samples_unmatched) < 5:
                                samples_unmatched.append(f"{entry.get('prod')} | {pres_key}")
        elif isinstance(first_val, list):
            for pres_key, items in fam_obj.items():
                if not isinstance(items, list): continue
                for entry in items:
                    if not isinstance(entry, dict): continue
                    if update_entry(entry, lookup, pres_key, lookup_by_prod):
                        matched += 1
                    else:
                        not_matched += 1
                        if len(samples_unmatched) < 5:
                            samples_unmatched.append(f"{entry.get('prod')} | {pres_key}")
    return matched, not_matched, samples_unmatched


def update_hardcoded_html_labels(text, label_prev, label_curr):
    """Reemplaza occurrences de 'PVP Dic 2025' / 'PVP Feb 2026' (hardcoded
    en mujer/SNC HTML) por los nuevos labels. Devuelve (new_text, n_replacements)."""
    n = 0
    for old in ['PVP Dic 2025', 'PVP Dic. 2025']:
        if old in text:
            n += text.count(old)
            text = text.replace(old, label_prev)
    for old in ['PVP Feb 2026', 'PVP Feb. 2026']:
        if old in text:
            n += text.count(old)
            text = text.replace(old, label_curr)
    return text, n


def merge_line(data_js_path, lookup, lookup_by_prod, label_prev, label_curr, dry_run=False):
    text = data_js_path.read_text(encoding='utf-8-sig', errors='replace')
    d1, d2 = parse_data_js(text)
    if d2 is None:
        return {'skipped': 'no OTC_DASHBOARD'}
    precios = d2.get('precios')
    if not isinstance(precios, dict):
        return {'skipped': 'no precios'}

    matched, not_matched, samples = update_precios_obj(precios, lookup, lookup_by_prod)

    d2.setdefault('meta', {})['price_prev_label'] = label_prev
    d2.setdefault('meta', {})['price_curr_label'] = label_curr

    if not dry_run:
        new_text = serialize_data_js(text, d1, d2)
        data_js_path.write_text(new_text, encoding='utf-8', newline='')

    return {'matched': matched, 'not_matched': not_matched, 'samples_unmatched': samples}


def merge_inline_line(html_path, lookup, lookup_by_prod, label_prev, label_curr, dry_run=False):
    """Para mujer/SNC con `const D = {...};` inline + labels hardcoded en HTML."""
    text = html_path.read_text(encoding='utf-8', errors='replace')
    pos = parse_inline_d(text)
    if not pos:
        return {'skipped': 'no const D'}
    obj_start, D, obj_end = pos
    precios = D.get('precios')
    if not isinstance(precios, dict):
        return {'skipped': 'no precios in inline D'}

    matched, not_matched, samples = update_precios_obj(precios, lookup, lookup_by_prod)
    D.setdefault('meta', {})['price_prev_label'] = label_prev
    D.setdefault('meta', {})['price_curr_label'] = label_curr

    new_text = serialize_inline_d(text, obj_start, D, obj_end)
    new_text, html_replacements = update_hardcoded_html_labels(new_text, label_prev, label_curr)

    if not dry_run:
        html_path.write_text(new_text, encoding='utf-8', newline='')

    return {
        'matched': matched, 'not_matched': not_matched,
        'samples_unmatched': samples, 'html_label_replacements': html_replacements,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pricefile', required=True)
    ap.add_argument('--repo', default=str(Path(__file__).resolve().parent.parent))
    ap.add_argument('--lines', nargs='+',
                    default=LINES_DEFAULT + ['mujer', 'SNC'],
                    help='Lineas a procesar. Para mujer/SNC se actualiza el inline D + labels hardcoded en HTML.')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pf = Path(args.pricefile)
    if not pf.is_file():
        print(f'ERROR: archivo no existe: {pf}', file=sys.stderr); return 2

    print(f'Leyendo: {pf}')
    lookup, lookup_by_prod, label_prev, label_curr, n_total = parse_pricefile(pf)
    print(f'  Filas con precio: {n_total}')
    print(f'  Productos unicos (prod, presentacion): {len(lookup)}')
    print(f'  Label previo:  {label_prev}')
    print(f'  Label actual:  {label_curr}')

    repo = Path(args.repo)
    print(f'\nMergeando precios en {len(args.lines)} lineas...')
    for line in args.lines:
        if line in INLINE_LINES:
            html_path = repo / line / 'index.html'
            extras = []
            # SNC tiene tambien psq_dashboard.html como copia legacy
            if line == 'SNC':
                extras.append(repo / 'SNC' / 'psq_dashboard.html')
            paths = [html_path] + extras
            for p in paths:
                if not p.is_file():
                    print(f'  [{line}/{p.name}] SKIP: no existe'); continue
                try:
                    res = merge_inline_line(p, lookup, lookup_by_prod, label_prev, label_curr, dry_run=args.dry_run)
                except Exception as e:
                    print(f'  [{line}/{p.name}] ERROR: {e}'); continue
                if 'skipped' in res:
                    print(f'  [{line}/{p.name}] SKIP: {res["skipped"]}'); continue
                m, nm = res['matched'], res['not_matched']
                total = m + nm
                pct = round(m * 100 / total, 1) if total else 0
                print(f'  [{line}/{p.name}] OK: {m}/{total} entries actualizados ({pct}%), {nm} sin match, {res["html_label_replacements"]} labels HTML reemplazados')
                for s in res['samples_unmatched'][:3]:
                    print(f'    sample no-match: {s}')
            continue

        data_js = repo / line / 'data.js'
        if not data_js.is_file():
            print(f'  [{line}] SKIP: no data.js'); continue
        try:
            res = merge_line(data_js, lookup, lookup_by_prod, label_prev, label_curr, dry_run=args.dry_run)
        except Exception as e:
            print(f'  [{line}] ERROR: {e}'); continue
        if 'skipped' in res:
            print(f'  [{line}] SKIP: {res["skipped"]}'); continue
        m, nm = res['matched'], res['not_matched']
        total = m + nm
        pct = round(m * 100 / total, 1) if total else 0
        print(f'  [{line}] OK: {m}/{total} entries actualizados ({pct}%), {nm} sin match')
        for s in res['samples_unmatched'][:3]:
            print(f'    sample no-match: {s}')

    if args.dry_run:
        print('\nDRY RUN: nada se escribio.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
