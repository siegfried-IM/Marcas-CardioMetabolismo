#!/usr/bin/env python3
"""
shared/merge-recetas-snc.py

Mergea recetas en SNC/index.html y SNC/psq_dashboard.html (que tienen
la data inline en `const D = {...};`, distinto a las otras lineas).

Input: pivot Excel con formato
  R1: [None, None, "Mes-Año", "<date>", "<date>", "<date>", ...]
  R2: ["Mercado (sin Mix)", "Droga", "Marca", "Cant. Recetas", "Cant. Medicos", "Cant. Recetas", ...]
  R3+: data rows

Acepta multiples meses por archivo. Solo actualiza:
  - D.recetas[brand][<MMM YYYY>] = {recetas, medicos}
  - D.rec_comp[<owner_sie>][brand].monthly[<MMM YYYY>] = recetas
  - D.rec_ms[<owner_sie>].sie/mkt/ms[<MMM YYYY>]

NO toca: budget, mol_perf, kpiStrip, stock, canales, convenios, precios.
NO toca otras lineas.

Uso:
  py shared/merge-recetas-snc.py --pivot "<pivot.xlsx>"
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

# Pivot market name -> SNC rec_ms owner SIE key(s).
# Cuando hay varios owners (caso MADOPAR + MADOPAR HBS), las recetas
# de cada brand se asignan al owner que matchea (por nombre); el market
# total se replica en ambos.
MARKET_TO_OWNERS = {
    '.PREGABALINA MULTIDOSIS':   ['PGB MULTIDOSIS SIE'],
    'ANTIDEPRESIVOS (VISDON)':   ['VISDON SIE'],
    'DIAZEPAM (VALIUM)':         ['VALIUM SIE'],
    'LEVETIRACETAM (LEVITAL)':   ['LEVITAL SIE'],
    'LURASIDONA (LURAP)':        ['LURAP SIE'],
    'MADOPAR (LEVO CARBI)':      ['MADOPAR SIE', 'MADOPAR HBS SIE'],
    'QUETIAPINAS (QTP)':         ['QTP SIE'],
    # Posibles markets futuros:
    'ANTIDEPRESIVOS (VALQUIR)':  ['VALQUIR SIE'],
    'BREXIPIPAZOL-ARIPIPRAZOL (BREXIL)': [],  # no SIE owner
}

SNC_HTMLS = ['SNC/index.html', 'SNC/psq_dashboard.html']


def normalize_brand(name):
    if not name: return ''
    return re.sub(r'\s+', ' ', str(name).upper()).strip()


def parse_pivot(pivot_path):
    """Devuelve dict: {market: {brand: {month_key_en: {'recetas': X, 'medicos': Y}}}}
       y la lista ordenada de meses encontrados."""
    wb = openpyxl.load_workbook(pivot_path, read_only=True, data_only=True)
    ws = wb.active

    # Row 1: column headers (dates), Row 2: subheaders ("Cant. Recetas" / "Cant. Medicos")
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    # Parse columns: each date appears in 1-2 consecutive cols (Recetas + Medicos).
    # Build col_map: { col_index: (month_key_en, 'recetas'|'medicos') }
    col_map = {}
    cur_month = None
    for i, h1 in enumerate(row1):
        if isinstance(h1, datetime):
            cur_month = f"{MES_EN[h1.month]} {h1.year}"
        h2 = (str(row2[i]) if i < len(row2) and row2[i] else '').lower()
        if 'receta' in h2:
            col_map[i] = (cur_month, 'recetas')
        elif 'médico' in h2 or 'medico' in h2:
            col_map[i] = (cur_month, 'medicos')

    months_set = set()
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'recetas': 0, 'medicos': 0})))

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]: continue
        market = str(row[0]).strip()
        droga = (row[1] or '')
        marca = row[2]
        # Skip "Totales" rows
        if str(droga).strip().lower() == 'totales' and not marca:
            continue
        if str(marca or '').strip().lower() == 'totales' or not marca:
            continue
        brand_norm = normalize_brand(marca)
        for col_idx, (month_key, kind) in col_map.items():
            if col_idx >= len(row): continue
            val = row[col_idx]
            try:
                v = int(val) if val is not None else 0
            except (TypeError, ValueError):
                v = 0
            data[market][brand_norm][month_key][kind] = v
            months_set.add(month_key)

    wb.close()
    return data, sorted(months_set, key=lambda k: (int(k.split()[1]), list(MES_EN.values()).index(k.split()[0])))


def find_inline_d(text):
    """Retorna (start_index_after_eq, parsed_obj, end_index_in_text)."""
    m = re.search(r'const\s+D\s*=\s*', text)
    if not m: return None
    obj_start = text.index('{', m.end())
    obj, end = json.JSONDecoder().raw_decode(text[obj_start:])
    return obj_start, obj, obj_start + end


def merge_into_d(D, pivot_data, months_in_pivot):
    """Patcha D in-place. Devuelve (n_brands_patched, n_owners_patched)."""
    rec_comp = D.setdefault('rec_comp', {})
    rec_ms = D.setdefault('rec_ms', {})
    recetas = D.setdefault('recetas', {})

    n_brands = 0
    n_owners = 0
    audit = []

    for market, brands in pivot_data.items():
        owners = MARKET_TO_OWNERS.get(market, [])
        if not owners:
            audit.append(f'  [skip] mercado sin owner SIE: {market}')
            continue

        for owner in owners:
            comp = rec_comp.setdefault(owner, {})
            rms = rec_ms.setdefault(owner, {'sie': {}, 'ms': {}, 'ms_quarterly': {}, 'mkt': {}})
            sie_dict = rms.setdefault('sie', {})
            mkt_dict = rms.setdefault('mkt', {})
            ms_dict = rms.setdefault('ms', {})

            for month in months_in_pivot:
                mkt_total = 0
                sie_count = 0
                for brand_norm, monthly in brands.items():
                    cell = monthly.get(month, {'recetas': 0, 'medicos': 0})
                    rec_val = cell.get('recetas', 0)
                    med_val = cell.get('medicos', 0)
                    mkt_total += rec_val
                    # Brand pertenece al owner si su nombre coincide con el owner
                    # o si el owner es el unico (single-SIE market).
                    is_owner_brand = (brand_norm == owner) or (
                        len(owners) == 1 and brand_norm.endswith(' SIE')
                    )
                    if is_owner_brand:
                        sie_count += rec_val
                    # Update rec_comp[owner][brand].monthly
                    pinfo = comp.setdefault(brand_norm, {'monthly': {}})
                    pinfo.setdefault('monthly', {})[month] = rec_val
                    # Update D.recetas[brand][month] = {recetas, medicos}
                    rentry = recetas.setdefault(brand_norm, {})
                    rentry[month] = {'recetas': rec_val, 'medicos': med_val}
                    n_brands += 1

                sie_dict[month] = sie_count
                mkt_dict[month] = mkt_total
                ms_dict[month] = round((sie_count / mkt_total) * 100, 2) if mkt_total else 0
            n_owners += 1
            audit.append(f'  [ok]   {market} -> {owner}: meses {months_in_pivot}')

    return n_brands, n_owners, audit


def serialize_back(text, obj_start, obj, obj_end_in_text, new_d):
    return text[:obj_start] + json.dumps(new_d, ensure_ascii=False) + text[obj_end_in_text:]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', required=True)
    ap.add_argument('--repo', default=str(Path(__file__).resolve().parent.parent))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pivot = Path(args.pivot)
    if not pivot.is_file():
        print(f'ERROR: pivot no existe: {pivot}', file=sys.stderr); return 2

    print(f'Leyendo pivot: {pivot}')
    data, months = parse_pivot(pivot)
    print(f'  Meses: {months}')
    print(f'  Mercados: {list(data.keys())}')

    repo = Path(args.repo)
    for rel in SNC_HTMLS:
        p = repo / rel
        if not p.is_file():
            print(f'\n[skip] no existe: {p}'); continue
        text = p.read_text(encoding='utf-8', errors='replace')
        pos = find_inline_d(text)
        if not pos:
            print(f'\n[skip] no encontre `const D = {{...}}` en {p}'); continue
        obj_start, D, obj_end = pos
        n_brands, n_owners, audit = merge_into_d(D, data, months)
        print(f'\n{rel}: patched {n_brands} brand-month entries, {n_owners} owners')
        for line in audit: print(line)
        if args.dry_run:
            continue
        new_text = serialize_back(text, obj_start, D, obj_end, D)
        p.write_text(new_text, encoding='utf-8', newline='')
        print(f'  -> escrito ({p.stat().st_size:,} bytes)')

    if args.dry_run:
        print('\nDRY RUN: nada se escribio.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
