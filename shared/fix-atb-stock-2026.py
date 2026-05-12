#!/usr/bin/env python3
"""Fix ATB stock Jan-Apr 2026: los valores en data.js estan 14-17x abajo del
pivot SAP correcto (probablemente cargado desde un pivot stale). Re-cargo
los 4 meses desde 'STOCK Y VENTAS.xlsx' del 2026-04.

Solo toca D.stock[fam][mk] para Jan/Feb/Mar/Apr 2026.
NO toca stock_alerts, stock_pres, coverage_labels — esos siguen siendo
arrays derivados, no se mueven."""
from __future__ import annotations
import re, json, sys
from pathlib import Path
import openpyxl
from datetime import datetime

REPO = Path(__file__).resolve().parent.parent
PIVOT = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\ATB\2026-04\fuentes-originales\STOCK Y VENTAS.xlsx')
DATA_JS = REPO / 'ATB' / 'data.js'
MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
TARGET = ['Jan 2026', 'Feb 2026', 'Mar 2026', 'Apr 2026']


def parse_fam_totals(path: Path):
    """Returns family_data[fam_upper][mk_en] = {stock, ventas, facturacion, dias}
    SOLO desde rows donde producto == 'Totales' (los totales de cada familia)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    col_map = {}
    for i, h1 in enumerate(row1):
        if not isinstance(h1, datetime): continue
        mk = f'{MES_EN[h1.month]} {h1.year}'
        h2 = (str(row2[i] or '')).strip().lower()
        metric = None
        if 'stock final' in h2: metric = 'stock'
        elif 'ventas' in h2: metric = 'ventas'
        elif 'facturaci' in h2: metric = 'facturacion'
        elif 'dias' in h2 or 'días' in h2: metric = 'dias'
        if metric: col_map[i] = (mk, metric)

    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row: continue
        lab = row[0] if len(row) > 0 else None
        fam = row[1] if len(row) > 1 else None
        prod = row[2] if len(row) > 2 else None
        if not lab or str(lab).strip() == 'Totales': continue
        if not fam or str(fam).strip() == 'Totales': continue
        if str(prod or '').strip() != 'Totales': continue
        fam_key = str(fam).strip().upper()
        out.setdefault(fam_key, {})
        for col_idx, (mk, metric) in col_map.items():
            if col_idx >= len(row): continue
            v = row[col_idx]
            try: vf = float(v) if v is not None else None
            except (TypeError, ValueError): vf = None
            if vf is None: continue
            out[fam_key].setdefault(mk, {})[metric] = int(round(vf))
    wb.close()
    return out


def main():
    if not PIVOT.is_file():
        print(f'ERROR: pivot no existe: {PIVOT}', file=sys.stderr); return 2
    fam_data = parse_fam_totals(PIVOT)
    print(f'Familias en pivot: {sorted(fam_data.keys())}')

    text = DATA_JS.read_text(encoding='utf-8-sig', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    ob = text.index('{', m.end())
    d2, end = json.JSONDecoder().raw_decode(text[ob:])
    prefix = text[:ob]; suffix = text[ob+end:]

    stock = d2.get('stock', {})
    fixed = 0
    for fam, months in stock.items():
        fam_up = fam.upper()
        src = fam_data.get(fam_up)
        if not src: continue
        for mk in TARGET:
            new = src.get(mk)
            if not new: continue
            cur = months.get(mk, {})
            cur_stock = cur.get('stock', 0)
            new_stock = new.get('stock', 0)
            # Only overwrite if values differ significantly (>2x diff)
            if cur_stock and new_stock and abs(cur_stock - new_stock) / max(cur_stock, new_stock) > 0.5:
                print(f'  {fam} {mk}: stock {cur_stock} -> {new_stock}, ventas {cur.get("ventas")} -> {new.get("ventas")}')
                months[mk] = {
                    'stock': new.get('stock', 0),
                    'ventas': new.get('ventas', 0),
                    'facturacion': new.get('facturacion', 0),
                    'dias': new.get('dias', 0),
                }
                fixed += 1
            elif not cur_stock and new_stock:
                # Was empty/missing
                print(f'  {fam} {mk}: FILL stock={new_stock} ventas={new.get("ventas")}')
                months[mk] = {
                    'stock': new.get('stock', 0),
                    'ventas': new.get('ventas', 0),
                    'facturacion': new.get('facturacion', 0),
                    'dias': new.get('dias', 0),
                }
                fixed += 1
    print(f'\nTotal entries arregladas: {fixed}')

    DATA_JS.write_text(prefix + json.dumps(d2, ensure_ascii=False) + suffix,
                       encoding='utf-8', newline='')
    print(f'-> {DATA_JS} ({DATA_JS.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
