#!/usr/bin/env python3
"""
shared/merge-ventas-internas.py

Actualiza budget[fam].YYYY.real (venta interna) en todas las lineas
desde un xlsx de Planilla de Ventas con formato:

  Familia | Ene-YYYY | Feb-YYYY | ... | Dic-YYYY+1

Acepta tanto el formato corto (1 fila de header con meses) como
el largo (2 filas de header). Detecta automaticamente.

Solo actualiza venta interna (real). NO toca presupuesto (budget).
NO toca rec_ms, rec_comp, recetas, mol_perf, stock, etc.

Familias se matchean por nombre exacto contra budget keys de cada
linea. Para mujer (que usa segmentos como 'SIN ESTROGENO' en el
inline D), aplica el mapeo brand->segment definido abajo.

Uso:
    py shared/merge-ventas-internas.py [--file <path>] [--cutoff YYYY-MM] [--dry-run]

  --cutoff: ultimo mes cerrado a incluir. Meses posteriores se
            ignoran (data parcial). Ej: --cutoff 2026-04 ignora May+.
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_FILE = Path(r'C:\Users\camarinaro\Downloads\Planilla de Ventas - 4 de mayo de 2026.xlsx')

# Map mes-año en Excel -> (year, month_idx 0..11)
MES_ES = {'ene':0,'feb':1,'mar':2,'abr':3,'may':4,'jun':5,
          'jul':6,'ago':7,'sep':8,'sept':8,'oct':9,'nov':10,'dic':11}

# Lineas: archivo a actualizar y donde esta budget
LINES = [
    {'key':'cardio',  'kind':'data.js', 'path':'cardio/data.js'},
    {'key':'antibio', 'kind':'data.js', 'path':'ATB/data.js'},
    {'key':'OTC',     'kind':'data.js', 'path':'OTC/data.js'},
    {'key':'respi',   'kind':'data.js', 'path':'respiratorio/data.js'},
    {'key':'SNC',     'kind':'inline',  'path':'SNC/index.html'},
    {'key':'mujer',   'kind':'inline',  'path':'mujer/index.html'},
    {'key':'dermato', 'kind':'inline',  'path':'dermatologia/dermato_dashboard.html'},
]

# Mapeo de segmento (mujer inline D) -> familia(s) del Excel
# El budget de mujer inline D suma todos los brands del segmento.
MUJER_SEGMENT_TO_FAMS = {
    'SIN ESTROGENO':   ['ISIS FREE'],
    'ALTA DOSIS':      ['ISIS'],
    'BAJA DOSIS 21+7': ['ISIS MINI'],
    'BAJA DOSIS 24':   ['ISIS MINI 24'],
    'COMPLEX':         ['SIDERBLUT COMPLEX', 'SIDERBLUT FOLIC'],
    'SOLO':            ['SIDERBLUT', 'SIDERBLUT POLI', 'FERINSOL'],
    'D3':              ['TRIP', 'CALCITOL D3', 'CALCITRIOL'],
    'D3 PLUS':         [],   # no hay match directo
    '45':              [],   # idem
    'MAGNESIO':        [],
    'DELTROX':         ['DELTROX'],
    'BASE':            ['CALCIO BASE DUPOMAR'],
    'BASE D':          ['CALCIO BASE DUPOMAR D', 'CALCIO BASE DUPOMAR D3',
                        'CALCIO CITRATO DUPOMAR D3'],
    'CLIMATIX':        ['CLIMATIX'],
}


def parse_xlsx(path, cutoff=None):
    """Devuelve dict[familia] = {(year, month_idx): value}.
    cutoff = (year, month_idx) o None. Meses posteriores se ignoran."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Detectar fila con headers de meses (a veces hay 2 filas, queremos
    # la que tenga "Ene-YYYY", "Feb-YYYY", etc.)
    header_row = None
    data_start = 2
    for ri in (1, 2):
        row = list(next(ws.iter_rows(min_row=ri, max_row=ri, values_only=True)))
        if not row: continue
        # ¿Tiene al menos un header tipo Mes-YYYY?
        if any(re.match(r'\w+[\s\-/]\d{4}', str(h).strip()) for h in row if h):
            header_row = row
            data_start = ri + 1
            break
    if header_row is None:
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        data_start = 2

    col_to_ym = {}  # col_idx -> (year, month_idx 0..11)
    for i, h in enumerate(header_row):
        if not h or i == 0: continue
        s = str(h).strip()
        m = re.match(r'(\w+)[\s\-/](\d{4})', s)
        if not m: continue
        mes = m.group(1).lower().rstrip('.')
        year = int(m.group(2))
        midx = MES_ES.get(mes)
        if midx is None: continue
        # Aplicar cutoff
        if cutoff is not None:
            cy, cm = cutoff
            if (year, midx) > (cy, cm):
                continue
        col_to_ym[i] = (year, midx)

    out = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or len(row) < 1 or not row[0]: continue
        fam = str(row[0]).strip()
        if not fam: continue
        # Skip header-like rows (e.g., 'Familia' as data row in 2-header layout)
        if fam.lower() in ('familia', 'familias', 'family'): continue
        out[fam] = {}
        for ci, ym in col_to_ym.items():
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try:
                out[fam][ym] = int(round(float(v)))
            except (ValueError, TypeError): pass
    wb.close()
    years_seen = sorted({y for fam_data in out.values() for (y, _) in fam_data})
    return out, years_seen


def load_data_js(p):
    text = p.read_text(encoding='utf-8-sig', errors='replace')
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text)
    if not m1: raise ValueError('OTC_DATA not found in ' + str(p))
    obj_start1 = text.index('{', m1.end())
    d1, end1 = json.JSONDecoder().raw_decode(text[obj_start1:])
    abs_end1 = obj_start1 + end1
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text[abs_end1:])
    if not m2: return text, d1, None, None
    obj_start2 = abs_end1 + text[abs_end1:].index('{', m2.end())
    d2, end2 = json.JSONDecoder().raw_decode(text[obj_start2:])
    abs_end2 = obj_start2 + end2
    return text, d1, d2, (obj_start1, abs_end1, obj_start2, abs_end2)


def write_data_js(text, d1, d2, bounds):
    obj_start1, abs_end1, obj_start2, abs_end2 = bounds
    return (text[:obj_start1]
            + json.dumps(d1, ensure_ascii=False)
            + text[abs_end1:obj_start2]
            + json.dumps(d2, ensure_ascii=False)
            + text[abs_end2:])


def load_inline(p):
    text = p.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    if not m: raise ValueError('const D not found in ' + str(p))
    obj_start = m.start() + len('const D = ')
    obj_start = text.index('{', obj_start)
    D, end = json.JSONDecoder().raw_decode(text[obj_start:])
    abs_end = obj_start + end
    return text, D, obj_start, abs_end


def write_inline(text, D, abs_start, abs_end):
    return text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]


def update_budget(budget, fam_to_excel, years_seen, line_key):
    """Devuelve count de familias actualizadas y lista de no matches."""
    updated, unmatched = 0, []
    for budget_key in list(budget.keys()):
        # Para mujer, si la key es un segmento, sumamos las familias mapeadas
        if line_key == 'mujer' and budget_key in MUJER_SEGMENT_TO_FAMS:
            target_fams = MUJER_SEGMENT_TO_FAMS[budget_key]
            if not target_fams:
                continue  # segmento sin mapeo, skip silently
            # Sumar valores de todas las target_fams
            sum_data = defaultdict(int)
            had_any = False
            for tf in target_fams:
                if tf in fam_to_excel:
                    had_any = True
                    for ym, v in fam_to_excel[tf].items():
                        sum_data[ym] += v
            if not had_any:
                unmatched.append(budget_key)
                continue
            new_data = dict(sum_data)
        else:
            # Match directo
            if budget_key not in fam_to_excel:
                unmatched.append(budget_key)
                continue
            new_data = fam_to_excel[budget_key]

        # Asegurar que existan los años en budget[key]
        for year in years_seen:
            yk = str(year)
            year_obj = budget[budget_key].setdefault(yk, {})
            real_arr = year_obj.get('real')
            if not isinstance(real_arr, list) or len(real_arr) != 12:
                real_arr = [None] * 12
            # Actualizar cada mes que tenga valor
            for (y, midx), v in new_data.items():
                if y == year:
                    real_arr[midx] = v
            year_obj['real'] = real_arr
            # Si no existe budget array, asegurar que exista (no lo tocamos)
            year_obj.setdefault('budget', [0]*12)
        updated += 1
    return updated, unmatched


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=str(DEFAULT_FILE))
    ap.add_argument('--cutoff', help="Ultimo mes cerrado, formato 'YYYY-MM'. Meses posteriores se ignoran.")
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    fp = Path(args.file)
    if not fp.is_file():
        print(f'ERROR: archivo no existe: {fp}', file=sys.stderr); return 2

    cutoff = None
    if args.cutoff:
        m = re.match(r'(\d{4})-(\d{2})', args.cutoff)
        if not m:
            print(f'ERROR: --cutoff debe ser YYYY-MM, recibido: {args.cutoff}', file=sys.stderr); return 2
        cutoff = (int(m.group(1)), int(m.group(2)) - 1)
        print(f'Cutoff: incluir hasta {args.cutoff} (mes idx {cutoff[1]})')

    print(f'Leyendo: {fp}')
    fam_data, years_seen = parse_xlsx(fp, cutoff=cutoff)
    print(f'  {len(fam_data)} familias en xlsx, años: {years_seen}')

    for line in LINES:
        path = REPO / line['path']
        if not path.is_file():
            print(f'  [{line["key"]}] SKIP: no existe {path}'); continue
        try:
            if line['kind'] == 'data.js':
                text, d1, d2, bounds = load_data_js(path)
                if d2 is None or 'budget' not in d2:
                    print(f'  [{line["key"]}] SKIP: no OTC_DASHBOARD.budget'); continue
                budget = d2['budget']
            else:
                text, D, abs_start, abs_end = load_inline(path)
                if 'budget' not in D:
                    print(f'  [{line["key"]}] SKIP: no D.budget'); continue
                budget = D['budget']

            updated, unmatched = update_budget(budget, fam_data, years_seen, line['key'])

            if args.dry_run:
                print(f'  [{line["key"]}] DRY: actualizaria {updated} familias, sin match: {len(unmatched)}'.encode('ascii','replace').decode())
                continue

            if line['kind'] == 'data.js':
                new_text = write_data_js(text, d1, d2, bounds)
            else:
                new_text = write_inline(text, D, abs_start, abs_end)
            path.write_text(new_text, encoding='utf-8', newline='')
            print(f'  [{line["key"]}] OK: {updated} familias actualizadas, {len(unmatched)} sin match'.encode('ascii','replace').decode())
            if unmatched and len(unmatched) <= 5:
                print(f'    sin match: {unmatched}')
        except Exception as e:
            print(f'  [{line["key"]}] ERROR: {e}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
