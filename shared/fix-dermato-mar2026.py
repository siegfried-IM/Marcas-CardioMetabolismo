#!/usr/bin/env python3
"""
shared/fix-dermato-mar2026.py

Dos cosas:
1) Sincroniza mol_perf de dermato con AR_PM master extendiendo los
   monthly_vals hasta Mar 2026 (estaba en Feb 2026). Recomputa
   quarterly_vals, ytd, mat, ms_*. Actualiza meta.

2) Remueve MICROSONA BB de:
   - sieProds (lista)
   - colors (mapa)
   - mol_perf[*].products (cualquier family)
   - rec_ms / rec_comp / recetas (si existe)
   - budget / stock / precios / convenios / canales / kpiStrip /
     brandKpis (si existe)

NO toca otras lineas. NO toca otras secciones que no sean las
listadas.

Uso: py shared/fix-dermato-mar2026.py [--dry-run]
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'dermatologia' / 'dermato_dashboard.html'
DEFAULT_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Apr-27-2026.xlsx')

REMOVE_BRAND = 'MICROSONA BB'

# Moleculas de dermato (matchean Molecules Long en AR_PM)
DERMATO_MOLECULES = [
    'MINOCYCLINE',
    'ADAPALENE_BENZOYL PEROXIDE',
    'CLOBETASOL',
    'CLOTRIMAZOLE',
    'CICLOPIROX',
    'BETAMETHASONE_CLOTRIMAZOLE',
    'HYDROCORTISONE',
    'MOMETASONE',
    'GENTAMICIN_HYDROCORTISONE_KETOCONAZOLE',
    'MUPIROCIN',
    'HYDROCORTISONE_MUPIROCIN',
    'ISOTRETINOIN',
]

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
MES_EN = {v:k for k,v in MES_INV.items()}


def msort(mk):
    p = mk.split()
    if len(p) != 2: return 0
    return int(p[1]) * 100 + MES_INV.get(p[0], 0)


def quarter_key(mk):
    parts = mk.split()
    if len(parts) != 2: return ''
    m = MES_INV.get(parts[0])
    if not m: return ''
    return f'Q{(m-1)//3+1} {parts[1]}'


def load_pm_for_molecules(path, target_mols):
    """Devuelve dict: norm_mol_name -> {product: {month_key: units}}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    col_manuf = col_product = col_pack = col_atc = col_mol = None
    month_cols = []
    for i, h in enumerate(row1):
        if not h: continue
        s = str(h).strip()
        s_norm = s.replace('\n', ' ').strip()
        if s_norm.lower().startswith('manufacturer'): col_manuf = i
        elif s_norm.lower().startswith('product'):    col_product = i
        elif s_norm.lower().startswith('pack'):       col_pack = i
        elif s_norm.lower().startswith('atc'):        col_atc = i
        elif s_norm.lower().startswith('molecules'):  col_mol = i
        if s.startswith('Units') and ('\n' in s or len(s.split()) >= 2):
            after = s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):].strip()
            after = after.strip()
            if after.upper().startswith('MAT') or after.upper().startswith('YTD'):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV:
                month_cols.append((i, f'{m.group(1)} {m.group(2)}'))

    if col_product is None: col_product = 1
    if col_mol is None: col_mol = 5

    # Normalizar molecules: la dashboard usa 'ADAPALENE_BENZOYL_PEROXIDE'
    # mientras AR_PM tiene 'ADAPALENE_BENZOYL PEROXIDE' (con espacio).
    def norm_mol(s):
        return str(s).strip().upper().replace(' ', '_')

    target_set = set(norm_mol(m) for m in target_mols)
    by_mol = defaultdict(lambda: defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float)}))

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        manuf = row[col_manuf] if col_manuf is not None and col_manuf < len(row) else None
        product = row[col_product] if col_product < len(row) else None
        mol = row[col_mol] if col_mol < len(row) else None
        if not product or not mol: continue
        mol_norm = norm_mol(mol)
        if mol_norm not in target_set: continue
        bucket = by_mol[mol_norm][product]
        bucket['manuf'] = manuf
        for ci, mk in month_cols:
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try: bucket['monthly'][mk] += float(v)
            except (ValueError, TypeError): pass
    wb.close()

    months_asc = [mk for _, mk in sorted(month_cols, key=lambda x: msort(x[1]))]
    out = {}
    for k, prods in by_mol.items():
        out[k] = {p: {'manuf': info['manuf'],
                      'monthly': {mk: int(round(v)) for mk, v in info['monthly'].items()}}
                  for p, info in prods.items()}
    return months_asc, out


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk: out[qk] += v
    return dict(out)


def aggregate_ytd_per_year(monthly, cierre_month):
    by_year = defaultdict(int)
    for mk, v in monthly.items():
        parts = mk.split()
        if len(parts) != 2: continue
        m_num = MES_INV.get(parts[0])
        if not m_num: continue
        if m_num <= cierre_month:
            by_year[parts[1]] += v
    return {f'{MES_EN[cierre_month]} {y}': v for y, v in by_year.items()}


def aggregate_mat_yearly(monthly, cierre_month):
    """MAT por año = rolling 12 meses ending <cierre_month> de YYYY."""
    years_with = set()
    for mk in monthly:
        p = mk.split()
        if len(p) == 2 and p[0] in MES_INV:
            years_with.add(int(p[1]))
    out = {}
    for y in sorted(years_with):
        total = 0
        for back in range(11, -1, -1):
            total_idx = (y * 12 + (cierre_month - 1)) - back
            yy, mm = divmod(total_idx, 12)
            mk = f'{MES_EN[mm + 1]} {yy}'
            total += int(monthly.get(mk, 0) or 0)
        out[f'{MES_EN[cierre_month]} {y}'] = total
    return out


def aggregate_mat_monthly(monthly):
    """MAT mensual rolling: para cada mes, suma de los 12 ending alli."""
    sorted_mks = sorted(monthly.keys(), key=msort)
    out = {}
    for i, mk in enumerate(sorted_mks):
        if i < 11: continue
        window = sorted_mks[i-11:i+1]
        out[mk] = sum(int(monthly.get(m, 0) or 0) for m in window)
    return out


def remove_microsona_bb(D):
    """Remueve MICROSONA BB de todas las secciones."""
    n = 0
    # sieProds
    if REMOVE_BRAND in D.get('sieProds', []):
        D['sieProds'].remove(REMOVE_BRAND); n += 1
    # colors
    if REMOVE_BRAND in D.get('colors', {}):
        del D['colors'][REMOVE_BRAND]; n += 1
    # secciones keyed por brand
    for sec_key in ('budget', 'rec_ms', 'rec_comp', 'recetas',
                    'stock', 'stock_alerts', 'stock_pres',
                    'precios', 'canales', 'convenios',
                    'kpiStrip', 'brandKpis', 'familyMap'):
        sec = D.get(sec_key)
        if isinstance(sec, dict) and REMOVE_BRAND in sec:
            del sec[REMOVE_BRAND]; n += 1
    # mol_perf: remover de products list
    for fam_key, fam_obj in D.get('mol_perf', {}).items():
        if not isinstance(fam_obj, dict): continue
        prods = fam_obj.get('products', [])
        before = len(prods)
        fam_obj['products'] = [p for p in prods
                                if not (REMOVE_BRAND in p.get('prod','').upper())]
        if len(fam_obj['products']) < before:
            n += (before - len(fam_obj['products']))
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--html', default=str(HTML))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pm = Path(args.master)
    if not pm.is_file():
        print(f'ERROR: master no existe: {pm}', file=sys.stderr); return 2

    print(f'Leyendo PM: {pm}')
    months, by_mol = load_pm_for_molecules(pm, DERMATO_MOLECULES)
    print(f'  Meses: {months[0]} .. {months[-1]} ({len(months)})')
    for mol, prods in by_mol.items():
        print(f'  [{mol}]: {len(prods)} productos')

    hp = Path(args.html)
    text = hp.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    if not m:
        print('ERROR: const D no encontrado'); return 3
    abs_start = m.start() + len('const D = ')
    abs_start = text.index('{', abs_start)
    D, end = json.JSONDecoder().raw_decode(text[abs_start:])
    abs_end = abs_start + end

    # === 1) Update mol_perf monthly_vals con data de PM ===
    cierre_month = MES_INV.get(months[-1].split()[0], 12)
    cierre_year  = int(months[-1].split()[1])
    cierre_mk    = months[-1]

    n_updated_prods = 0
    for fam_key, fam_obj in D.get('mol_perf', {}).items():
        if not isinstance(fam_obj, dict): continue
        # Buscar bucket del mol en by_mol (norm)
        fam_norm = str(fam_key).upper().replace(' ', '_')
        pm_prods = by_mol.get(fam_norm, {})
        if not pm_prods:
            print(f'  WARN: mol_perf[{fam_key}] sin match en PM (norm={fam_norm})'.encode('ascii','replace').decode())
            continue
        # Map por product name
        # pm_prods key = product name como aparece en AR_PM (ej. 'PALDAR (SIE)')
        # mol_perf product = mismo formato
        pm_by_name = pm_prods  # ya keyed por name
        # Update each product en mol_perf
        for p in fam_obj.get('products', []):
            name = p.get('prod', '')
            pm_data = pm_by_name.get(name)
            if not pm_data: continue
            # Reemplazar monthly_vals con la serie completa del PM
            new_monthly = pm_data['monthly']
            p['monthly_vals'] = new_monthly
            p['quarterly_vals'] = aggregate_quarterly(new_monthly)
            p['ytd'] = aggregate_ytd_per_year(new_monthly, cierre_month)
            p['mat'] = aggregate_mat_yearly(new_monthly, cierre_month)
            # Sumar tambien MAT monthly rolling para el chart de Mercado IQVIA
            p['mat'].update(aggregate_mat_monthly(new_monthly))
            n_updated_prods += 1

        # Recompute family-level
        fam_monthly = defaultdict(int)
        for p in fam_obj.get('products', []):
            for mk, v in p.get('monthly_vals', {}).items():
                fam_monthly[mk] += int(v or 0)
        fam_monthly = dict(fam_monthly)
        fam_obj['monthly'] = fam_monthly
        fam_obj['quarterly'] = aggregate_quarterly(fam_monthly)
        fam_obj['ytd'] = aggregate_ytd_per_year(fam_monthly, cierre_month)
        fam_obj['mat'] = aggregate_mat_yearly(fam_monthly, cierre_month)
        fam_obj['mat'].update(aggregate_mat_monthly(fam_monthly))

        # Recompute MS por producto vs family
        for p in fam_obj.get('products', []):
            mv = p.get('monthly_vals', {})
            qv = p.get('quarterly_vals', {})
            yv = p.get('ytd', {})
            mtv = p.get('mat', {})
            p['ms_monthly'] = {mk: round(mv.get(mk,0)/fv*100, 2) if fv > 0 else 0
                                for mk, fv in fam_monthly.items()}
            p['ms_quarterly'] = {qk: round(qv.get(qk,0)/fv*100, 2) if fv > 0 else 0
                                  for qk, fv in fam_obj['quarterly'].items()}
            p['ms_ytd'] = {y: round(yv.get(y,0)/fv*100, 2) if fv > 0 else 0
                            for y, fv in fam_obj['ytd'].items()}
            p['ms_mat'] = {mk: round(mtv.get(mk,0)/fv*100, 2) if fv > 0 else 0
                            for mk, fv in fam_obj['mat'].items()}

    print(f'\nProductos actualizados: {n_updated_prods}')

    # === 2) Update meta ===
    meta = D.setdefault('meta', {})
    cmk = f'{MES_EN[cierre_month]} {cierre_year}'
    pmk = f'{MES_EN[cierre_month]} {cierre_year-1}'
    meta['latest_month']    = cmk
    meta['current_ytd_key'] = cmk
    meta['prev_ytd_key']    = pmk
    meta['current_mat_key'] = cmk
    meta['prev_mat_key']    = pmk
    meta['kpi_ytd_label']      = f"YTD {MES_EN[cierre_month]}'{str(cierre_year)[-2:]}"
    meta['kpi_ytd_prev_label'] = f"YTD {MES_EN[cierre_month]}'{str(cierre_year-1)[-2:]}"
    meta['kpi_mat_label']      = f"MAT {MES_EN[cierre_month]}'{str(cierre_year)[-2:]}"
    meta['kpi_mat_prev_label'] = f"MAT {MES_EN[cierre_month]}'{str(cierre_year-1)[-2:]}"
    print(f'meta.latest_month: {meta["latest_month"]}')

    # === 3) Remover MICROSONA BB ===
    n_removed = remove_microsona_bb(D)
    print(f'MICROSONA BB removido en {n_removed} secciones')

    if args.dry_run:
        print('\nDRY RUN: no se escribio.')
        return 0

    new_text = text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    hp.write_text(new_text, encoding='utf-8', newline='')
    print(f'\n-> {hp} reescrito ({hp.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
