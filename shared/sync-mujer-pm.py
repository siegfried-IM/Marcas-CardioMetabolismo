#!/usr/bin/env python3
"""
shared/sync-mujer-pm.py

Sincroniza la inline D de mujer/index.html con la data IQVIA Premium del
AR_PM master. Respeta la SEGMENTACION existente (familias en mol_perf:
SIN ESTROGENO, ALTA DOSIS, BAJA DOSIS 21+7, ..., CLIMATIX) y solo
reemplaza el time-series de cada producto.

Pasos:
  1) Lee AR_PM_mujer.xlsx (sliced) o, fallback, AR_PM master.
  2) Por cada product en inline D mol_perf[fam].products[]:
       - Match por Product (col 1). Si no, por Pack (col 2). Si no, skip.
       - Reemplaza monthly_vals con todos los meses del PM.
       - Recomputa quarterly_vals, ytd (per year), mat (per year).
       - Recomputa ms_monthly, ms_quarterly, ms_ytd, ms_mat usando
         el total de la familia (suma de todos los products).
  3) Recomputa los aggregates al nivel mol_perf[fam]: ytd, mat, monthly,
     quarterly (suma de todos los products del fam).
  4) Escribe inline D al index.html.

Uso:
    py shared/sync-mujer-pm.py [--master <path>] [--dry-run]
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'mujer' / 'index.html'
DEFAULT_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Apr-27-2026.xlsx')

# ATC-4 codes en los que mujer compite (mismos que el slicer)
MUJER_ATCS: set[str] = {
    'V03X0', 'A12C1', 'A11C2', 'G02X9', 'G03A1', 'A12A0',
    'A11X9', 'M05B3', 'G01D0', 'B03A2', 'B03A1', 'G03X0', 'G03A5',
}

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}


def msort(mk):
    p = mk.split()
    if len(p) != 2: return 0
    return int(p[1]) * 100 + MES_INV.get(p[0], 0)


def quarter_key(mk):
    parts = mk.split()
    if len(parts) != 2: return ''
    m = MES_INV.get(parts[0])
    if not m: return ''
    q = (m - 1) // 3 + 1
    return f'Q{q} {parts[1]}'


def load_pm(path, atc_filter=None):
    """Lee el AR_PM master. Detecta columnas correctamente diferenciando
    Units monthly de Units YTD/MAT (en el master los headers son
    'Units\nMar 2026' vs 'Units\nMAT Mar 2026' vs 'Units\nYTD Mar 2026').

    Devuelve (month_keys_asc, by_product, by_pack).
    Si atc_filter es un set de ATC-4 codes, solo incluye rows que matchean."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    # Identificar columnas de Units monthly (NO MAT ni YTD)
    # Header tipo: 'Units\n<Mes> <YYYY>'
    month_cols = []
    # Tambien indices de campos clave
    col_manuf = col_product = col_pack = col_atc = col_mol = None
    for i, h in enumerate(row1):
        if not h: continue
        s = str(h).strip()
        s_norm = s.replace('\n', ' ').strip()
        # Campos categoricos
        if s_norm.lower().startswith('manufacturer'): col_manuf = i
        elif s_norm.lower().startswith('product'):    col_product = i
        elif s_norm.lower().startswith('pack'):       col_pack = i
        elif s_norm.lower().startswith('atc'):        col_atc = i
        elif s_norm.lower().startswith('molecules'):  col_mol = i
        # Units monthly: matcheo 'Units' + newline + 'Mes YYYY' (sin MAT/YTD)
        if s.startswith('Units') and ('\n' in s or len(s.split()) >= 2):
            after = s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):].strip()
            after = after.strip()
            if after.upper().startswith('MAT') or after.upper().startswith('YTD'):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV:
                month_cols.append((i, f'{m.group(1)} {m.group(2)}'))

    if col_manuf is None: col_manuf = 0
    if col_product is None: col_product = 1
    if col_pack is None:    col_pack = 2
    if col_atc is None:     col_atc = 3
    if col_mol is None:     col_mol = 5

    by_product = defaultdict(list)
    by_pack = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        manuf   = row[col_manuf]   if col_manuf < len(row)   else None
        product = row[col_product] if col_product < len(row) else None
        pack    = row[col_pack]    if col_pack < len(row)    else None
        atc     = row[col_atc]     if col_atc < len(row)     else None
        mol     = row[col_mol]     if col_mol < len(row)     else None
        if not product: continue
        # Filtrar por ATC si corresponde
        if atc_filter:
            atc_str = str(atc or '')
            atc4 = atc_str.split('-')[0].strip() if atc_str else ''
            if atc4 not in atc_filter:
                continue
        monthly = {}
        for ci, mk in month_cols:
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try:
                monthly[mk] = float(v)
            except (ValueError, TypeError):
                pass
        rec = {'manuf': manuf, 'product': product, 'pack': pack,
               'atc': atc, 'mol': mol, 'monthly': monthly}
        by_product[product].append(rec)
        if pack:
            by_pack[pack].append(rec)
    wb.close()
    months_asc = [mk for _, mk in sorted(month_cols, key=lambda x: msort(x[1]))]
    return months_asc, by_product, by_pack


def find_pm_match(prod_name, by_product, by_pack):
    """Devuelve list de rows del PM que matchean."""
    if prod_name in by_product:
        return by_product[prod_name]
    if prod_name in by_pack:
        return by_pack[prod_name]
    return []


def aggregate_monthly(rows, all_months):
    """Suma monthly_vals across rows. Devuelve dict {mk: int}."""
    out = {}
    for mk in all_months:
        total = 0.0
        present = False
        for r in rows:
            v = r['monthly'].get(mk)
            if v is not None:
                total += v
                present = True
        if present:
            out[mk] = int(round(total))
    return out


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk: out[qk] += v
    return dict(out)


def aggregate_ytd_per_year(monthly, cierre_month=None):
    """Para cada año del data, suma Jan..<cierre_month>.
    Si cierre_month es None, usa Dic (12). Devuelve dict keyed por
    '<mes_cierre> <YYYY>' (formato consistente con dashboards)."""
    if not monthly: return {}
    if cierre_month is None: cierre_month = 12
    inv = {v: k for k, v in MES_INV.items()}  # 1->'Jan', etc.
    mes_label = MES_INV.get(cierre_month) if cierre_month in MES_INV.values() else None
    # cierre_month puede ser num int. Mapear a label:
    if isinstance(cierre_month, int):
        # MES_INV: {'Jan':1,...} so reverse
        num_to_label = {v: k for k, v in MES_INV.items()}
        mes_label = num_to_label.get(cierre_month, 'Dec')
    by_year = defaultdict(int)
    for mk, v in monthly.items():
        parts = mk.split()
        if len(parts) != 2: continue
        m_num = MES_INV.get(parts[0])
        if not m_num: continue
        if m_num <= cierre_month:
            by_year[parts[1]] += v
    return {f'{mes_label} {y}': v for y, v in by_year.items()}


def aggregate_mat(monthly, cierre_month=None):
    """MAT por año = rolling 12 meses terminando en <cierre_month> de YYYY.
    Devuelve dict keyed por '<mes_cierre> <YYYY>'."""
    if not monthly: return {}
    if cierre_month is None: cierre_month = 12
    num_to_label = {v: k for k, v in MES_INV.items()}
    mes_label = num_to_label.get(cierre_month, 'Dec')

    # Identificar años con data
    years_with = set()
    for mk in monthly:
        parts = mk.split()
        if len(parts) == 2 and parts[0] in MES_INV:
            years_with.add(int(parts[1]))

    out = {}
    for y in sorted(years_with):
        # MAT termina en <cierre_month> y, empieza 11 meses antes
        total = 0
        for back in range(11, -1, -1):
            total_idx = (y * 12 + (cierre_month - 1)) - back
            yy, mm = divmod(total_idx, 12)
            mk = f'{num_to_label[mm + 1]} {yy}'
            total += int(monthly.get(mk, 0) or 0)
        out[f'{mes_label} {y}'] = total
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--html', default=str(HTML))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pm_path = Path(args.master)
    if not pm_path.is_file():
        print(f'ERROR: master no existe: {pm_path}', file=sys.stderr); return 2

    print(f'Leyendo: {pm_path} (filtrando ATCs de mujer)')
    all_months, by_product, by_pack = load_pm(pm_path, atc_filter=MUJER_ATCS)
    print(f'  {len(all_months)} meses ({all_months[0]} .. {all_months[-1]})')
    print(f'  {len(by_product)} productos unicos en AR_PM')
    last_mk = all_months[-1]
    # Cierre month como int (1-12) para aggregates ytd/mat
    cierre_m = MES_INV.get(last_mk.split()[0], 12)

    # Read inline D
    html_path = Path(args.html)
    text = html_path.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    if not m:
        print('ERROR: const D no encontrado'); return 3
    abs_start = m.start() + len('const D = ')
    abs_start = text.index('{', abs_start)
    D, end = json.JSONDecoder().raw_decode(text[abs_start:])
    abs_end = abs_start + end

    mol_perf = D.get('mol_perf', {})
    if not mol_perf:
        print('ERROR: D.mol_perf vacio'); return 4

    n_products_updated = 0
    n_unmatched = 0
    unmatched_list = []

    # Step 1: Update each product's time-series from PM
    for fam, fam_obj in mol_perf.items():
        if not isinstance(fam_obj, dict): continue
        for p in fam_obj.get('products', []):
            prod_name = p.get('prod', '')
            matches = find_pm_match(prod_name, by_product, by_pack)
            if not matches:
                n_unmatched += 1
                unmatched_list.append((fam, prod_name))
                continue
            new_monthly = aggregate_monthly(matches, all_months)
            if not new_monthly: continue
            p['monthly_vals'] = new_monthly
            p['quarterly_vals'] = aggregate_quarterly(new_monthly)
            p['ytd'] = aggregate_ytd_per_year(new_monthly, cierre_month=cierre_m)
            p['mat'] = aggregate_mat(new_monthly, cierre_month=cierre_m)
            n_products_updated += 1

    # Step 2: Recompute family-level aggregates (sum of all products' monthly_vals)
    for fam, fam_obj in mol_perf.items():
        if not isinstance(fam_obj, dict): continue
        prods = fam_obj.get('products', [])
        # Family monthly = sum of all products
        fam_monthly = defaultdict(int)
        for p in prods:
            for mk, v in p.get('monthly_vals', {}).items():
                fam_monthly[mk] += v
        fam_monthly = dict(fam_monthly)
        fam_obj['monthly'] = fam_monthly
        fam_obj['quarterly'] = aggregate_quarterly(fam_monthly)
        fam_obj['ytd'] = aggregate_ytd_per_year(fam_monthly, cierre_month=cierre_m)
        fam_obj['mat'] = aggregate_mat(fam_monthly, cierre_month=cierre_m)

        # Step 3: Recompute MS for each product (vs family total)
        for p in prods:
            mv = p.get('monthly_vals', {})
            ms_monthly = {}
            for mk, fv in fam_monthly.items():
                pv = mv.get(mk, 0)
                ms_monthly[mk] = round(pv / fv * 100, 2) if fv > 0 else 0
            p['ms_monthly'] = ms_monthly
            qv = p.get('quarterly_vals', {})
            fam_q = fam_obj['quarterly']
            ms_q = {}
            for qk, fv in fam_q.items():
                pv = qv.get(qk, 0)
                ms_q[qk] = round(pv / fv * 100, 2) if fv > 0 else 0
            p['ms_quarterly'] = ms_q
            ytd = p.get('ytd', {})
            fam_ytd = fam_obj['ytd']
            p['ms_ytd'] = {y: (round(ytd.get(y, 0) / fam_ytd[y] * 100, 2) if fam_ytd[y] > 0 else 0)
                           for y in fam_ytd}
            mat = p.get('mat', {})
            fam_mat = fam_obj['mat']
            p['ms_mat'] = {y: (round(mat.get(y, 0) / fam_mat[y] * 100, 2) if fam_mat[y] > 0 else 0)
                           for y in fam_mat}

    print(f'\nProductos actualizados: {n_products_updated}')
    print(f'Productos sin match: {n_unmatched}')
    if unmatched_list[:10]:
        for fam, p in unmatched_list[:10]:
            print(f'  unmatched: [{fam}] {p}'.encode('ascii','replace').decode())

    # Sanity: check Mar 2026 presence
    sample_fam = list(mol_perf.keys())[0]
    sample = mol_perf[sample_fam]['products'][0]
    print(f'\nSample post-sync [{sample_fam}][{sample["prod"]}]:'.encode('ascii','replace').decode())
    print(f'  monthly_vals last: {list(sample["monthly_vals"].items())[-3:]}')

    if args.dry_run:
        print('\nDRY RUN: no se escribio.')
        return 0

    new_text = text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    html_path.write_text(new_text, encoding='utf-8', newline='')
    print(f'\n-> {html_path} reescrito ({html_path.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
